import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import yfinance as yf
from fpdf import FPDF
import time
import numpy as np
import PyPDF2
import json
import base64
from openai import OpenAI
from geopy.geocoders import Nominatim
import os
import re
import requests
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path

# --- CONFIGURAZIONE PAGINA ---
APP_PAGE_TITLE = os.getenv("SMES_REPORTING_PAGE_TITLE", "Supporto alla Rendicontazione PMI")
APP_DISPLAY_TITLE = os.getenv("SMES_REPORTING_DISPLAY_TITLE", "🌍 Supporto alla Rendicontazione PMI")
OFFLINE_MODE = os.getenv("SMES_REPORTING_OFFLINE_MODE", "").strip().lower() in {"1", "true", "yes", "on"}

st.set_page_config(page_title=APP_PAGE_TITLE, layout="wide")

# --- CACHE PERSISTENTE E THROTTLING (Anti-Rate-Limiting) ---
APP_ROOT = Path(__file__).resolve().parent
APP_DATA_ROOT = Path(os.getenv("SMES_REPORTING_DATA_DIR", str(APP_ROOT))).expanduser()
CACHE_DIR = str(APP_DATA_ROOT / ".yfinance_cache")
CACHE_TIMEOUT_HOURS = 24
THROTTLE_DELAY = 2
PROJECTS_DIR = str(APP_DATA_ROOT / ".saved_projects")

os.makedirs(APP_DATA_ROOT, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(PROJECTS_DIR, exist_ok=True)

PROJECT_STATE_KEYS = [
    'revenue', 'opex', 'totale_attivo', 'dipendenti', 'company_name', 'quotata',
    'sector', 'industry', 'selected_country', 'scope1', 'scope2', 'scope3',
    'perc_red', 'em_final', 'rata_prestito', 'policy_multiplier', 'capex_totale',
    'tax_portfolio', 'cbam_portfolio', 'gap_answers', 'impresa_area',
    'sucursale_eu_200', 'hq_address', 'hq_geocoded_address', 'hq_lat', 'hq_lon',
    'status_normativo', 'vsme_module_choice', 'deliverable_type',
    'b1_info_omesse', 'b1_info_omesse_quali', 'b1_perimetro',
    'ragione_sociale', 'codice_nace_ateco', 'b1_paese_operazioni_asset_significativi'
]


def sanitize_project_name(project_name):
    sanitized = re.sub(r'[^A-Za-z0-9._ -]+', '', str(project_name)).strip()
    sanitized = re.sub(r'\s+', '_', sanitized)
    return sanitized[:80] or f"progetto_{int(time.time())}"


def get_project_file(project_name):
    return os.path.join(PROJECTS_DIR, f"{sanitize_project_name(project_name)}.json")


def format_unbounded_number(value):
    if value in (None, ""):
        return "0"
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        # Unparseable value (e.g. a stale Italian-formatted string stored as the
        # value key): return "0" rather than propagating the bad string, which
        # would cause an infinite display-reset loop in unbounded_number_input.
        return "0"
    if numeric_value.is_integer():
        return str(int(numeric_value))

    # No thousand separators; use comma as decimal separator (Italian format).
    decimal_part = f"{numeric_value:.10f}".split(".")[1].rstrip("0")
    if not decimal_part:
        return str(int(numeric_value))
    return f"{int(numeric_value)},{decimal_part}"


def parse_unbounded_number(raw_value):
    text = str(raw_value or "").strip()
    if not text:
        return 0

    cleaned = re.sub(r"[€\s_]", "", text)
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        integer_part, decimal_part = cleaned.rsplit(",", 1)
        if decimal_part.isdigit() and len(decimal_part) <= 2:
            cleaned = f"{integer_part}.{decimal_part}"
        else:
            cleaned = cleaned.replace(",", "")
    elif "." in cleaned:
        # Dots only: treat as Italian thousands separator when every group
        # after the first dot is exactly 3 digits (e.g. "15.000", "1.234.567").
        parts = cleaned.lstrip("-").split(".")
        if len(parts) > 1 and all(len(p) == 3 and p.isdigit() for p in parts[1:]) and parts[0].isdigit():
            cleaned = cleaned.replace(".", "")

    try:
        numeric_value = float(cleaned)
    except ValueError:
        return None

    if numeric_value.is_integer():
        return int(numeric_value)
    return numeric_value


def sync_formatted_number_input(display_key, value_key, allow_float):
    parsed_value = parse_unbounded_number(st.session_state.get(display_key, ""))
    if parsed_value is None:
        # Invalid input — reset the display to the last known valid value.
        current = st.session_state.get(value_key, 0)
        st.session_state[display_key] = "" if current == 0 else format_unbounded_number(current)
        return

    if not allow_float:
        parsed_value = int(parsed_value)

    st.session_state[value_key] = parsed_value
    st.session_state[display_key] = format_unbounded_number(parsed_value)


def unbounded_number_input(label, value_key, display_key, help=None, placeholder=None, allow_float=True):
    current_value = st.session_state.get(value_key, 0)
    # Coerce any string stored in the value key (e.g. from an old autosave that
    # serialised an Italian-formatted string instead of a number) to a proper
    # numeric value so the display can be formatted correctly.
    if isinstance(current_value, str):
        coerced = parse_unbounded_number(current_value)
        current_value = coerced if coerced is not None else 0
        st.session_state[value_key] = current_value
    formatted_value = format_unbounded_number(current_value)
    if display_key not in st.session_state:
        st.session_state[display_key] = "" if current_value == 0 else formatted_value
    else:
        parsed_display_value = parse_unbounded_number(st.session_state[display_key])
        # Reset when the display is unparseable (corrupted value) OR when it
        # represents the correct underlying number but is not in canonical format.
        if parsed_display_value is None or (
            parsed_display_value == current_value and st.session_state[display_key] != formatted_value
        ):
            st.session_state[display_key] = "" if current_value == 0 else formatted_value

    st.text_input(
        label,
        key=display_key,
        help=help,
        placeholder=placeholder,
        on_change=sync_formatted_number_input,
        args=(display_key, value_key, allow_float),
    )
    return st.session_state.get(value_key, 0)


def _to_json_safe(value):
    if value is None:
        return None
    if isinstance(value, dict):
        return {str(key): _to_json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_to_json_safe(item) for item in value]
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, np.generic):
        return value.item()
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return str(value)


def serialize_dataframe(df):
    if df is None:
        return {"columns": [], "records": []}
    safe_df = df.copy()
    for column in safe_df.columns:
        safe_df[column] = safe_df[column].map(_to_json_safe)
    return {
        "columns": list(safe_df.columns),
        "records": safe_df.to_dict(orient='records')
    }


def deserialize_dataframe(payload):
    if not payload:
        return pd.DataFrame()
    records = payload.get("records", [])
    columns = payload.get("columns", [])
    if not records and columns:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records, columns=columns or None)


def serialize_gap_documents(documents):
    serialized = {}
    for key, doc in (documents or {}).items():
        if not isinstance(doc, dict):
            continue
        binary_data = doc.get('data', b'') or b''
        if isinstance(binary_data, str):
            binary_data = binary_data.encode('utf-8')
        serialized[key] = {
            'filename': doc.get('filename', ''),
            'timestamp': doc.get('timestamp', ''),
            'data_b64': base64.b64encode(binary_data).decode('ascii') if binary_data else '',
        }
    return serialized


def deserialize_gap_documents(payload):
    documents = {}
    for key, doc in (payload or {}).items():
        if not isinstance(doc, dict):
            continue
        encoded_data = doc.get('data_b64', '')
        documents[key] = {
            'filename': doc.get('filename', ''),
            'timestamp': doc.get('timestamp', ''),
            'data': base64.b64decode(encoded_data) if encoded_data else b'',
        }
    return documents


def serialize_vsme_disclosure_tables(tables):
    return {
        key: serialize_dataframe(df)
        for key, df in (tables or {}).items()
        if isinstance(df, pd.DataFrame)
    }


def deserialize_vsme_disclosure_tables(payload):
    return {
        key: deserialize_dataframe(df_payload)
        for key, df_payload in (payload or {}).items()
    }


def build_project_payload(project_name):
    state = {key: _to_json_safe(st.session_state.get(key)) for key in PROJECT_STATE_KEYS}
    state['portfolio_df'] = serialize_dataframe(st.session_state.get('portfolio_df', pd.DataFrame()))
    state['gap_documents'] = serialize_gap_documents(st.session_state.get('gap_documents', {}))
    state['vsme_disclosure_tables'] = serialize_vsme_disclosure_tables(st.session_state.get('vsme_disclosure_tables', {}))
    return {
        'project_name': project_name,
        'saved_at': datetime.now().isoformat(),
        'state': state,
    }


def normalize_project_payload(payload, fallback_name=None):
    state = payload.get('state', {}) if isinstance(payload, dict) else {}
    project_name = payload.get('project_name') if isinstance(payload, dict) else None
    normalized_name = project_name or fallback_name or f"progetto_{int(time.time())}"
    return {
        'project_name': normalized_name,
        'saved_at': (payload.get('saved_at') if isinstance(payload, dict) else None) or datetime.now().isoformat(),
        'state': state,
    }


def project_signature_from_payload(payload):
    return json.dumps(payload.get('state', {}), ensure_ascii=False, sort_keys=True)


def register_saved_project(payload):
    st.session_state['current_project_name'] = payload.get('project_name', '')
    st.session_state['last_project_saved_at'] = payload.get('saved_at', '')
    st.session_state['last_project_signature'] = project_signature_from_payload(payload)


def save_project_payload(payload):
    normalized_payload = normalize_project_payload(payload)
    with open(get_project_file(normalized_payload['project_name']), 'w', encoding='utf-8') as handle:
        json.dump(normalized_payload, handle, ensure_ascii=False, indent=2)
    register_saved_project(normalized_payload)
    return normalized_payload


def save_project(project_name):
    payload = build_project_payload(project_name)
    return save_project_payload(payload)


def export_project_snapshot(project_name=None):
    resolved_name = (
        (project_name or "").strip()
        or st.session_state.get('current_project_name', '').strip()
        or st.session_state.get('company_name', '').strip()
        or "progetto"
    )
    payload = normalize_project_payload(build_project_payload(resolved_name), fallback_name=resolved_name)
    file_name = f"{sanitize_project_name(payload['project_name'])}.json"
    file_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode('utf-8')
    return file_name, file_bytes


def list_saved_projects():
    projects = []
    for filename in sorted(os.listdir(PROJECTS_DIR)):
        if not filename.endswith('.json'):
            continue
        file_path = os.path.join(PROJECTS_DIR, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as handle:
                payload = json.load(handle)
            projects.append({
                'label': payload.get('project_name') or filename[:-5],
                'file_path': file_path,
                'saved_at': payload.get('saved_at', ''),
            })
        except Exception:
            projects.append({
                'label': filename[:-5],
                'file_path': file_path,
                'saved_at': '',
            })
    return projects


def load_project_payload(payload, fallback_name=None):
    normalized_payload = normalize_project_payload(payload, fallback_name=fallback_name)
    state = normalized_payload.get('state', {})
    for key in PROJECT_STATE_KEYS:
        if key in state:
            st.session_state[key] = state[key]

    loaded_portfolio_df = deserialize_dataframe(state.get('portfolio_df'))
    if not loaded_portfolio_df.empty:
        st.session_state['portfolio_df'] = process_portfolio_dataframe(
            loaded_portfolio_df.drop(columns=['Display_Size', 'Risk_Score'], errors='ignore')
        )
    else:
        st.session_state['portfolio_df'] = pd.DataFrame()
    st.session_state['gap_documents'] = deserialize_gap_documents(state.get('gap_documents'))
    st.session_state['vsme_disclosure_tables'] = deserialize_vsme_disclosure_tables(state.get('vsme_disclosure_tables'))
    register_saved_project(normalized_payload)

    # Clear cached display values for number inputs so they re-initialize from
    # the freshly loaded value keys on the next render.
    for key in ('dipendenti_input', 'revenue_input', 'totale_attivo_input'):
        st.session_state.pop(key, None)

    for key in list(st.session_state.keys()):
        if key.startswith('editor_vsme_table_') or key.startswith('show_pdf_'):
            del st.session_state[key]

    for answer_key, answer_data in st.session_state.get('gap_answers', {}).items():
        if isinstance(answer_data, dict) and 'ans' in answer_data:
            st.session_state[answer_key] = answer_data['ans']


def load_project(project_option):
    with open(project_option['file_path'], 'r', encoding='utf-8') as handle:
        payload = json.load(handle)
    load_project_payload(payload, fallback_name=project_option['label'])


def import_project(uploaded_file):
    raw_payload = json.loads(uploaded_file.getvalue().decode('utf-8'))
    fallback_name = os.path.splitext(uploaded_file.name)[0]
    normalized_payload = normalize_project_payload(raw_payload, fallback_name=fallback_name)
    saved_payload = save_project_payload(normalized_payload)
    load_project_payload(saved_payload, fallback_name=saved_payload['project_name'])
    return saved_payload


def autosave_project_if_needed(project_name):
    if not st.session_state.get('project_autosave_enabled', True):
        return False
    normalized_name = sanitize_project_name(project_name)
    if not normalized_name:
        return False

    payload = build_project_payload(project_name)
    current_signature = project_signature_from_payload(payload)
    if current_signature == st.session_state.get('last_project_signature'):
        return False

    save_project_payload(payload)
    return True


def delete_project(project_option):
    if os.path.exists(project_option['file_path']):
        os.remove(project_option['file_path'])

def get_cache_file(ticker):
    """Restituisce il percorso del file cache per un ticker"""
    return os.path.join(CACHE_DIR, f"{ticker.upper()}_cache.json")

def is_cache_valid(ticker):
    """Controlla se la cache è ancora valida"""
    cache_file = get_cache_file(ticker)
    if not os.path.exists(cache_file):
        return False
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
            timestamp = datetime.fromisoformat(cache_data.get('timestamp', '2020-01-01'))
            age_hours = (datetime.now() - datetime.fromisoformat(cache_data.get('timestamp'))).total_seconds() / 3600
            return age_hours < CACHE_TIMEOUT_HOURS
    except:
        return False

def load_from_cache(ticker):
    """Carica dati dalla cache locale"""
    cache_file = get_cache_file(ticker)
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)['data']
    except:
        return None

def save_to_cache(ticker, data):
    """Salva dati nella cache locale"""
    cache_file = get_cache_file(ticker)
    cache_data = {
        'ticker': ticker.upper(),
        'timestamp': datetime.now().isoformat(),
        'data': data
    }
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
    except:
        pass

def retry_with_backoff(func, ticker, max_retries=3):
    """Riprova con exponential backoff"""
    if OFFLINE_MODE:
        return None
    for attempt in range(max_retries):
        try:
            time.sleep(THROTTLE_DELAY)
            result = func(ticker)
            if result:
                return result
        except requests.exceptions.ConnectionError:
            wait_time = (2 ** attempt)
            if attempt < max_retries - 1:
                st.info(f"⏳ Tentativo {attempt + 1}/{max_retries}... (attesa {wait_time}s)")
                time.sleep(wait_time)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
    return None

# --- FUNZIONI YFINANCE CON CACHE ---
@lru_cache(maxsize=128)
def _fetch_from_yfinance(ticker):
    """Fetch da yfinance (interno, con caching LRU)"""
    try:
        company = yf.Ticker(ticker)
        info = company.info
        
        data = {
            'ticker': ticker.upper(),
            'company_name': info.get('longName', 'N/A'),
            'country': info.get('country', 'N/A'),
            'sector': info.get('sector', 'N/A'),
            'industry': info.get('industry', 'N/A'),
            'website': info.get('website', 'N/A'),
            'total_assets': info.get('totalAssets', 0),
            'revenue': info.get('totalRevenue', 0),
            'operating_expense': info.get('operatingExpense', 0),
            'capex': info.get('capitalExpenditure', 0),
            'employees': info.get('fullTimeEmployees', 0),
            'margin': info.get('operatingMargins', 0),
            'currency': info.get('currency', 'EUR'),
            'source': 'yfinance'
        }
        return data
    except Exception as e:
        return None

def get_company_info(ticker):
    """Recupera i dati dell'azienda con cache e retry"""
    ticker = ticker.upper()

    if OFFLINE_MODE:
        return load_from_cache(ticker)
    
    # 1. Prova cache locale
    if is_cache_valid(ticker):
        cached_data = load_from_cache(ticker)
        if cached_data:
            return cached_data
    
    # 2. Prova a scaricare da yfinance con retry
    result = retry_with_backoff(_fetch_from_yfinance, ticker)
    
    if result:
        save_to_cache(ticker, result)
        return result
    
    # 3. Fallback a cache anche se scaduta
    cached_data = load_from_cache(ticker)
    if cached_data:
        return cached_data
    
    return None

# --- CONFIGURAZIONE PAGINA (originale) ---

# --- COSTANTI VSME ---
VSME_SCALE_OPTIONS = ["Sì", "Sì, ma necessita integrazione", "No, ma pianificato", "No"]
VSME_PLOT_RESPONSE_ORDER = ["Sì", "Sì, ma necessita integrazione", "No, ma pianificato", "No"]


def normalize_vsme_response_label(answer):
    normalized = str(answer or "").strip().lower()
    if normalized in {"yes", "sì", "si"}:
        return "Sì"
    if "integrazione" in normalized or "integration" in normalized:
        return "Sì, ma necessita integrazione"
    if normalized.startswith("no") and ("pianific" in normalized or "planned" in normalized):
        return "No, ma pianificato"
    if normalized == "no":
        return "No"
    return str(answer or "").strip()


def to_vsme_plot_response_label(answer):
    return normalize_vsme_response_label(answer)
VSME_DEFAULT_FILE = "Gap Analysis Template_VSME_Standard_Tool v3.xlsx"
VSME_DATA_COLLECTION_FILE = "Raccolta dati VSME_Template app.xlsx"
VSME_DISCLOSURE_PILLARS = {
    "GEN": ["B1", "B2", "C1", "C2"],
    "E": ["B3", "B4", "B5", "B6", "B7", "C3", "C4"],
    "S": ["B8", "B9", "B10", "C5", "C6", "C7"],
    "G": ["B11", "C8", "C9"],
}


def _vsme_code_sort_key(code):
    match = re.match(r'([BC])(\d+)', str(code).strip())
    if not match:
        return (99, str(code))
    module_rank = 0 if match.group(1) == "B" else 1
    return (module_rank, int(match.group(2)))


def _extract_excel_preview(ws, start_row=1, end_row=None, max_cols=12, max_rows=None):
    if end_row is None:
        end_row = ws.max_row

    row_values = []
    for row_idx in range(start_row, end_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_cols + 1)]
        if any(value not in (None, "") for value in values):
            row_values.append(values)

    if not row_values:
        return pd.DataFrame()

    max_non_empty_col = 0
    for values in row_values:
        for idx, value in enumerate(values, start=1):
            if value not in (None, ""):
                max_non_empty_col = max(max_non_empty_col, idx)

    trimmed_rows = []
    visible_rows = row_values if max_rows is None else row_values[:max_rows]
    for values in visible_rows:
        trimmed_rows.append([
            "" if value is None else str(value).strip()
            for value in values[:max_non_empty_col]
        ])

    column_labels = [chr(64 + idx) for idx in range(1, max_non_empty_col + 1)]
    return pd.DataFrame(trimmed_rows, columns=column_labels)


def _extract_disclosure_preview(ws, code, max_cols=12, max_rows=None):
    code_starts = {}
    for row_idx in range(1, min(ws.max_row, 200) + 1):
        for col_idx in range(1, 4):
            value = ws.cell(row_idx, col_idx).value
            if not isinstance(value, str):
                continue
            match = re.match(r'^([BC]\d+)(?:\b|[\s\-.])', value.strip())
            if match and match.group(1) not in code_starts:
                code_starts[match.group(1)] = row_idx

    if code in code_starts:
        start_row = code_starts[code]
        following_starts = [row for item_code, row in code_starts.items() if item_code != code and row > start_row]
        end_row = min(following_starts) - 1 if following_starts else ws.max_row
        return _extract_excel_preview(ws, start_row=start_row, end_row=end_row, max_cols=max_cols, max_rows=max_rows)

    return _extract_excel_preview(ws, max_cols=max_cols, max_rows=max_rows)


@st.cache_data
def load_vsme_disclosure_reference(gap_file_path=VSME_DEFAULT_FILE, data_collection_path=VSME_DATA_COLLECTION_FILE):
    try:
        import openpyxl

        gap_wb = openpyxl.load_workbook(gap_file_path, data_only=True)
        data_wb = openpyxl.load_workbook(data_collection_path, data_only=True)

        disclosure_catalog = {pillar: [] for pillar in VSME_DISCLOSURE_PILLARS}
        section_names = {
            "GEN": {"general information", "general information "},
            "E": {"envi metrics", "environmental"},
            "S": {"social metrics", "social"},
            "G": {"governance metrics", "governance"},
        }

        sheet2 = gap_wb["Sheet2"]
        for pillar, codes in VSME_DISCLOSURE_PILLARS.items():
            code_details = {}
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                _, _, section, disclosure_code, disclosure_title, item_type = row[:6]
                if not section or not disclosure_code or str(disclosure_code).strip() not in codes:
                    continue

                normalized_section = re.sub(r'\s+', ' ', str(section).strip().lower())
                if normalized_section not in section_names[pillar]:
                    continue

                code = str(disclosure_code).strip()
                item = code_details.setdefault(code, {
                    "code": code,
                    "title": str(disclosure_title).strip(),
                    "datapoints": 0,
                })
                if str(item_type).strip().lower() == "datapoint":
                    item["datapoints"] += 1

            for code in codes:
                if code not in code_details:
                    continue

                sheet_names = []
                for sheet_name in data_wb.sheetnames:
                    if re.search(rf'^{re.escape(code)}(?:$|[-_])', sheet_name, flags=re.IGNORECASE):
                        sheet_names.append(sheet_name)

                tables = []
                for sheet_name in sheet_names:
                    preview = _extract_disclosure_preview(data_wb[sheet_name], code)
                    if preview.empty:
                        continue
                    tables.append({
                        "sheet_name": sheet_name,
                        "data": preview,
                    })

                disclosure_catalog[pillar].append({
                    "code": code,
                    "title": code_details[code]["title"],
                    "datapoints": code_details[code]["datapoints"],
                    "sheet_names": sheet_names,
                    "tables": tables,
                })

            disclosure_catalog[pillar] = sorted(
                disclosure_catalog[pillar],
                key=lambda item: _vsme_code_sort_key(item["code"]),
            )

        return disclosure_catalog, None
    except Exception as exc:
        return {pillar: [] for pillar in VSME_DISCLOSURE_PILLARS}, str(exc)

def load_vsme_checklist_from_excel(file_path=VSME_DEFAULT_FILE):
    """Carica le domande del checklist VSME dal file Excel."""
    try:
        import openpyxl
        # Carica il workbook con data_only=True per ottenere valori calcolati, non formule
        wb = openpyxl.load_workbook(file_path, data_only=True)
        excluded_questions = set()

        def parse_datapoints(value):
            if value is None or value == "":
                return 0
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                cleaned = value.strip().replace(",", ".")
                try:
                    return int(float(cleaned))
                except ValueError:
                    return 0
            return 0

        vsme_scale_options = VSME_SCALE_OPTIONS
        if "Support" in wb.sheetnames:
            ws_support = wb["Support"]
            excel_scale = []
            for row_idx in range(1, 10):
                value = ws_support.cell(row_idx, 1).value
                if value and isinstance(value, str) and value.strip():
                    excel_scale.append(value.strip())
                elif excel_scale:
                    break
            if len(excel_scale) >= 2:
                normalized_scale = []
                for option in excel_scale:
                    normalized_option = normalize_vsme_response_label(option)
                    if normalized_option and normalized_option not in normalized_scale:
                        normalized_scale.append(normalized_option)
                if all(opt in normalized_scale for opt in VSME_SCALE_OPTIONS):
                    vsme_scale_options = VSME_SCALE_OPTIONS
                elif len(normalized_scale) >= 2:
                    vsme_scale_options = normalized_scale
        
        sheet_map = {
            'GEN': 'Checklist_General_Information',
            'E': 'Checklist_Environment',
            'S': 'Checklist_Social',
            'G': 'Checklist_Governance',
        }
        
        checklist_data = {'base': {}, 'comprehensive': {}}
        
        for pillar, sheet_name in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            checklist_data['base'][pillar] = []
            checklist_data['comprehensive'][pillar] = []
            
            # Leggi da riga 4 (indice 3) in poi
            for row_idx in range(4, ws.max_row + 1):
                cell_b = ws.cell(row_idx, 2).value
                cell_c = ws.cell(row_idx, 3).value
                cell_h = ws.cell(row_idx, 8).value
                cell_j = ws.cell(row_idx, 10).value
                
                # Column B = documento base, Column D = datapoints base
                if cell_b and isinstance(cell_b, str) and cell_b.strip():
                    datapoints_base = parse_datapoints(ws.cell(row_idx, 4).value)
                    if datapoints_base == 0:
                        # Fallback su colonna C per compatibilita con eventuali template legacy.
                        datapoints_base = parse_datapoints(cell_c)
                    label_base = f"{cell_b.strip()} ({datapoints_base} datapoints)"
                    if cell_b.strip().lower() not in excluded_questions:
                        checklist_data['base'][pillar].append(label_base)
                
                # Column H = documento comprehensive, Column J = datapoints comprehensive
                if cell_h and isinstance(cell_h, str) and cell_h.strip():
                    datapoints_comp = parse_datapoints(cell_j)
                    label_comp = f"{cell_h.strip()} ({datapoints_comp} datapoints)"
                    if cell_h.strip().lower() not in excluded_questions:
                        checklist_data['comprehensive'][pillar].append(label_comp)
        
        return checklist_data, vsme_scale_options, None
    except Exception as e:
        return None, VSME_SCALE_OPTIONS, str(e)

# --- SINCRONIZZAZIONE (Session State) ---
st.session_state.setdefault('revenue', 0)
st.session_state.setdefault('opex', 0)
st.session_state.setdefault('totale_attivo', 0)
st.session_state.setdefault('dipendenti', 0)
st.session_state.setdefault('company_name', '')
st.session_state.setdefault('ragione_sociale', '')
st.session_state.setdefault('codice_nace_ateco', '')
st.session_state.setdefault('quotata', False)
st.session_state.setdefault('sector', 'Industrials')
st.session_state.setdefault('industry', 'Machinery')
st.session_state.setdefault('selected_country', 'Italia')
st.session_state.setdefault('scope1', 0)
st.session_state.setdefault('scope2', 0)
st.session_state.setdefault('scope3', 0)
st.session_state.setdefault('perc_red', 0)
st.session_state.setdefault('em_final', 0)
st.session_state.setdefault('rata_prestito', 0)
st.session_state.setdefault('policy_multiplier', 1.0)
st.session_state.setdefault('capex_totale', 0)
st.session_state.setdefault('tax_portfolio', [])
st.session_state.setdefault('cbam_portfolio', [])
st.session_state.setdefault('gap_answers', {})
st.session_state.setdefault('gap_documents', {})
st.session_state.setdefault('vsme_disclosure_tables', {})
st.session_state.setdefault('portfolio_df', pd.DataFrame())
st.session_state.setdefault('impresa_area', 'UE')
st.session_state.setdefault('sucursale_eu_200', 'No')
st.session_state.setdefault('hq_address', '')
st.session_state.setdefault('hq_geocoded_address', '')
st.session_state.setdefault('hq_lat', None)
st.session_state.setdefault('hq_lon', None)
st.session_state.setdefault('current_project_name', '')
st.session_state.setdefault('last_project_saved_at', '')
st.session_state.setdefault('last_project_signature', '')
st.session_state.setdefault('project_autosave_enabled', True)
st.session_state.setdefault('b1_info_omesse', 'No')
st.session_state.setdefault('b1_info_omesse_quali', '')
st.session_state.setdefault('b1_perimetro', 'Individuale')
st.session_state.setdefault('b1_paese_operazioni_asset_significativi', '')

if 'materiality_scores' not in st.session_state:
    st.session_state.materiality_scores = {
        "E1 - Cambiamento Climatico": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E2 - Inquinamento": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E3 - Risorse Idriche e Marine": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E4 - Biodiversità ed Ecosistemi": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E5 - Uso Risorse ed Economia Circolare": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "S1 - Forza Lavoro Propria": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S2 - Lavoratori nella Catena del Valore": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S3 - Comunità Interessate": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S4 - Consumatori ed Utenti Finali": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "G1 - Condotta Aziendale": {"pilastro": "G", "impatto": 1, "finanza": 1},
    }

if 'gics_sectors' not in st.session_state:
    st.session_state.gics_sectors = {
        "Industrials": ["Aerospace & Defense", "Building Products", "Construction & Engineering", "Electrical Equipment", "Industrial Conglomerates", "Machinery", "Commercial Services", "Transportation"],
        "Energy": ["Oil, Gas & Consumable Fuels", "Energy Equipment & Services"],
        "Materials": ["Chemicals", "Construction Materials", "Containers & Packaging", "Metals & Mining", "Paper & Forest Products"],
        "Consumer Discretionary": ["Automobiles & Components", "Consumer Durables & Apparel", "Consumer Services", "Retailing"],
        "Consumer Staples": ["Food & Staples Retailing", "Food, Beverage & Tobacco", "Household & Personal Products"],
        "Health Care": ["Health Care Equipment & Services", "Pharmaceuticals & Biotechnology"],
        "Financials": ["Banks", "Financial Services", "Insurance"],
        "Information Technology": ["Software & Services", "Technology Hardware & Equipment", "Semiconductors"],
        "Communication Services": ["Telecommunication Services", "Media & Entertainment"],
        "Utilities": ["Electric Utilities", "Gas Utilities", "Multi-Utilities", "Water Utilities", "Renewable Electricity"],
        "Real Estate": ["Equity REITs", "Real Estate Management & Development"],
        "Altro / Non Specificato": ["Altro"]
    }

def get_tot_emissions(): return st.session_state.scope1 + st.session_state.scope2 + st.session_state.scope3
def sync_from_perc(): st.session_state.em_final = int(get_tot_emissions() * (1 - st.session_state.perc_red / 100.0))
def sync_from_scopes(): sync_from_perc()

# --- SCALE DI VALUTAZIONE ---
ESRS_SCALE_OPTIONS = [
    "Per niente (0% - Nessuna azione intrapresa)",
    "In fase iniziale (Attività pianificata o discussa)",
    "Parzialmente (Policy esistente, implementazione frammentaria)",
    "In gran parte (Processo attivo, non ancora auditabile)",
    "Quasi completamente (Manca solo verifica finale o XBRL)",
    "Completamente (100% - Allineato a ESRS e verificabile)"
]
ESRS_SCALE_VALUES = {
    "Per niente (0% - Nessuna azione intrapresa)": 0,
    "In fase iniziale (Attività pianificata o discussa)": 1,
    "Parzialmente (Policy esistente, implementazione frammentaria)": 2,
    "In gran parte (Processo attivo, non ancora auditabile)": 3,
    "Quasi completamente (Manca solo verifica finale o XBRL)": 4,
    "Completamente (100% - Allineato a ESRS e verificabile)": 5
}

# --- FUNZIONI DATI E MAPPA ---
def process_portfolio_dataframe(df):
    cols = [c.lower() for c in df.columns]
    
    if 'address' in cols: df.rename(columns={df.columns[cols.index('address')]: 'Address'}, inplace=True)
    if 'operator' in cols: df.rename(columns={df.columns[cols.index('operator')]: 'Operator'}, inplace=True)
    if 'installation name' in cols: df.rename(columns={df.columns[cols.index('installation name')]: 'Name'}, inplace=True)
    elif 'nome' in cols: df.rename(columns={df.columns[cols.index('nome')]: 'Name'}, inplace=True)
    if 'installation id' in cols: df.rename(columns={df.columns[cols.index('installation id')]: 'ID'}, inplace=True)
    
    lat_col = next((c for c in df.columns if c.lower() in ['lat', 'latitude', 'latitudine']), None)
    lon_col = next((c for c in df.columns if c.lower() in ['lon', 'lng', 'longitude', 'longitudine']), None)
    
    if lat_col and lon_col:
        df.rename(columns={lat_col: 'Lat', lon_col: 'Lon'}, inplace=True)
        df = df.dropna(subset=['Lat', 'Lon'])
    else:
        st.error("Il file deve contenere colonne 'Lat' e 'Lon'.")
        return pd.DataFrame()

    if 'Name' not in df.columns: df['Name'] = "Impianto " + df.index.astype(str)
    if 'ID' not in df.columns: df['ID'] = df.index.astype(str)
    if 'Address' not in df.columns: df['Address'] = "Indirizzo non disponibile"
    
    if 'Operator' not in df.columns: 
        df['Operator'] = "Operatore Non Specificato"
    else:
        df['Operator'] = df['Operator'].astype(str).str.upper()
        df['Operator'] = df['Operator'].str.replace(r'[^\w\s]', '', regex=True)
        df['Operator'] = df['Operator'].str.replace(r'\s+', ' ', regex=True).str.strip()

    alloc_col = next((c for c in df.columns if 'alloc' in c.lower() or 'capacit' in c.lower()), None)
    if alloc_col: df['Size'] = pd.to_numeric(df[alloc_col], errors='coerce').fillna(5000)
    else: df['Size'] = 5000
    
    max_size = df['Size'].max()
    if max_size > 0: df['Display_Size'] = (df['Size'] / max_size) * 30 + 5
    else: df['Display_Size'] = 10
    
    return df

@st.cache_data(ttl=3600)
def get_live_eu_ets_price():
    if OFFLINE_MODE:
        return 70.00
    try: return round(float(yf.Ticker("KEZ=F").history(period="1d")['Close'].iloc[-1]), 2)
    except: return 70.00


def parse_coordinate_input(raw_value):
    text = str(raw_value or "").strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None

@st.cache_data
def load_taxonomy_json(file_content_or_path="taxonomy.json"):
    eligible_prefixes = set()
    try:
        if os.path.exists(file_content_or_path):
            with open(file_content_or_path, 'r', encoding='utf-8', errors='ignore') as f: data = json.load(f)
            for activity in data.get('activities', []):
                for code in activity.get('nace_codes', []):
                    if code:
                        num_part = re.sub(r'^[A-Z]+', '', code.strip(), flags=re.IGNORECASE).replace('.', '')
                        if len(num_part) == 1: num_part = "0" + num_part
                        if num_part: eligible_prefixes.add(num_part)
        return eligible_prefixes
    except: return set()

@st.cache_data
def load_nace_hierarchy(file_content_or_path="NACE_Rev.2.1.rdf"):
    try:
        if os.path.exists(file_content_or_path):
            with open(file_content_or_path, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
            labels_by_code = {}
            for match in re.finditer(r'<rdf:Description[^>]*rdf:about="http://data\.europa\.eu/ux2/nace2\.1/(?:[^"]*_)?([A-V]|\d{2,4})"[^>]*>(.*?)</rdf:Description>', content, re.DOTALL):
                code, block = match.group(1).strip(), match.group(2)
                if code not in labels_by_code: labels_by_code[code] = {}
                for l_m in re.finditer(r'<skos:prefLabel[^>]*xml:lang="(it|en)"[^>]*>(.*?)</skos:prefLabel>', block, re.DOTALL | re.IGNORECASE):
                    labels_by_code[code][l_m.group(1).lower()] = re.sub(r'\s+', ' ', l_m.group(2)).strip()

            sections, divisions, groups, classes = {}, {}, {}, {}
            for code, langs in labels_by_code.items():
                label = langs.get('it', langs.get('en', f"Attività {code}"))
                label = re.sub(rf'^{code}\s*-?\s*', '', label)
                if code.isalpha() and len(code) == 1: sections[code] = {'label': f"{code} - {label}", 'children': {}}
                elif code.isdigit() and len(code) == 2: divisions[code] = {'label': f"{code} - {label}", 'children': {}}
                elif len(code) == 3 and code.isdigit(): groups[code] = {'label': f"{code[:2]}.{code[2:]} - {label}", 'children': {}}
                elif len(code) == 4 and code.isdigit(): classes[code] = {'label': f"{code[:2]}.{code[2:]} - {label}", 'code': f"{code[:2]}.{code[2:]}"}

            def get_section(d):
                d = int(d)
                if 1<=d<=3: return 'A'
                if 5<=d<=9: return 'B'
                if 10<=d<=33: return 'C'
                if d==35: return 'D'
                if 36<=d<=39: return 'E'
                if 41<=d<=43: return 'F'
                if 45<=d<=47: return 'G'
                if 49<=d<=53: return 'H'
                if 55<=d<=56: return 'I'
                if 58<=d<=63: return 'J'
                if 64<=d<=66: return 'K'
                if d==68: return 'L'
                if 69<=d<=75: return 'M'
                if 77<=d<=82: return 'N'
                return 'UNKNOWN'

            for d_code, d_data in divisions.items():
                s_code = get_section(d_code)
                if s_code in sections: sections[s_code]['children'][d_code] = d_data
            for g_code, g_data in groups.items():
                if g_code[:2] in divisions: divisions[g_code[:2]]['children'][g_code] = g_data
            for c_code, c_data in classes.items():
                if c_code[:3] in groups: groups[c_code[:3]]['children'][c_code] = c_data
                    
            ui_db = {}
            for s in sections.values():
                if not s['children']: continue
                ui_db[s['label']] = {}
                for d in s['children'].values():
                    if not d['children']: continue
                    ui_db[s['label']][d['label']] = {}
                    for g in d['children'].values():
                        if not g['children']: continue
                        ui_db[s['label']][d['label']][g['label']] = {c['label']: c['code'] for c in g['children'].values()}
            return ui_db
    except: return {}

@st.cache_data
def generate_offline_data():
    data = []
    for c in ['Stati Uniti', 'Cina', 'Germania', 'Italia', 'India']:
        for s in ['Net Zero 2050 (Ordinata)', 'Transizione Ritardata (Shock)', 'Politiche Attuali (BAU)']:
            for y in range(2020, 2055, 5):
                p = (y - 2020) * 12 if 'Net Zero' in s else ((y - 2020) * 2 if 'BAU' in s else (10 if y < 2030 else (y - 2030) * 20 + 20))
                if c in ['Germania', 'Italia']: p *= 1.4 
                elif c in ['India', 'Cina']: p *= 0.5 
                data.append({'Scenario': s, 'Paese': c, 'Anno': y, 'Prezzo Carbonio Base': p})
    return pd.DataFrame(data)
df_base = generate_offline_data()

@st.cache_data(ttl=600)
def load_cbam_hierarchy(file_path="cn_codes_clean.csv"):
    tree = {}
    try:
        if os.path.exists(file_path):
            df_cn = pd.read_csv(file_path, dtype=str)
            df_cn['CN_Code'] = df_cn['CN_Code'].fillna("00000000").str.zfill(8)
            df_cn['Description'] = df_cn['Description'].fillna("Nessuna descrizione")
            cur_s, cur_c, cur_h = "Sconosciuta", "Sconosciuto", "Voce Sconosciuta"
            for _, row in df_cn.iterrows():
                code, desc = row['CN_Code'], str(row['Description']).strip()
                if desc.upper().startswith("SEZIONE"): cur_s, cur_c, cur_h = desc, "Sconosciuto", "Voce Sconosciuta"
                elif desc.upper().startswith("CAPITOLO"): cur_c, cur_h = desc, "Voce Sconosciuta"
                elif code.endswith('0000') and code[2:4] != '00': cur_h = f"{code[:4]} - {desc[5:].strip() if desc.startswith(code[:4]) else desc}"
                else:
                    if code[2:4] == '00' and not desc.upper().startswith(("SEZIONE", "CAPITOLO")): continue
                    if cur_s not in tree: tree[cur_s] = {}
                    if cur_c not in tree[cur_s]: tree[cur_s][cur_c] = {}
                    if cur_h not in tree[cur_s][cur_c]: tree[cur_s][cur_c][cur_h] = {}
                    tree[cur_s][cur_c][cur_h][f"{code} - {desc}"] = code
            return tree
    except: pass
    return {"SEZIONE V (FALLBACK)": {"CAPITOLO 25": {"2523 - Cementi": {"25231000 - Cemento": "25231000"}}}}
cbam_tree = load_cbam_hierarchy()
def check_cbam_category(cn_code):
    cn = str(cn_code).strip()
    if cn.startswith(('25070080', '2523')): return "Cemento"
    if cn == '27160000': return "Elettricità"
    if cn == '28041000': return "Idrogeno"
    if cn.startswith(('2808', '2814', '28342100', '3102', '3105')): return "Fertilizzanti"
    if cn.startswith(('72', '7301', '7302', '7303', '7304', '7305', '7306', '7307', '7308', '7309', '7310', '7311', '7318', '7326')): return "Ferro e Acciaio"
    if cn.startswith(('7601', '7603', '7604', '7605', '7606', '7607', '7608', '7609')): return "Alluminio"
    return "Non Soggetto"


# =====================================================================
# SIDEBAR GLOBALE (INPUT DATI AZIENDALI)
# =====================================================================
with st.sidebar:
    st.markdown("## Test di Assoggettabilità")

    if OFFLINE_MODE:
        st.info("Modalità offline completa: tutte le tab restano disponibili. Le parti che richiedono internet usano cache locale o input manuali.")

    st.session_state.company_name = st.text_input("Nome Azienda", value=st.session_state.get('company_name', ''))
    st.session_state.ragione_sociale = st.text_input("Ragione sociale", value=st.session_state.get('ragione_sociale', ''))
    st.session_state.codice_nace_ateco = st.text_input("Codice NACE/ATECO", value=st.session_state.get('codice_nace_ateco', ''))

    # Scelta UE o Extra-UE
    if 'impresa_area' not in st.session_state:
        st.session_state.impresa_area = 'UE'
    
    impresa_area = st.radio("Sede Legale", ["UE", "Extra-UE"], key='impresa_area', horizontal=True)
    st.text_input("Indirizzo sede", key="hq_address", placeholder="Es. Via Roma 1, Milano, Italia")
    if OFFLINE_MODE:
        c_hq_lat, c_hq_lon = st.columns(2)
        with c_hq_lat:
            hq_lat_text = st.text_input(
                "Latitudine sede",
                value="" if st.session_state.get('hq_lat') is None else f"{float(st.session_state.get('hq_lat')):.6f}",
                key="hq_lat_text",
                placeholder="45.464203"
            )
        with c_hq_lon:
            hq_lon_text = st.text_input(
                "Longitudine sede",
                value="" if st.session_state.get('hq_lon') is None else f"{float(st.session_state.get('hq_lon')):.6f}",
                key="hq_lon_text",
                placeholder="9.189982"
            )

        parsed_hq_lat = parse_coordinate_input(hq_lat_text)
        parsed_hq_lon = parse_coordinate_input(hq_lon_text)
        if hq_lat_text.strip() and parsed_hq_lat is None:
            st.warning("Latitudine sede non valida.")
        if hq_lon_text.strip() and parsed_hq_lon is None:
            st.warning("Longitudine sede non valida.")
        st.session_state.hq_lat = parsed_hq_lat
        st.session_state.hq_lon = parsed_hq_lon
    
    # Campi condizionali per UE
    if st.session_state.impresa_area == "UE":
        st.session_state.dipendenti = unbounded_number_input(
            "Numero dipendenti",
            value_key="dipendenti",
            display_key="dipendenti_input",
            placeholder="0",
            allow_float=False,
        )
        st.session_state.revenue = unbounded_number_input(
            "Fatturato netto (€)",
            value_key="revenue",
            display_key="revenue_input",
            placeholder="0",
        )
        st.session_state.totale_attivo = unbounded_number_input(
            "Totale bilancio / Attivo (€)",
            value_key="totale_attivo",
            display_key="totale_attivo_input",
            placeholder="0",
        )
        st.checkbox("Azienda quotata in borsa?", key="quotata")

        dip = st.session_state.dipendenti
        rev = st.session_state.revenue
        attivo = st.session_state.totale_attivo
        quotata = st.session_state.quotata

        has_dip = dip > 0
        has_rev = rev > 0
        has_attivo = attivo > 0
        n_provided = sum([has_dip, has_rev, has_attivo])

        if n_provided >= 2:
            # ---------- Grandi Imprese (CSRD) ----------
            # Richiede entrambi: dipendenti > 1.000 E fatturato netto > €450 mln
            esrs_obbligato = has_dip and has_rev and dip > 1000 and rev > 450_000_000

            # ---------- PMI Quotate (LSME) ----------
            # 51-1.000 dip AND 11-450 mln rev AND 5,5-25 mln attivo (2 su 3) + quotata in borsa
            lsme_checks = []
            if has_dip:
                lsme_checks.append(51 <= dip <= 1000)
            if has_rev:
                lsme_checks.append(11_000_000 <= rev <= 450_000_000)
            if has_attivo:
                lsme_checks.append(5_500_000 <= attivo <= 25_000_000)
            lsme_match = sum(lsme_checks) >= 2 and quotata

            # ---------- Escluse da CSRD (251-1.000 dipendenti) ----------
            csrd_escluse = has_dip and 251 <= dip <= 1000

            # ---------- Medie Imprese (VSME) ----------
            # < 250 dip AND < 50 mln rev AND < 25 mln attivo (2 su 3)
            medie_checks = []
            if has_dip:
                medie_checks.append(dip < 250)
            if has_rev:
                medie_checks.append(rev < 50_000_000)
            if has_attivo:
                medie_checks.append(attivo < 25_000_000)
            medie_match = sum(medie_checks) >= 2

            # ---------- Piccole Imprese (VSME) ----------
            # ≤ 50 dip AND ≤ 11 mln rev AND ≤ 5,5 mln attivo (2 su 3)
            piccole_checks = []
            if has_dip:
                piccole_checks.append(dip <= 50)
            if has_rev:
                piccole_checks.append(rev <= 11_000_000)
            if has_attivo:
                piccole_checks.append(attivo <= 5_500_000)
            piccole_match = sum(piccole_checks) >= 2

            # ---------- Micro Imprese (VSME) ----------
            # ≤ 10 dip AND ≤ 900k rev AND ≤ 450k attivo (2 su 3)
            micro_checks = []
            if has_dip:
                micro_checks.append(dip <= 10)
            if has_rev:
                micro_checks.append(rev <= 900_000)
            if has_attivo:
                micro_checks.append(attivo <= 450_000)
            micro_match = sum(micro_checks) >= 2

            if esrs_obbligato:
                st.session_state.status_normativo = "CSRD_GRANDE"
                st.error(
                    "**Esito: OBBLIGO ESRS** – L'impresa è soggetta agli standard europei completi (ESRS) "
                    "e non può utilizzare questa app. "
                    "(Dipendenti > 1.000 e fatturato netto > €450 mln)",
                    icon="⚖️",
                )
            elif lsme_match:
                st.session_state.status_normativo = "LSME"
                st.error(
                    "**Esito: Standard LSME** – Le PMI quotate in borsa devono rendicontare secondo lo "
                    "standard LSME (Listed SME) e non possono utilizzare questa app. "
                    "(51-1.000 dip., 11-450 mln € fatturato, 5,5-25 mln € attivo – 2 su 3 – e quotata in borsa)",
                    icon="📋",
                )
            elif csrd_escluse:
                st.session_state.status_normativo = "VSME"
                st.success(
                    "**Esito: Esclusa da CSRD** ⏳ – L'impresa (tra 251 e 1.000 dipendenti) è esclusa "
                    "dall'obbligo CSRD. Può fare la rendicontazione VSME con questa app in attesa di "
                    "nuova normativa dedicata.",
                )
            elif micro_match:
                st.session_state.status_normativo = "VSME"
                st.success(
                    "**Esito: Micro Impresa VSME** ✅ – L'impresa può utilizzare questa app per la "
                    "rendicontazione VSME volontaria. "
                    "(≤ 10 dipendenti, ≤ €900k fatturato, ≤ €450k attivo – 2 su 3)",
                )
            elif piccole_match:
                st.session_state.status_normativo = "VSME"
                st.success(
                    "**Esito: Piccola Impresa VSME** ✅ – L'impresa può utilizzare questa app per la "
                    "rendicontazione VSME volontaria. "
                    "(≤ 50 dipendenti, ≤ €11 mln fatturato, ≤ €5,5 mln attivo – 2 su 3)",
                )
            elif medie_match:
                st.session_state.status_normativo = "VSME"
                st.success(
                    "**Esito: Media Impresa VSME** ✅ – L'impresa può utilizzare questa app per la "
                    "rendicontazione VSME volontaria. "
                    "(< 250 dipendenti, < €50 mln fatturato, < €25 mln attivo – 2 su 3)",
                )
            else:
                st.session_state.status_normativo = "VSME"
                st.info(
                    "**Esito: Percorso volontario VSME** – L'impresa può utilizzare questa app per la "
                    "rendicontazione volontaria.",
                    icon="ℹ️",
                )
        elif n_provided == 1:
            st.caption(
                "Inserisci almeno 2 parametri su 3 (dipendenti, fatturato e totale bilancio) "
                "per completare il test di assoggettabilità."
            )
    
    # Campi condizionali per Extra-UE
    elif st.session_state.impresa_area == "Extra-UE":
        st.session_state.revenue = unbounded_number_input(
            "Fatturato netto in UE (€)",
            value_key="revenue",
            display_key="revenue_input",
            placeholder="0",
        )
        sucursale_eu_200 = st.radio(
            "Sucursale UE con fatturato netto > 200 mln €?",
            ["Sì", "No"],
            key='sucursale_eu_200',
            horizontal=True
        )

        if st.session_state.revenue > 0:
            extra_ue_esrs = st.session_state.revenue > 450000000 and sucursale_eu_200 == "Sì"
            if extra_ue_esrs:
                st.session_state.status_normativo = "CSRD_GRANDE"
                st.error("**Esito:** OBBLIGO ESRS (Impresa Extra-UE: Fatturato netto in UE > 450M € e sucursale UE > 200 mln €)", icon="⚖️")
            else:
                st.session_state.status_normativo = "VSME"
                st.success("**Esito:** Rendicontazione VSME (fatturato netto in UE <= 450M € oppure sucursale UE > 200 mln € = No)")
    project_name_input = st.text_input(
        "Salva progetto con nome",
        value=st.session_state.get('current_project_name', '') or st.session_state.get('company_name', ''),
        key="project_name_input"
    )
    effective_project_name = project_name_input.strip() or st.session_state.get('current_project_name', '').strip() or st.session_state.get('company_name', '').strip()
    app_not_eligible = st.session_state.get('status_normativo') in ('LSME', 'CSRD_GRANDE')
    if app_not_eligible:
        st.info("Il salvataggio progetto è disponibile solo per le aziende idonee all'utilizzo dell'app.", icon="🔒")
    if st.button("Salva", use_container_width=True, disabled=app_not_eligible):
        if not effective_project_name:
            st.warning("Inserisci un nome prima di salvare.")
        else:
            payload = save_project(effective_project_name)
            st.success(f"Progetto salvato: {payload['project_name']}")
            st.rerun()

    project_options = list_saved_projects()
    selected_project = st.selectbox(
        "Progetti salvati",
        options=project_options,
        format_func=lambda item: f"{item['label']} ({item['saved_at'][:16].replace('T', ' ')})" if item and item.get('saved_at') else item['label'],
        index=None,
        placeholder="Seleziona un progetto salvato",
        key="selected_saved_project_option"
    )

    c_avvia, c_cancella = st.columns(2)
    with c_avvia:
        if st.button("Avvia", use_container_width=True, disabled=selected_project is None):
            load_project(selected_project)
            st.success(f"Progetto avviato: {selected_project['label']}")
            st.rerun()
    with c_cancella:
        if st.button("Cancella", use_container_width=True, disabled=selected_project is None):
            deleted_label = selected_project['label']
            delete_project(selected_project)
            if st.session_state.get('current_project_name') == deleted_label:
                st.session_state['current_project_name'] = ''
                st.session_state['last_project_saved_at'] = ''
                st.session_state['last_project_signature'] = ''
            st.warning(f"Progetto eliminato: {deleted_label}")
            st.rerun()

    if st.session_state.get('current_project_name'):
        saved_at = st.session_state.get('last_project_saved_at', '')
        if saved_at:
            st.caption(f"Progetto corrente: {st.session_state['current_project_name']} | salvato il {saved_at[:16].replace('T', ' ')}")
        else:
            st.caption(f"Progetto corrente: {st.session_state['current_project_name']}")

    st.caption("Backup progetto per uso online o multi-dispositivo")
    export_file_name, export_file_bytes = export_project_snapshot(effective_project_name)
    st.download_button(
        "Scarica progetto JSON",
        data=export_file_bytes,
        file_name=export_file_name,
        mime="application/json",
        use_container_width=True,
    )
    uploaded_project_file = st.file_uploader(
        "Importa progetto JSON",
        type=['json'],
        key="uploaded_project_json",
    )
    if st.button("Importa progetto", use_container_width=True, disabled=uploaded_project_file is None):
        try:
            imported_payload = import_project(uploaded_project_file)
        except Exception as exc:
            st.error(f"Importazione non riuscita: {exc}")
        else:
            st.success(f"Progetto importato: {imported_payload['project_name']}")
            st.rerun()

    if os.getenv("CODESPACES") or os.getenv("STREAMLIT_RUNTIME"):
        st.caption("Per mantenere lo stesso progetto anche online, scarica il JSON e reimportalo quando riapri l'app.")


# --- CORPO PRINCIPALE E TABS ---
st.title(APP_DISPLAY_TITLE)
if OFFLINE_MODE:
    st.caption("Versione locale offline completa: report, checklist, mappa da coordinate o upload, export PDF e salvataggio progetto sono disponibili senza internet.")


def clean_question_label(question):
    return re.sub(r'\s*\(\d+ datapoints\)$', '', str(question or '').strip())


def to_italian_question_label(question_label):
    """Traduce in italiano le etichette chiave richieste per la UI."""
    label = str(question_label or '').strip()
    normalized = label.lower()
    translation_map = {
        "interview management esg": "Gestione delle interviste ESG",
        "balance/economics": "Bilancio/Economia",
        "hr reporting": "Rendicontazione HR",
        "governance overview": "Panoramica della governance",
        "environmental certifications": "Certificazioni ambientali",
        "energy consumption": "Consumo energetico",
        "ghg emissions - scope 1 & 2": "Emissioni GHG - Scope 1 e 2",
        "biodiversity metrics": "Metriche di biodiversità",
        "environmental monitoring data_water": "Uso di acqua",
    }
    return translation_map.get(normalized, label)


def to_italian_disclosure_title(disclosure_code, disclosure_title):
    title = str(disclosure_title or '').strip()
    code = str(disclosure_code or '').strip().upper()
    normalized_title = title.lower()

    if code == "B3" and normalized_title == "energy and greenhouse gas emissions":
        return "Energia ed emissioni di gas a effetto serra"
    if code == "B4":
        return "Inquinamento di acqua, aria e suolo"

    return title


def is_interview_management_question(question):
    normalized_question = clean_question_label(question).lower()
    return "interview management esg" in normalized_question


def is_b3_energy_consumption_question(question):
    normalized_question = clean_question_label(question).lower()
    return "energy consumption" in normalized_question or "consumi energetici" in normalized_question


def is_b3_ghg_scope12_question(question):
    normalized_question = clean_question_label(question).lower()
    return (
        ("ghg" in normalized_question and "scope 1" in normalized_question and "scope 2" in normalized_question)
        or "ghg emissions - scope 1 & 2" in normalized_question
        or "emissioni ghg scope 1 e scope 2" in normalized_question
    )


def is_b3_checklist_question(question):
    return is_b3_energy_consumption_question(question) or is_b3_ghg_scope12_question(question)


def is_b4_environmental_monitoring_question(question):
    normalized_question = clean_question_label(question).lower()
    return (
        "environmental monitoring data" in normalized_question
        or "dati di monitoraggio ambientale" in normalized_question
    )


def is_b5_biodiversity_metrics_question(question):
    normalized_question = clean_question_label(question).lower()
    return (
        "biodiversity metrics" in normalized_question
        or "metriche di biodivers" in normalized_question
    )


def is_b6_water_question(question):
    normalized_question = clean_question_label(question).lower()
    return (
        "environmental monitoring data_water" in normalized_question
        or "uso di acqua" in normalized_question
        or "water" in normalized_question and "monitoring" in normalized_question
    )


def b6_water_auto_answer(prev_year, curr_year):
    tables = st.session_state.get('vsme_disclosure_tables', {})
    df = tables.get('vsme_table_B6_uso_acqua')
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return None

    B6_PRELIEVO_ROWS = [
        "Prelievo da acquedotto",
        "Prelievo da pozzo",
        "Prelievo acque superficiali (es. laghi, fiumi)",
        "Prelievo di acqua da altre fonti (specificare)",
    ]
    B6_SCARICO_ROWS = [
        "Quantità di acqua scaricata",
    ]
    work_df = df.where(pd.notna(df), "")
    label_col = "Voce"
    if label_col not in work_df.columns:
        return None

    def has_any_value(row_labels, cols):
        for lbl in row_labels:
            row = work_df[work_df[label_col] == lbl]
            if row.empty:
                continue
            for col in cols:
                if col in row.columns and str(row.iloc[0][col]).strip():
                    return True
        return False

    year_cols = [c for c in [prev_year, curr_year] if c in work_df.columns]
    if not year_cols:
        return None

    part1_filled = has_any_value(B6_PRELIEVO_ROWS, year_cols)
    part2_filled = has_any_value(B6_SCARICO_ROWS, year_cols)

    if part1_filled and part2_filled:
        return 'Sì'
    if part1_filled or part2_filled:
        return 'Sì, ma necessita integrazione'
    return None


def _parse_decimal_for_calc(raw_value):
    raw = str(raw_value or "").strip().replace(" ", "")
    if not raw:
        return None
    try:
        if "," in raw and "." in raw:
            raw = raw.replace(".", "").replace(",", ".")
        elif "," in raw:
            raw = raw.replace(",", ".")
        return float(raw)
    except Exception:
        return None


def _compute_variation_percent(prev_val, curr_val):
    prev_num = _parse_decimal_for_calc(prev_val)
    curr_num = _parse_decimal_for_calc(curr_val)
    if prev_num is None or curr_num is None or prev_num == 0:
        return ""
    pct = (curr_num - prev_num) / abs(prev_num) * 100
    return f"{pct:+.2f}%"


def _latest_site_year_column(df, site_idx=1):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return None

    pattern = re.compile(rf"^Sede\s+{site_idx}\s+-\s+(\d{{4}})$")
    matched = []
    for col in df.columns:
        col_str = str(col)
        match = pattern.match(col_str)
        if match:
            try:
                matched.append((int(match.group(1)), col_str))
            except ValueError:
                continue

    if not matched:
        fallback_col = "Sede 1 - 2025"
        return fallback_col if fallback_col in df.columns else None

    matched.sort(key=lambda item: item[0])
    return matched[-1][1]


def _table_column_completion_state(df, target_col, required_row_labels=None, label_col=None):
    if df is None or not isinstance(df, pd.DataFrame) or df.empty or target_col not in df.columns:
        return 'missing'

    work_df = df.copy().where(pd.notna(df), '')
    if required_row_labels:
        if not label_col or label_col not in work_df.columns:
            return 'missing'
        work_df = work_df[work_df[label_col].astype(str).isin(required_row_labels)]
        if work_df.empty:
            return 'missing'

    total_cells = len(work_df.index)
    if total_cells == 0:
        return 'missing'

    filled_cells = work_df[target_col].apply(lambda v: str(v).strip() != '').sum()
    if filled_cells == 0:
        return 'missing'
    if filled_cells == total_cells:
        return 'full'
    return 'partial'


def b3_energy_consumption_auto_answer():
    tables = st.session_state.get('vsme_disclosure_tables', {})
    energy_df = tables.get('vsme_table_B3_energia_emissioni')
    target_col = _latest_site_year_column(energy_df, site_idx=1)
    table_state = _table_column_completion_state(energy_df, target_col)

    if table_state == 'full':
        return 'Sì'
    if table_state == 'partial':
        return 'Sì, ma necessita integrazione'
    return None


def b3_ghg_scope12_auto_answer():
    tables = st.session_state.get('vsme_disclosure_tables', {})
    ghg_df = tables.get('vsme_table_B3_ghg_emissioni')
    target_col = _latest_site_year_column(ghg_df, site_idx=1)
    required_rows = [
        "Emissioni GHG Scope 1",
        "Emissioni GHG Scope 2 (basato sulla localizzazione)",
        "Intensità GHG",
    ]
    table_state = _table_column_completion_state(
        ghg_df,
        target_col,
        required_row_labels=required_rows,
        label_col="Emissioni di gas a effetto serra",
    )

    if table_state == 'full':
        return 'Sì'
    if table_state == 'partial':
        return 'Sì, ma necessita integrazione'
    return None


def b4_environmental_monitoring_auto_answer(prev_year, curr_year):
    tables = st.session_state.get('vsme_disclosure_tables', {})
    b4_df = tables.get('vsme_table_B4_inquinamento')
    if b4_df is None or not isinstance(b4_df, pd.DataFrame) or b4_df.empty:
        return None

    emission_cols = [f"Emissioni [kg] - {prev_year}", f"Emissioni [kg] - {curr_year}"]
    existing_cols = [col for col in emission_cols if col in b4_df.columns]
    if not existing_cols:
        return None

    normalized_df = b4_df.where(pd.notna(b4_df), "")
    has_any_emission_value = False
    for col in existing_cols:
        if normalized_df[col].apply(lambda value: str(value).strip() != "").any():
            has_any_emission_value = True
            break

    return "Sì" if has_any_emission_value else None


def b5_biodiversity_metrics_auto_answer(curr_year):
    tables = st.session_state.get('vsme_disclosure_tables', {})

    first_table_state = 'missing'
    site_df = tables.get('vsme_table_B5_sedi_biodiversita')
    if site_df is not None and isinstance(site_df, pd.DataFrame) and not site_df.empty and "Metrica" in site_df.columns:
        normalized_site_df = site_df.where(pd.notna(site_df), "")
        site_cols = [col for col in normalized_site_df.columns if col != "Metrica"]
        area_row = normalized_site_df[normalized_site_df["Metrica"] == "Area (ha)"]
        protected_row = normalized_site_df[normalized_site_df["Metrica"] == "Area protetta (biodiversity sensitive area)"]

        has_area = False
        has_protected = False
        if site_cols and not area_row.empty:
            has_area = any(str(area_row.iloc[0].get(col, "")).strip() for col in site_cols)
        if site_cols and not protected_row.empty:
            has_protected = any(str(protected_row.iloc[0].get(col, "")).strip() for col in site_cols)

        if has_area and has_protected:
            first_table_state = 'full'
        elif has_area or has_protected:
            first_table_state = 'partial'

    second_table_state = 'missing'
    metrics_df = tables.get('vsme_table_B5_metriche_biodiversita')
    if metrics_df is not None and isinstance(metrics_df, pd.DataFrame) and not metrics_df.empty and curr_year in metrics_df.columns:
        normalized_metrics_df = metrics_df.where(pd.notna(metrics_df), "")
        filled_count = normalized_metrics_df[curr_year].apply(lambda value: str(value).strip() != "").sum()
        if filled_count == 4:
            second_table_state = 'full'
        elif filled_count > 0:
            second_table_state = 'partial'

    if first_table_state == 'full' and second_table_state == 'full':
        return 'Sì'
    if first_table_state in {'full', 'partial'} or second_table_state in {'full', 'partial'}:
        return 'Sì, ma necessita integrazione'
    return None


def get_question_datapoints(question):
    if is_interview_management_question(question):
        return 4
    if is_balance_economics_question(question):
        return 4
    if is_hr_reporting_question(question):
        return 1
    if is_governance_overview_question(question):
        return 2
    if is_environmental_certifications_question(question):
        return 1
    if is_hse_action_plan_question(question):
        return 1
    if is_hse_policies_question(question):
        return 1
    if is_transition_plan_question(question):
        return 2
    if is_b3_energy_consumption_question(question):
        return 1
    if is_b3_ghg_scope12_question(question):
        return 3
    if is_b4_environmental_monitoring_question(question):
        return 1
    if is_b5_biodiversity_metrics_question(question):
        return 5
    if is_b6_water_question(question):
        return 2
    match = re.search(r'\((\d+) datapoints\)', str(question or ''))
    return int(match.group(1)) if match else 0


def answer_to_status(answer):
    normalized_answer = normalize_vsme_response_label(answer).lower()
    if normalized_answer in {'sì', 'si'}:
        return 'ready', 0
    if normalized_answer == 'sì, ma necessita integrazione':
        return 'partial', 1
    if normalized_answer in {'no, ma pianificato', 'no'}:
        return 'missing', 2
    return 'partial', 1


def consolidated_table_completion_state():
    tables = st.session_state.get('vsme_disclosure_tables', {})
    df = tables.get('vsme_table_B1_consolidato')
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return 'missing'

    normalized_df = df.where(pd.notna(df), '')
    total_cells = normalized_df.shape[0] * normalized_df.shape[1]
    if total_cells == 0:
        return 'missing'

    filled_cells = 0
    for value in normalized_df.to_numpy().flatten():
        if str(value).strip() != '':
            filled_cells += 1

    if filled_cells == 0:
        return 'missing'
    if filled_cells == total_cells:
        return 'full'
    return 'partial'


def interview_management_auto_answer(selected_mode):
    # Datapoint 1 (OPTION A/B) e datapoint 2-3 (informazioni omesse + perimetro)
    # sono sempre disponibili in app. Datapoint 4 dipende dalla tabella Consolidato.
    perimetro = st.session_state.get('b1_perimetro', 'Individuale')

    if perimetro != 'Consolidato':
        return 'Sì'

    table_state = consolidated_table_completion_state()
    if table_state == 'full':
        return 'Sì'
    if table_state == 'partial':
        return 'Sì, ma necessita integrazione'
    return None


def is_balance_economics_question(question):
    return "balance/economics" in clean_question_label(question).lower()


def balance_economics_auto_answer():
    # Datapoint 1: Ragione sociale (sidebar)
    # Datapoint 2: Codice NACE/ATECO (sidebar)
    # Datapoint 3: Totale degli attivi (tabella Dimensioni, riga 0)
    # Datapoint 4: Fatturato (tabella Dimensioni, riga 1)
    filled = 0

    if str(st.session_state.get('ragione_sociale', '')).strip():
        filled += 1
    if str(st.session_state.get('codice_nace_ateco', '')).strip():
        filled += 1

    tables = st.session_state.get('vsme_disclosure_tables', {})
    df = tables.get('vsme_table_B1_dimensioni')
    if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
        year_cols = sorted([col for col in df.columns if str(col).isdigit() and len(str(col)) == 4])
        target_year_cols = year_cols[-2:] if len(year_cols) >= 2 else ['2024', '2025']
        if df.shape[0] > 0:
            row0 = df.iloc[0]
            if any(str(row0.get(col, '')).strip() for col in target_year_cols):
                filled += 1
        if df.shape[0] > 1:
            row1 = df.iloc[1]
            if any(str(row1.get(col, '')).strip() for col in target_year_cols):
                filled += 1

    if filled == 4:
        return 'Sì'
    if filled > 0:
        return 'Sì, ma necessita integrazione'
    return None


def is_hr_reporting_question(question):
    return "hr reporting" in clean_question_label(question).lower()


def hr_reporting_auto_answer():
    # Datapoint: Numero di dipendenti (tabella Dimensioni, riga 2)
    tables = st.session_state.get('vsme_disclosure_tables', {})
    df = tables.get('vsme_table_B1_dimensioni')
    if df is None or not isinstance(df, pd.DataFrame) or df.empty or df.shape[0] <= 2:
        return None

    row = df.iloc[2]
    year_cols = sorted([col for col in df.columns if str(col).isdigit() and len(str(col)) == 4])
    target_year_cols = year_cols[-2:] if len(year_cols) >= 2 else ['2024', '2025']
    val_prev = str(row.get(target_year_cols[0], '')).strip()
    val_curr = str(row.get(target_year_cols[1], '')).strip()

    if val_prev and val_curr:
        return 'Sì'
    if val_prev or val_curr:
        return 'Sì, ma necessita integrazione'
    return None


def is_governance_overview_question(question):
    return "governance overview" in clean_question_label(question).lower()


def governance_overview_auto_answer():
    # Datapoint 1: Paese delle principali operazioni e localizzazione degli asset significativi
    # Datapoint 2: Geolocalizzazione siti con coordinate GPS
    country_filled = bool(str(st.session_state.get('b1_paese_operazioni_asset_significativi', '')).strip())

    tables = st.session_state.get('vsme_disclosure_tables', {})
    geo_df = tables.get('vsme_table_B1_geolocalizzazione_siti')
    has_geo_full = False
    has_geo_partial = False

    if geo_df is not None and isinstance(geo_df, pd.DataFrame) and not geo_df.empty:
        normalized_df = geo_df.where(pd.notna(geo_df), '')
        for _, row in normalized_df.iterrows():
            site_name = str(row.get('Sito (di proprietà, in locazione o gestito)', '')).strip()
            gps_coords = str(row.get('Coordinate GPS', '')).strip()
            if site_name and gps_coords:
                has_geo_full = True
                break
            if site_name or gps_coords:
                has_geo_partial = True

    filled = int(country_filled) + int(has_geo_full)
    if filled == 2:
        return 'Sì'
    if filled > 0 or has_geo_partial:
        return 'Sì, ma necessita integrazione'
    return None


def is_environmental_certifications_question(question):
    normalized = clean_question_label(question).lower()
    return "environmental certifications" in normalized or "environmental certification" in normalized


def environmental_certifications_auto_answer():
    # Datapoint: presenza di almeno una certificazione/marchio sostenibilita con dati completi.
    tables = st.session_state.get('vsme_disclosure_tables', {})
    cert_df = tables.get('vsme_table_B1_certificazioni_sostenibilita')
    if cert_df is None or not isinstance(cert_df, pd.DataFrame) or cert_df.empty:
        return None

    has_full = False
    has_partial = False
    normalized_df = cert_df.where(pd.notna(cert_df), '')
    for _, row in normalized_df.iterrows():
        description = str(row.get('Breve descrizione', '')).strip()
        body = str(row.get('Organismo di certificazione', '')).strip()
        date_value = str(row.get('Data', '')).strip()
        score = str(row.get('Punteggio', '')).strip()
        filled_count = sum(bool(v) for v in [description, body, date_value, score])
        if filled_count == 4:
            has_full = True
            break
        if filled_count > 0:
            has_partial = True

    if has_full:
        return 'Sì'
    if has_partial:
        return 'Sì, ma necessita integrazione'
    return None


# ---------------------------------------------------------------
# B2 TABLE HELPERS
# ---------------------------------------------------------------

B2_TEMI = [
    "Cambiamento climatico",
    "Inquinamento",
    "Acqua e risorse marine",
    "Biodiversità ed ecosistemi",
    "Economia circolare",
    "Forza lavoro propria",
    "Lavoratori nella catena del valore",
    "Comunità interessate",
    "Consumatori e utenti finali",
    "Condotta aziendale",
]

B2_COLONNE = [
    "Esistono pratiche che affrontano dei temi di sostenibilità?",
    "Esistono politiche di sostenibilità che affrontano dei temi di sostenibilità?",
    "Le politiche sono disponibili pubblicamente?",
    "Le politiche hanno degli obiettivi?",
    "Esistono iniziative future che affrontano dei temi di sostenibilità?",
]


def get_b2_table():
    """Ritorna il DataFrame della tabella B2 dalla sessione, inizializzandolo se assente."""
    tables = st.session_state.get('vsme_disclosure_tables', {})
    key = 'vsme_table_B2_pratiche_politiche'
    all_cols = ["Tema"] + B2_COLONNE
    if key not in tables or not isinstance(tables[key], pd.DataFrame) or tables[key].empty:
        tables[key] = pd.DataFrame(
            [{"Tema": tema, **{col: "" for col in B2_COLONNE}} for tema in B2_TEMI]
        )
        st.session_state['vsme_disclosure_tables'] = tables
    else:
        df = tables[key].copy()
        for col in all_cols:
            if col not in df.columns:
                df[col] = ""
        # Converti NaN / None / 'nan' / 'None' in stringa vuota
        for col in B2_COLONNE:
            df[col] = df[col].fillna("").astype(str).replace({'nan': '', 'None': ''})
        tables[key] = df
    return tables[key]


def _b2_col_has_si(col_name):
    """Ritorna True se almeno una riga della colonna col_name ha valore 'Sì'."""
    df = get_b2_table()
    if col_name not in df.columns:
        return False
    return any(str(v).strip() == "Sì" for v in df[col_name])


def is_b2_checklist_question(question):
    """True per le domande auto-calcolate dalla tabella B2 (scope B2, non B1)."""
    return (
        is_hse_action_plan_question(question)
        or is_hse_policies_question(question)
        or is_transition_plan_question(question)
    )


def is_hse_action_plan_question(question):
    return "hse action plan" in clean_question_label(question).lower()


def hse_action_plan_auto_answer():
    pratiche_col = B2_COLONNE[0]
    if _b2_col_has_si(pratiche_col):
        return 'Sì'
    return None


def is_hse_policies_question(question):
    return "hse policies" in clean_question_label(question).lower()


def hse_policies_auto_answer():
    politiche_col = B2_COLONNE[1]
    if _b2_col_has_si(politiche_col):
        return 'Sì'
    return None


def is_transition_plan_question(question):
    return "transition plan" in clean_question_label(question).lower()


def transition_plan_auto_answer():
    # dp1: ≥1 Sì in colonna "obiettivi"
    # dp2: ≥1 Sì in colonna "iniziative future"
    obiettivi_col = B2_COLONNE[3]
    iniziative_col = B2_COLONNE[4]
    dp1 = _b2_col_has_si(obiettivi_col)
    dp2 = _b2_col_has_si(iniziative_col)
    if dp1 and dp2:
        return 'Sì'
    if dp1 or dp2:
        return 'Sì, ma necessita integrazione'
    return None


def summarize_gap_answers(module_prefixes=None):
    rows = []
    for key, data in st.session_state.get("gap_answers", {}).items():
        if module_prefixes and not any(key.startswith(prefix) for prefix in module_prefixes):
            continue

        answer = str(data.get("ans", "")).strip()
        question_raw = str(data.get("q", "")).strip()
        question = clean_question_label(question_raw)
        if not answer:
            continue

        status, severity = answer_to_status(answer)
        datapoints = get_question_datapoints(question_raw)

        rows.append({
            "question": question,
            "answer": answer,
            "status": status,
            "severity": severity,
            "datapoints": datapoints,
        })

    total = len(rows)
    ready = sum(1 for row in rows if row["status"] == "ready")
    partial = sum(1 for row in rows if row["status"] == "partial")
    missing = sum(1 for row in rows if row["status"] == "missing")

    total_datapoints = sum(row["datapoints"] for row in rows)
    ready_datapoints = sum(row["datapoints"] for row in rows if row["status"] == "ready")
    partial_datapoints = sum(row["datapoints"] for row in rows if row["status"] == "partial")
    missing_datapoints = sum(row["datapoints"] for row in rows if row["status"] == "missing")
    coverage = ((ready_datapoints + partial_datapoints) / total_datapoints * 100) if total_datapoints else 0.0

    top_gaps = []
    seen_questions = set()
    for row in sorted(rows, key=lambda item: (-item["severity"], -item["datapoints"], item["question"])):
        if row["status"] == "ready" or row["question"] in seen_questions:
            continue
        top_gaps.append(row["question"])
        seen_questions.add(row["question"])
        if len(top_gaps) == 5:
            break

    return {
        "total": total,
        "ready": ready,
        "partial": partial,
        "missing": missing,
        "total_datapoints": total_datapoints,
        "ready_datapoints": ready_datapoints,
        "partial_datapoints": partial_datapoints,
        "missing_datapoints": missing_datapoints,
        "coverage": coverage,
        "top_gaps": top_gaps,
    }

def build_priority_actions(max_risk_score=None, top_risk_asset=None):
    gap_summary = summarize_gap_answers(["vsme_base", "vsme_comprehensive"])
    actions = []

    if gap_summary["missing"] > 0:
        actions.append("Recuperare i dati VSME assenti con priorita alta, partendo dalle domande senza evidenze.")
    if gap_summary["partial"] > 0:
        actions.append("Consolidare i dati parziali con fonti interne e documenti di supporto caricati in app.")
    if not st.session_state.get("hq_address", "").strip():
        actions.append("Inserire l'indirizzo sede per completare l'inquadramento geografico di base.")
    if st.session_state.get("portfolio_df", pd.DataFrame()).empty:
        actions.append("Caricare almeno una sede o un portfolio asset per rendere utile la lettura del rischio fisico.")
    elif max_risk_score is not None and max_risk_score >= 70:
        asset_label = top_risk_asset or "piu esposto"
        actions.append(f"Approfondire l'asset {asset_label} con una verifica tecnica dedicata sul rischio fisico.")
    if get_tot_emissions() == 0:
        actions.append("Raccogliere i consumi energetici e logistici minimi per avviare una baseline GHG credibile.")
    elif st.session_state.get("perc_red", 0) < 20:
        actions.append("Definire un primo piano di riduzione emissioni per migliorare la resilienza allo scenario di transizione.")

    deduplicated_actions = []
    seen_actions = set()
    for action in actions:
        if action in seen_actions:
            continue
        deduplicated_actions.append(action)
        seen_actions.add(action)

    return deduplicated_actions[:5]

st.markdown(
    """
    <style>
    /* ---- Sticky tab bar (tutte le sezioni) ---- */
    [data-baseweb="tab-list"] {
        position: sticky !important;
        top: 0 !important;
        z-index: 100 !important;
        background: #ffffff !important;
        padding-top: 4px !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.08) !important;
    }
    /* ---- Stili tab principali ---- */
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"] {
        font-size: 1rem;
        font-weight: 600;
        border-radius: 8px 8px 0 0;
        padding: 0.5rem 0.95rem;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(1) {
        color: #1f4e79;
        background: #eaf3ff;
        border: 1px solid #bfd7f2;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(1) {
        color: #0f3554;
        background: #d6e9ff;
        border-bottom-color: #d6e9ff;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(2) {
        color: #2f6b3b;
        background: #e7f4ea;
        border: 1px solid #b9d8c0;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(2) {
        color: #214d2a;
        background: #d7eddc;
        border-bottom-color: #d7eddc;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(3) {
        color: #8a6a00;
        background: #fff4d6;
        border: 1px solid #ecd89b;
    }
    div[data-testid="stVerticalBlock"] > div:has(.app-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(3) {
        color: #6c5200;
        background: #f8e7b0;
        border-bottom-color: #f8e7b0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
st.markdown("<div class='app-main-tabs-marker'></div>", unsafe_allow_html=True)
t_triage, t_rischi, t_down = st.tabs([
    "🔎 Diagnosi Azienda", "🗺️ Piano di Azione & Rischi", "📦 Deliverable"
])

# =====================================================================
# TAB 0: TRIAGE NORMATIVO, GAP ANALYSIS E DOPPIA MATERIALITÀ
# =====================================================================
with t_triage:
    status_normativo = st.session_state.get("status_normativo", "VSME")
    st.caption("Percorso 1 di 3")
    st.subheader("Diagnosi azienda")
    st.write("Questa sezione serve a inquadrare il caso, verificare cosa si applica e identificare i gap documentali e informativi piu rilevanti.")
    
    def render_gap_list(questions, pillar_code, tab_context, scale_options, prefix="gap", question_filter=None, manage_cleanup=True):
        with tab_context:
            if manage_cleanup:
                valid_answer_keys = {f"{prefix}_{pillar_code}_{i}" for i in range(len(questions))}
                prefix_scope = f"{prefix}_{pillar_code}_"

                for key in list(st.session_state.get("gap_answers", {}).keys()):
                    if key.startswith(prefix_scope) and key not in valid_answer_keys:
                        del st.session_state.gap_answers[key]

                for doc_key in list(st.session_state.get("gap_documents", {}).keys()):
                    if doc_key.startswith(prefix_scope) and doc_key not in valid_answer_keys:
                        del st.session_state.gap_documents[doc_key]

                for state_key in list(st.session_state.keys()):
                    if not state_key.startswith("show_pdf_"):
                        continue
                    doc_key = state_key.replace("show_pdf_", "", 1)
                    if doc_key.startswith(prefix_scope) and doc_key not in valid_answer_keys:
                        del st.session_state[state_key]

            display_index = 1

            for i, q in enumerate(questions):
                if question_filter is not None and not question_filter(q):
                    continue

                # Creare colonne: selectbox + bottone PDF allineati
                col_q, col_pdf = st.columns([4, 1], vertical_alignment="bottom")
                
                with col_q:
                    answer_key = f"{prefix}_{pillar_code}_{i}"
                    is_interview_row = is_interview_management_question(q)
                    is_balance_row = is_balance_economics_question(q)
                    is_hr_row = is_hr_reporting_question(q)
                    is_governance_row = is_governance_overview_question(q)
                    is_env_cert_row = is_environmental_certifications_question(q)
                    is_hse_ap_row = is_hse_action_plan_question(q)
                    is_hse_pol_row = is_hse_policies_question(q)
                    is_transition_row = is_transition_plan_question(q)
                    is_b3_energy_row = is_b3_energy_consumption_question(q)
                    is_b3_ghg_row = is_b3_ghg_scope12_question(q)
                    is_b4_monitoring_row = is_b4_environmental_monitoring_question(q)
                    is_b5_biodiversity_row = is_b5_biodiversity_metrics_question(q)
                    is_b6_water_row = is_b6_water_question(q)
                    is_auto_row = (is_interview_row or is_balance_row or is_hr_row
                                   or is_governance_row or is_env_cert_row
                                   or is_hse_ap_row or is_hse_pol_row or is_transition_row
                                   or is_b3_energy_row or is_b3_ghg_row or is_b4_monitoring_row
                                   or is_b5_biodiversity_row or is_b6_water_row)

                    if is_auto_row:
                        auto_datapoints = get_question_datapoints(q)
                        translated_question = to_italian_question_label(clean_question_label(q))
                        question_label = f"{display_index}. {translated_question} ({auto_datapoints} datapoints)"
                        if is_interview_row:
                            auto_answer = interview_management_auto_answer(selected_mode)
                        elif is_balance_row:
                            auto_answer = balance_economics_auto_answer()
                        elif is_hr_row:
                            auto_answer = hr_reporting_auto_answer()
                        elif is_governance_row:
                            auto_answer = governance_overview_auto_answer()
                        elif is_env_cert_row:
                            auto_answer = environmental_certifications_auto_answer()
                        elif is_hse_ap_row:
                            auto_answer = hse_action_plan_auto_answer()
                        elif is_hse_pol_row:
                            auto_answer = hse_policies_auto_answer()
                        elif is_b3_energy_row:
                            auto_answer = b3_energy_consumption_auto_answer()
                        elif is_b3_ghg_row:
                            auto_answer = b3_ghg_scope12_auto_answer()
                        elif is_b4_monitoring_row:
                            compilation_year = datetime.now().year
                            prev_year = str(compilation_year - 2)
                            curr_year = str(compilation_year - 1)
                            auto_answer = b4_environmental_monitoring_auto_answer(prev_year, curr_year)
                        elif is_b5_biodiversity_row:
                            compilation_year = datetime.now().year
                            curr_year = str(compilation_year - 1)
                            auto_answer = b5_biodiversity_metrics_auto_answer(curr_year)
                        elif is_b6_water_row:
                            compilation_year = datetime.now().year
                            prev_year = str(compilation_year - 2)
                            curr_year = str(compilation_year - 1)
                            auto_answer = b6_water_auto_answer(prev_year, curr_year)
                        else:
                            auto_answer = transition_plan_auto_answer()

                        if auto_answer is not None:
                            st.session_state[answer_key] = auto_answer
                            val = st.selectbox(
                                question_label,
                                [auto_answer],
                                key=answer_key,
                                disabled=True,
                            )
                        else:
                            if is_b4_monitoring_row or is_b6_water_row:
                                manual_options = ["Sì, ma necessita integrazione", "No, ma pianificato", "No"]
                            else:
                                manual_options = ["No, ma pianificato", "No"]
                            if st.session_state.get(answer_key) not in manual_options:
                                st.session_state[answer_key] = manual_options[0]
                            val = st.selectbox(
                                question_label,
                                manual_options,
                                key=answer_key,
                            )

                        stored_question = f"{clean_question_label(q)} ({auto_datapoints} datapoints)"
                    else:
                        translated_q = to_italian_question_label(clean_question_label(q))
                        current_answer = st.session_state.get(answer_key)
                        if current_answer is not None:
                            normalized_current = normalize_vsme_response_label(current_answer)
                            if normalized_current in scale_options and normalized_current != current_answer:
                                st.session_state[answer_key] = normalized_current
                        val = st.selectbox(
                            f"{display_index}. {translated_q}",
                            scale_options,
                            key=answer_key
                        )
                        stored_question = q

                    st.session_state.gap_answers[answer_key] = {
                        "ans": val,
                        "pillar": pillar_code,
                        "q": stored_question,
                    }
                    display_index += 1
                
                with col_pdf:
                    doc_key = f"{prefix}_{pillar_code}_{i}"
                    
                    # Determina il label del bottone in base allo stato
                    if doc_key in st.session_state.gap_documents and st.session_state.gap_documents[doc_key]:
                        btn_label = "PDF ✅"
                    else:
                        btn_label = "Carica PDF"
                    
                    # Bottone di caricamento PDF
                    if st.button(btn_label, key=f"btn_pdf_{prefix}_{pillar_code}_{i}", use_container_width=True):
                        st.session_state[f"show_pdf_{doc_key}"] = not st.session_state.get(f"show_pdf_{doc_key}", False)
                    
                    # Mostra uploader se bottone cliccato
                    if st.session_state.get(f"show_pdf_{doc_key}", False):
                        uploaded_file = st.file_uploader(
                            "Carica PDF",
                            type="pdf",
                            key=f"uploader_{prefix}_{pillar_code}_{i}"
                        )
                        if uploaded_file:
                            # Salva il file in memoria/session state
                            st.session_state.gap_documents[doc_key] = {
                                'filename': uploaded_file.name,
                                'data': uploaded_file.getvalue(),
                                'timestamp': datetime.now().isoformat()
                            }
                            st.success(f"✅ PDF caricato: {uploaded_file.name}")

    def render_disclosure_readiness(questions, pillar_code, prefix, disclosure_label, question_filter):
        ordered_options = list(VSME_PLOT_RESPONSE_ORDER)
        response_totals = {option: 0 for option in ordered_options}
        detail_rows = []

        for i, q in enumerate(questions):
            if question_filter is not None and not question_filter(q):
                continue

            answer_key = f"{prefix}_{pillar_code}_{i}"
            answer_data = st.session_state.get("gap_answers", {}).get(answer_key)
            if not answer_data:
                continue

            question_label = answer_data.get("q", q)
            clean_q = clean_question_label(question_label)
            display_q = to_italian_question_label(clean_q)
            datapoints = get_question_datapoints(question_label)
            response_label = to_vsme_plot_response_label(answer_data.get("ans", ""))
            if response_label not in response_totals:
                response_totals[response_label] = 0
            response_totals[response_label] += datapoints

            detail_rows.append({
                "Domanda": display_q,
                "Datapoints": datapoints,
                "Risposta": response_label,
            })

        total_datapoints = sum(response_totals.values())
        distribution_rows = []
        for option in ordered_options:
            dp = response_totals.get(option, 0)
            pct = (dp / total_datapoints * 100) if total_datapoints else 0.0
            distribution_rows.append({
                "Risposta": option,
                "Datapoints": dp,
                "% sul totale disclosure": pct,
            })

        df_distribution = pd.DataFrame(distribution_rows)
        color_map = {
            "Sì": "#2ecc71",
            "Sì, ma necessita integrazione": "#f39c12",
            "No, ma pianificato": "#3498db",
            "No": "#e74c3c",
        }

        st.markdown(f"##### Readiness - {disclosure_label}")

        details_df = pd.DataFrame(detail_rows)
        if not details_df.empty and total_datapoints:
            details_df["% sul totale disclosure"] = details_df["Datapoints"].apply(lambda x: f"{(x / total_datapoints * 100):.2f}%")
            details_df = details_df[["Domanda", "Datapoints", "% sul totale disclosure", "Risposta"]]
        st.dataframe(details_df, width='stretch', hide_index=True)

        fig = px.bar(
            df_distribution,
            x="Risposta",
            y="% sul totale disclosure",
            color="Risposta",
            text=df_distribution["% sul totale disclosure"].apply(lambda x: f"{x:.2f}%"),
            color_discrete_map=color_map,
            category_orders={"Risposta": ordered_options},
        )
        fig.update_layout(
            showlegend=False,
            yaxis_title="Percentuale sul totale disclosure",
            xaxis_title="Risposta",
        )
        st.plotly_chart(fig, width='stretch', key=f"readiness_{prefix}_{pillar_code}_{disclosure_label}")

        table_distribution = df_distribution.copy()
        table_distribution["% sul totale disclosure"] = table_distribution["% sul totale disclosure"].apply(lambda x: f"{x:.2f}%")
        st.dataframe(table_distribution, width='stretch', hide_index=True)

    def render_vsme_disclosure_tables(
        pillar_code,
        selected_mode,
        tab_context,
        disclosure_reference,
        questions_for_pillar=None,
        gap_scale_options=None,
        gap_prefix="gap",
    ):
        pillar_styles = {
            "GEN": {
                "tab_bg": "#eaf3ff",
                "tab_bg_selected": "#d6e9ff",
                "tab_border": "#bfd7f2",
                "tab_text": "#1f4e79",
                "tab_text_selected": "#0f3554",
                "badge_bg": "#d9ecff",
                "badge_text": "#1f4e79",
            },
            "E": {
                "tab_bg": "#e7f4ea",
                "tab_bg_selected": "#d7eddc",
                "tab_border": "#b9d8c0",
                "tab_text": "#2f6b3b",
                "tab_text_selected": "#214d2a",
                "badge_bg": "#dff1e3",
                "badge_text": "#2f6b3b",
            },
            "S": {
                "tab_bg": "#fde8e4",
                "tab_bg_selected": "#f9d7cf",
                "tab_border": "#f2c4ba",
                "tab_text": "#a54b3f",
                "tab_text_selected": "#7d362d",
                "badge_bg": "#fbe0db",
                "badge_text": "#a54b3f",
            },
            "G": {
                "tab_bg": "#fff4d6",
                "tab_bg_selected": "#f8e7b0",
                "tab_border": "#ecd89b",
                "tab_text": "#8a6a00",
                "tab_text_selected": "#6c5200",
                "badge_bg": "#fff0bf",
                "badge_text": "#8a6a00",
            },
        }
        style = pillar_styles.get(pillar_code, pillar_styles["GEN"])

        if selected_mode == "absolute":
            module_prefixes = ("B", "C")
        elif selected_mode == "base":
            module_prefixes = ("B",)
        elif selected_mode == "comprehensive":
            module_prefixes = ("B", "C")
        else:
            module_prefixes = ("C",)

        disclosures = [
            item for item in disclosure_reference.get(pillar_code, [])
            if item["code"].startswith(module_prefixes)
        ]

        if not disclosures:
            return

        marker_class = f"vsme-subtabs-marker-{pillar_code}-{selected_mode}"

        with tab_context:
            if pillar_code == "GEN" and questions_for_pillar is not None and gap_scale_options is not None:
                # Pulizia chiavi una volta sola per GEN; il rendering filtrato B1/B2 avviene nei rispettivi tab.
                render_gap_list(
                    questions_for_pillar,
                    pillar_code,
                    st.container(),
                    gap_scale_options,
                    gap_prefix,
                    question_filter=lambda _q: False,
                    manage_cleanup=True,
                )
            elif pillar_code == "E" and questions_for_pillar is not None and gap_scale_options is not None:
                # Pulizia chiavi una volta sola per Ambiente; il rendering filtrato B3 avviene nel relativo tab.
                render_gap_list(
                    questions_for_pillar,
                    pillar_code,
                    st.container(),
                    gap_scale_options,
                    gap_prefix,
                    question_filter=lambda _q: False,
                    manage_cleanup=True,
                )

            css_rules = [
                "font-size: 1rem;",
                "font-weight: 600;",
                f"color: {style['tab_text']} !important;",
                f"background: {style['tab_bg']} !important;",
                f"border: 1px solid {style['tab_border']} !important;",
                "border-radius: 8px 8px 0 0;",
                "padding: 0.45rem 0.9rem;",
            ]
            css_selected_rules = [
                f"color: {style['tab_text_selected']} !important;",
                f"background: {style['tab_bg_selected']} !important;",
                f"border-bottom-color: {style['tab_bg_selected']} !important;",
            ]

            st.markdown(
                f"""
                <style>
                div[data-testid="stVerticalBlock"] > div:has(.{marker_class}) ~ div div[data-baseweb="tab-list"] {{
                    position: sticky !important;
                    top: 2.5rem !important;
                    z-index: 90 !important;
                    background: #ffffff !important;
                    padding-top: 4px !important;
                    box-shadow: 0 2px 6px rgba(0,0,0,0.08) !important;
                }}
                div[data-testid="stVerticalBlock"] > div:has(.{marker_class}) ~ div div[data-baseweb="tab-list"] button[role="tab"] {{
                    {' '.join(css_rules)}
                }}
                div[data-testid="stVerticalBlock"] > div:has(.{marker_class}) ~ div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"] {{
                    {' '.join(css_selected_rules)}
                }}
                </style>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(f"<div class='{marker_class}'></div>", unsafe_allow_html=True)
            st.markdown(
                "<p style='font-size:1rem; font-weight:600; margin-bottom:0.45rem; color:#1f1f1f;'>Struttura dati da compilare</p>",
                unsafe_allow_html=True,
            )
            disclosure_tabs = st.tabs([f"{disclosure['code']}" for disclosure in disclosures])

            for disclosure_tab, disclosure in zip(disclosure_tabs, disclosures):
                with disclosure_tab:
                    display_disclosure_title = to_italian_disclosure_title(disclosure.get('code'), disclosure.get('title'))
                    st.markdown(
                        f"<p style='display:inline-block; font-size:1rem; font-weight:600; margin-bottom:0.45rem; padding:0.25rem 0.6rem; background:{style['badge_bg']}; color:{style['badge_text']}; border-radius:6px;'>{disclosure['code']} - {display_disclosure_title}</p>",
                        unsafe_allow_html=True,
                    )

                    b1_is_custom = disclosure['code'] == "B1"
                    b2_is_custom = disclosure['code'] == "B2"
                    b3_is_custom = disclosure['code'] == "B3"
                    b4_is_custom = disclosure['code'] == "B4"
                    b5_is_custom = disclosure['code'] == "B5"
                    b6_is_custom = disclosure['code'] == "B6"
                    if b1_is_custom:
                        st.radio(
                            'Informazioni omesse in quanto ritenute riservate o sensibili?',
                            ['Si', 'No'],
                            key='b1_info_omesse',
                            horizontal=True,
                        )
                        if st.session_state.b1_info_omesse == 'Si':
                            st.text_input('Quali?', key='b1_info_omesse_quali')

                        st.radio(
                            'Perimetro rendicontazione',
                            ['Individuale', 'Consolidato'],
                            key='b1_perimetro',
                            horizontal=True,
                        )

                        if st.session_state.b1_perimetro == 'Consolidato':
                            editor_storage_key = "vsme_table_B1_consolidato"
                            b1_default_columns = [
                                'Tipo sede (es. sede legale, magazzino, stabilimento industriale, ecc.)',
                                'N. siti',
                                'Indirizzo',
                                'Codice Postale',
                                'Città',
                                'Paese',
                            ]
                            if editor_storage_key not in st.session_state.vsme_disclosure_tables:
                                st.session_state.vsme_disclosure_tables[editor_storage_key] = pd.DataFrame([
                                    {col: '' for col in b1_default_columns}
                                    for _ in range(2)
                                ])
                            else:
                                existing_df = st.session_state.vsme_disclosure_tables[editor_storage_key].copy()
                                for col in b1_default_columns:
                                    if col not in existing_df.columns:
                                        existing_df[col] = ''
                                existing_df = existing_df[b1_default_columns]
                                if existing_df.empty:
                                    existing_df = pd.DataFrame([{col: '' for col in b1_default_columns} for _ in range(2)])
                                existing_df = existing_df.where(pd.notna(existing_df), '')
                                st.session_state.vsme_disclosure_tables[editor_storage_key] = existing_df

                            st.markdown(
                                """
                                <style>
                                div[data-testid="stDataEditor"] thead th,
                                div[data-testid="stDataFrame"] thead th {
                                    color: #000000 !important;
                                }
                                div[data-testid="stDataEditor"] [role="columnheader"] {
                                    color: #000000 !important;
                                }
                                </style>
                                """,
                                unsafe_allow_html=True,
                            )

                            consolidato_col_config = {
                                'Tipo sede (es. sede legale, magazzino, stabilimento industriale, ecc.)': st.column_config.TextColumn(
                                    'Tipo sede (es. sede legale, magazzino, stabilimento industriale, ecc.)',
                                    width='medium',
                                    help='Tipologia della sede (es. sede legale, stabilimento, magazzino, ecc.)'
                                ),
                                'N. siti': st.column_config.TextColumn(
                                    'N. siti',
                                    width='small',
                                    help='Numero di siti di questa tipologia'
                                ),
                                'Indirizzo': st.column_config.TextColumn(
                                    'Indirizzo',
                                    width='medium',
                                    help='Indirizzo fisico della sede'
                                ),
                                'Codice Postale': st.column_config.TextColumn(
                                    'Codice Postale',
                                    width='small',
                                    help='Codice postale della sede'
                                ),
                                'Città': st.column_config.TextColumn(
                                    'Città',
                                    width='small',
                                    help='Città dove è localizzata la sede'
                                ),
                                'Paese': st.column_config.TextColumn(
                                    'Paese',
                                    width='small',
                                    help='Paese di localizzazione della sede'
                                ),
                            }

                            edited_df = st.data_editor(
                                st.session_state.vsme_disclosure_tables[editor_storage_key],
                                key=f"editor_{editor_storage_key}",
                                column_config=consolidato_col_config,
                                width='stretch',
                                hide_index=True,
                                num_rows="fixed",
                            )
                            edited_df = edited_df.where(pd.notna(edited_df), '')
                            st.session_state.vsme_disclosure_tables[editor_storage_key] = edited_df

                            if st.button("+ Aggiungi riga", key="b1_add_row_btn"):
                                table_df = st.session_state.vsme_disclosure_tables[editor_storage_key].copy()
                                new_row = pd.DataFrame([{col: '' for col in b1_default_columns}])
                                table_df = pd.concat([table_df, new_row], ignore_index=True)
                                st.session_state.vsme_disclosure_tables[editor_storage_key] = table_df
                                st.rerun()

                        dimensioni_storage_key = "vsme_table_B1_dimensioni"
                        compilation_year = datetime.now().year
                        dim_prev_year_col = str(compilation_year - 2)
                        dim_curr_year_col = str(compilation_year - 1)

                        if dimensioni_storage_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[dimensioni_storage_key] = pd.DataFrame([
                                {"Dimensione": "Totale degli attivi", dim_prev_year_col: "", dim_curr_year_col: ""},
                                {"Dimensione": "Fatturato", dim_prev_year_col: "", dim_curr_year_col: ""},
                                {"Dimensione": "Numero di dipendenti", dim_prev_year_col: "", dim_curr_year_col: ""},
                            ])
                        else:
                            dim_df = st.session_state.vsme_disclosure_tables[dimensioni_storage_key].copy()
                            if dim_prev_year_col not in dim_df.columns and '2024' in dim_df.columns:
                                dim_df[dim_prev_year_col] = dim_df['2024']
                            if dim_curr_year_col not in dim_df.columns and '2025' in dim_df.columns:
                                dim_df[dim_curr_year_col] = dim_df['2025']

                            expected_columns = ["Dimensione", dim_prev_year_col, dim_curr_year_col]
                            for col in expected_columns:
                                if col not in dim_df.columns:
                                    dim_df[col] = ""
                            dim_df = dim_df[expected_columns].where(pd.notna(dim_df[expected_columns]), '')
                            if dim_df.empty:
                                dim_df = pd.DataFrame([
                                    {"Dimensione": "Totale degli attivi", dim_prev_year_col: "", dim_curr_year_col: ""},
                                    {"Dimensione": "Fatturato", dim_prev_year_col: "", dim_curr_year_col: ""},
                                    {"Dimensione": "Numero di dipendenti", dim_prev_year_col: "", dim_curr_year_col: ""},
                                ])
                            st.session_state.vsme_disclosure_tables[dimensioni_storage_key] = dim_df

                        st.markdown("##### Dimensioni")
                        dim_col_config = {
                            "Dimensione": st.column_config.TextColumn(
                                "Dimensione",
                                width='medium',
                                help="Indicatore di dimensione aziendale"
                            ),
                            dim_prev_year_col: st.column_config.TextColumn(
                                dim_prev_year_col,
                                width='small',
                                help=f"Valore relativo all'anno {compilation_year - 2}"
                            ),
                            dim_curr_year_col: st.column_config.TextColumn(
                                dim_curr_year_col,
                                width='small',
                                help=f"Valore relativo all'anno {compilation_year - 1}"
                            ),
                        }
                        dim_edited_df = st.data_editor(
                            st.session_state.vsme_disclosure_tables[dimensioni_storage_key],
                            key=f"editor_{dimensioni_storage_key}",
                            column_config=dim_col_config,
                            width='stretch',
                            hide_index=True,
                            num_rows="fixed",
                            disabled=["Dimensione"],
                        )
                        dim_edited_df = dim_edited_df.where(pd.notna(dim_edited_df), '')
                        st.session_state.vsme_disclosure_tables[dimensioni_storage_key] = dim_edited_df

                        st.markdown("##### Paese")
                        st.caption("delle principali operazioni e localizzazione degli asset significativi")
                        st.text_input(
                            "Paese",
                            key='b1_paese_operazioni_asset_significativi',
                            label_visibility="collapsed",
                            placeholder="Inserisci il paese"
                        )

                        geoloc_storage_key = "vsme_table_B1_geolocalizzazione_siti"
                        geoloc_columns = ["Sito (di proprietà, in locazione o gestito)", "Coordinate GPS"]
                        if geoloc_storage_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[geoloc_storage_key] = pd.DataFrame([
                                {col: '' for col in geoloc_columns}
                            ])
                        else:
                            geo_df = st.session_state.vsme_disclosure_tables[geoloc_storage_key].copy()
                            legacy_site_col = "Sito (owned, leased or managed)"
                            new_site_col = "Sito (di proprietà, in locazione o gestito)"
                            if legacy_site_col in geo_df.columns and new_site_col not in geo_df.columns:
                                geo_df[new_site_col] = geo_df[legacy_site_col]
                            for col in geoloc_columns:
                                if col not in geo_df.columns:
                                    geo_df[col] = ''
                            geo_df = geo_df[geoloc_columns].where(pd.notna(geo_df[geoloc_columns]), '')
                            if geo_df.empty:
                                geo_df = pd.DataFrame([{col: '' for col in geoloc_columns}])
                            st.session_state.vsme_disclosure_tables[geoloc_storage_key] = geo_df

                        if st.session_state.b1_perimetro == 'Consolidato':
                            consolidato_df = st.session_state.vsme_disclosure_tables.get("vsme_table_B1_consolidato")
                            if consolidato_df is not None and isinstance(consolidato_df, pd.DataFrame) and not consolidato_df.empty:
                                site_col = 'Tipo sede (es. sede legale, magazzino, stabilimento industriale, ecc.)'
                                consolidato_sites = []
                                if site_col in consolidato_df.columns:
                                    for value in consolidato_df[site_col].tolist():
                                        site_name = str(value or '').strip()
                                        if site_name:
                                            consolidato_sites.append(site_name)

                                if consolidato_sites:
                                    current_geo = st.session_state.vsme_disclosure_tables[geoloc_storage_key].copy()
                                    existing_names = {
                                        str(v).strip().lower()
                                        for v in current_geo["Sito (di proprietà, in locazione o gestito)"].tolist()
                                        if str(v).strip()
                                    }
                                    sites_to_insert = [
                                        site for site in consolidato_sites
                                        if site.lower() not in existing_names
                                    ]
                                    site_col = "Sito (di proprietà, in locazione o gestito)"
                                    for idx, row in current_geo.iterrows():
                                        if not sites_to_insert:
                                            break
                                        current_value = str(row.get(site_col, '')).strip()
                                        if not current_value:
                                            current_geo.at[idx, site_col] = sites_to_insert.pop(0)
                                    if sites_to_insert:
                                        rows_to_add = [
                                            {site_col: site, "Coordinate GPS": ''}
                                            for site in sites_to_insert
                                        ]
                                        current_geo = pd.concat([current_geo, pd.DataFrame(rows_to_add)], ignore_index=True)
                                    st.session_state.vsme_disclosure_tables[geoloc_storage_key] = current_geo

                        st.markdown("##### Geolocalizzazione siti")
                        geo_col_config = {
                            "Sito (di proprietà, in locazione o gestito)": st.column_config.TextColumn(
                                "Sito (di proprietà, in locazione o gestito)",
                                width='medium',
                                help="Nome o descrizione del sito"
                            ),
                            "Coordinate GPS": st.column_config.TextColumn(
                                "Coordinate GPS",
                                width='medium',
                                help="Coordinate GPS nel formato: latitudine, longitudine (es. 45.4642, 9.1900)"
                            ),
                        }
                        geo_edited_df = st.data_editor(
                            st.session_state.vsme_disclosure_tables[geoloc_storage_key],
                            key=f"editor_{geoloc_storage_key}",
                            column_config=geo_col_config,
                            width='stretch',
                            hide_index=True,
                            num_rows="fixed",
                        )
                        geo_edited_df = geo_edited_df.where(pd.notna(geo_edited_df), '')
                        st.session_state.vsme_disclosure_tables[geoloc_storage_key] = geo_edited_df

                        if st.button("+ Aggiungi riga sito", key="b1_add_row_geoloc_btn"):
                            geo_table_df = st.session_state.vsme_disclosure_tables[geoloc_storage_key].copy()
                            new_geo_row = pd.DataFrame([{col: '' for col in geoloc_columns}])
                            geo_table_df = pd.concat([geo_table_df, new_geo_row], ignore_index=True)
                            st.session_state.vsme_disclosure_tables[geoloc_storage_key] = geo_table_df
                            st.rerun()

                        cert_storage_key = "vsme_table_B1_certificazioni_sostenibilita"
                        cert_columns = ["Breve descrizione", "Organismo di certificazione", "Data", "Punteggio"]
                        if cert_storage_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[cert_storage_key] = pd.DataFrame([
                                {col: '' for col in cert_columns}
                            ])
                        else:
                            cert_df = st.session_state.vsme_disclosure_tables[cert_storage_key].copy()
                            for col in cert_columns:
                                if col not in cert_df.columns:
                                    cert_df[col] = ''
                            cert_df = cert_df[cert_columns].where(pd.notna(cert_df[cert_columns]), '')
                            if cert_df.empty:
                                cert_df = pd.DataFrame([{col: '' for col in cert_columns}])
                            st.session_state.vsme_disclosure_tables[cert_storage_key] = cert_df

                        st.markdown("##### Certificazioni o marchi relativi alla sostenibilità")
                        cert_col_config = {
                            "Breve descrizione": st.column_config.TextColumn(
                                "Breve descrizione",
                                width='medium',
                                help="Descrizione della certificazione o marchio"
                            ),
                            "Organismo di certificazione": st.column_config.TextColumn(
                                "Organismo di certificazione",
                                width='medium',
                                help="Nome dell'ente certificatore"
                            ),
                            "Data": st.column_config.TextColumn(
                                "Data",
                                width='small',
                                help="Data di ottenimento della certificazione (es. 2024-01-15)"
                            ),
                            "Punteggio": st.column_config.TextColumn(
                                "Punteggio",
                                width='small',
                                help="Punteggio o rating della certificazione (se applicabile)"
                            ),
                        }
                        cert_edited_df = st.data_editor(
                            st.session_state.vsme_disclosure_tables[cert_storage_key],
                            key=f"editor_{cert_storage_key}",
                            column_config=cert_col_config,
                            width='stretch',
                            hide_index=True,
                            num_rows="fixed",
                        )
                        cert_edited_df = cert_edited_df.where(pd.notna(cert_edited_df), '')
                        st.session_state.vsme_disclosure_tables[cert_storage_key] = cert_edited_df

                        if st.button("+ Aggiungi riga certificazione", key="b1_add_row_cert_btn"):
                            cert_table_df = st.session_state.vsme_disclosure_tables[cert_storage_key].copy()
                            new_cert_row = pd.DataFrame([{col: '' for col in cert_columns}])
                            cert_table_df = pd.concat([cert_table_df, new_cert_row], ignore_index=True)
                            st.session_state.vsme_disclosure_tables[cert_storage_key] = cert_table_df
                            st.rerun()

                        if pillar_code == "GEN" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B1")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=lambda q: not is_b2_checklist_question(q),
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B1",
                                question_filter=lambda q: not is_b2_checklist_question(q),
                            )

                    elif b2_is_custom:
                        st.caption(
                            "Per ogni tema selezionato, dovrà essere compilata la sezione C2 del Modulo Completo."
                        )

                        b2_key = 'vsme_table_B2_pratiche_politiche'
                        b2_df = get_b2_table()
                        b2_df = b2_df.where(pd.notna(b2_df), '')
                        for col in B2_COLONNE:
                            if col in b2_df.columns:
                                b2_df[col] = b2_df[col].astype(str).replace({'None': '', 'nan': ''})

                        b2_display_df = b2_df.copy()
                        for col in B2_COLONNE:
                            if col in b2_display_df.columns:
                                b2_display_df[col] = b2_display_df[col].apply(lambda v: str(v).strip() == "Sì")

                        st.markdown(
                            """
                            <style>
                            div[data-testid="stDataEditor"] [role="columnheader"] {
                                height: 96px !important;
                                min-height: 96px !important;
                                max-height: none !important;
                                align-items: flex-start !important;
                                padding-top: 6px !important;
                            }
                            div[data-testid="stDataEditor"] [role="columnheader"] div {
                                white-space: pre-wrap !important;
                                overflow-wrap: anywhere !important;
                                line-height: 1.15 !important;
                                height: auto !important;
                                max-height: none !important;
                            }
                            </style>
                            """,
                            unsafe_allow_html=True,
                        )

                        column_config_b2 = {
                            "Tema": st.column_config.TextColumn(
                                "Tema",
                                disabled=True,
                                width="medium",
                                help="Tema di sostenibilità di riferimento",
                            ),
                        }
                        b2_header_map = {
                            B2_COLONNE[0]: "Esistono pratiche che affrontano\ndei temi di sostenibilità?",
                            B2_COLONNE[1]: "Esistono politiche di sostenibilità che affrontano\ndei temi di sostenibilità?",
                            B2_COLONNE[2]: "Le politiche sono\ndisponibili pubblicamente?",
                            B2_COLONNE[3]: "Le politiche hanno\ndegli obiettivi?",
                            B2_COLONNE[4]: "Esistono iniziative future che affrontano\ndei temi di sostenibilità?",
                        }
                        b2_help_map = {
                            B2_COLONNE[0]: "Indicare la presenza di pratiche concrete",
                            B2_COLONNE[1]: "Indicare la presenza di politiche documentate",
                            B2_COLONNE[2]: "Indicare se le politiche sono pubblicamente accessibili",
                            B2_COLONNE[3]: "Indicare la presenza di obiettivi quantificati",
                            B2_COLONNE[4]: "Indicare la presenza di piani futuri pianificati",
                        }
                        for col in B2_COLONNE:
                            column_config_b2[col] = st.column_config.CheckboxColumn(
                                b2_header_map.get(col, col),
                                width="small",
                                help=b2_help_map.get(col, col),
                            )

                        b2_edited = st.data_editor(
                            b2_display_df,
                            key=f"editor_{b2_key}",
                            column_config=column_config_b2,
                            hide_index=True,
                            num_rows="fixed",
                            use_container_width=True,
                        )

                        saved_b2 = b2_df.copy()
                        for col in B2_COLONNE:
                            if col in b2_edited.columns:
                                saved_b2[col] = b2_edited[col].fillna(False).apply(lambda v: "Sì" if bool(v) else "No")
                        st.session_state.vsme_disclosure_tables[b2_key] = saved_b2

                        if pillar_code == "GEN" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B2")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=is_b2_checklist_question,
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B2",
                                question_filter=is_b2_checklist_question,
                            )

                    elif b3_is_custom:
                        compilation_year = datetime.now().year
                        prev_year = compilation_year - 1
                        prev_prev_year = compilation_year - 2
                        b3_key = 'vsme_table_B3_energia_emissioni'
                        b3_col_label = "Energia"
                        b3_col_udm = "UdM"
                        b3_rows = [
                            ("Energia elettrica acquistata senza GdO", "kWh"),
                            ("Energia elettrica acquistata con GdO", "kWh"),
                            ("Energia elettrica autoprodotta da fonti rinnovabili", "kWh"),
                            ("Benzina per autotrazione", "litri"),
                            ("Gasolio per autotrazione", "litri"),
                            ("GPL per autotrazione", "kg"),
                            ("Metano per autotrazione", "m³"),
                            ("Metano per alimentazione caldaie/macchinari", "m³"),
                            ("GPL per alimentazione caldaie/macchinari", "kg"),
                            ("Gasolio per alimentazione caldaie/macchinari", "litri"),
                        ]

                        def _sede_prev_col(site_idx):
                            return f"Sede {site_idx} - {prev_prev_year}"

                        def _sede_curr_col(site_idx):
                            return f"Sede {site_idx} - {prev_year}"

                        def _sede_var_col(site_idx):
                            if site_idx == 1:
                                return "Variazione %"
                            return f"Variazione % Sede {site_idx}"

                        def _compute_variazione_b3(row, prev_col, curr_col):
                            try:
                                prev_val = float(str(row[prev_col]).replace(",", ".").strip())
                                curr_val = float(str(row[curr_col]).replace(",", ".").strip())
                                if prev_val == 0:
                                    return ""
                                return f"{round((curr_val - prev_val) / abs(prev_val) * 100, 2):+.2f}%"
                            except (ValueError, TypeError, KeyError):
                                return ""

                        base_data = {b3_col_label: [], b3_col_udm: []}
                        for label, udm in b3_rows:
                            base_data[b3_col_label].append(label)
                            base_data[b3_col_udm].append(udm)
                        default_df = pd.DataFrame(base_data)
                        default_df[_sede_prev_col(1)] = ""
                        default_df[_sede_curr_col(1)] = ""

                        if b3_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[b3_key] = default_df.copy()

                        b3_df = st.session_state.vsme_disclosure_tables[b3_key].copy()
                        if b3_df.empty:
                            b3_df = default_df.copy()

                        if b3_col_label not in b3_df.columns:
                            b3_df[b3_col_label] = [label for label, _ in b3_rows]
                        if b3_col_udm not in b3_df.columns:
                            b3_df[b3_col_udm] = [udm for _, udm in b3_rows]

                        # Mantieni struttura righe fissa e ordinata.
                        b3_df = b3_df.set_index(b3_col_label, drop=False)
                        for label, udm in b3_rows:
                            if label not in b3_df.index:
                                b3_df.loc[label, b3_col_label] = label
                                b3_df.loc[label, b3_col_udm] = udm
                        b3_df = b3_df.loc[[label for label, _ in b3_rows]].reset_index(drop=True)
                        b3_df[b3_col_udm] = [udm for _, udm in b3_rows]

                        # Numero sedi gestito tramite session state dedicato.
                        b3_num_sites_key = 'b3_num_sites'
                        if b3_num_sites_key not in st.session_state:
                            st.session_state[b3_num_sites_key] = 1
                        site_numbers = list(range(1, st.session_state[b3_num_sites_key] + 1))

                        for site_idx in site_numbers:
                            prev_col = _sede_prev_col(site_idx)
                            curr_col = _sede_curr_col(site_idx)
                            if prev_col not in b3_df.columns:
                                b3_df[prev_col] = ""
                            if curr_col not in b3_df.columns:
                                b3_df[curr_col] = ""

                        # Opzioni UdM disponibili per energia/emissioni.
                        b3_udm_options = [
                            "kWh", "MWh", "GJ", "litri", "m³", "Nm³", "kg", "t",
                            "tCO2eq", "kgCO2eq", "MJ", "toe",
                        ]
                        # Assicura che il valore corrente sia sempre nella lista.
                        for udm_val in b3_df[b3_col_udm].unique():
                            if str(udm_val).strip() and str(udm_val).strip() not in b3_udm_options:
                                b3_udm_options.insert(0, str(udm_val).strip())

                        # Costruisci dataframe di visualizzazione con colonne variazione auto-calcolate.
                        ordered_display_cols = [b3_col_label, b3_col_udm]
                        column_config_b3 = {
                            b3_col_label: st.column_config.TextColumn(
                                b3_col_label,
                                disabled=True,
                                width="medium",
                                help=b3_col_label,
                            ),
                            b3_col_udm: st.column_config.SelectboxColumn(
                                b3_col_udm,
                                options=b3_udm_options,
                                required=False,
                                width="small",
                                help="Unità di misura — modificabile",
                            ),
                        }

                        for site_idx in sorted(site_numbers):
                            prev_col = _sede_prev_col(site_idx)
                            curr_col = _sede_curr_col(site_idx)
                            var_col = _sede_var_col(site_idx)
                            b3_df[var_col] = b3_df.apply(lambda row, p=prev_col, c=curr_col: _compute_variazione_b3(row, p, c), axis=1)
                            ordered_display_cols.extend([prev_col, curr_col, var_col])
                            column_config_b3[prev_col] = st.column_config.TextColumn(prev_col, width="small", help=f"Valore relativo all'anno {compilation_year - 2}")
                            column_config_b3[curr_col] = st.column_config.TextColumn(curr_col, width="small", help=f"Valore relativo all'anno {compilation_year - 1}")
                            column_config_b3[var_col] = st.column_config.TextColumn(var_col, disabled=True, width="small", help="Variazione % calcolata automaticamente")

                        df_b3_display = b3_df[ordered_display_cols].copy()

                        edited_b3 = st.data_editor(
                            df_b3_display,
                            key=f"editor_{b3_key}_{st.session_state[b3_num_sites_key]}",
                            column_config=column_config_b3,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                        )

                        editable_cols = [b3_col_label, b3_col_udm]
                        for site_idx in sorted(site_numbers):
                            editable_cols.extend([_sede_prev_col(site_idx), _sede_curr_col(site_idx)])
                        saved_b3 = edited_b3[[c for c in editable_cols if c in edited_b3.columns]].copy()
                        saved_b3 = saved_b3.where(pd.notna(saved_b3), "")
                        st.session_state.vsme_disclosure_tables[b3_key] = saved_b3

                        if st.button("+ Aggiungi sede", key="b3_add_site_btn"):
                            st.session_state[b3_num_sites_key] += 1
                            st.rerun()

                        # ── Tabella GHG Emissioni ─────────────────────────────────────────
                        st.markdown("---")
                        b3_ghg_key = 'vsme_table_B3_ghg_emissioni'
                        b3_ghg_col_label = "Emissioni di gas a effetto serra"
                        b3_ghg_col_udm = "UdM"

                        # 4 righe fisse; Intensità GHG è auto-calcolata.
                        b3_ghg_rows = [
                            ("Emissioni GHG Scope 1", "ton CO2eq"),
                            ("Emissioni GHG Scope 2 (basato sulla localizzazione)", "ton CO2eq"),
                            ("Emissioni GHG Scope 3 (basato sulla localizzazione)", "ton CO2eq"),
                            ("Intensità GHG", "ton CO2eq/€"),
                        ]
                        GHG_SCOPE1_LABEL = "Emissioni GHG Scope 1"
                        GHG_SCOPE2_LABEL = "Emissioni GHG Scope 2 (basato sulla localizzazione)"
                        GHG_INTENSITY_LABEL = "Intensità GHG"

                        def _ghg_sede_prev_col(site_idx):
                            return f"Sede {site_idx} - {prev_prev_year}"

                        def _ghg_sede_curr_col(site_idx):
                            return f"Sede {site_idx} - {prev_year}"

                        def _ghg_sede_var_col(site_idx):
                            return "Variazione %" if site_idx == 1 else f"Variazione % Sede {site_idx}"

                        def _compute_variazione_ghg(row, p, c):
                            try:
                                pv = float(str(row[p]).replace(",", ".").strip())
                                cv = float(str(row[c]).replace(",", ".").strip())
                                if pv == 0:
                                    return ""
                                return f"{round((cv - pv) / abs(pv) * 100, 2):+.2f}%"
                            except (ValueError, TypeError, KeyError):
                                return ""

                        ghg_base = {b3_ghg_col_label: [], b3_ghg_col_udm: []}
                        for lbl, udm in b3_ghg_rows:
                            ghg_base[b3_ghg_col_label].append(lbl)
                            ghg_base[b3_ghg_col_udm].append(udm)
                        ghg_default_df = pd.DataFrame(ghg_base)
                        ghg_default_df[_ghg_sede_prev_col(1)] = ""
                        ghg_default_df[_ghg_sede_curr_col(1)] = ""

                        if b3_ghg_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[b3_ghg_key] = ghg_default_df.copy()

                        ghg_df = st.session_state.vsme_disclosure_tables[b3_ghg_key].copy()
                        if ghg_df.empty:
                            ghg_df = ghg_default_df.copy()

                        if b3_ghg_col_label not in ghg_df.columns:
                            ghg_df[b3_ghg_col_label] = [l for l, _ in b3_ghg_rows]
                        if b3_ghg_col_udm not in ghg_df.columns:
                            ghg_df[b3_ghg_col_udm] = [u for _, u in b3_ghg_rows]

                        # Struttura righe fissa.
                        ghg_df = ghg_df.set_index(b3_ghg_col_label, drop=False)
                        for lbl, udm in b3_ghg_rows:
                            if lbl not in ghg_df.index:
                                ghg_df.loc[lbl, b3_ghg_col_label] = lbl
                                ghg_df.loc[lbl, b3_ghg_col_udm] = udm
                        ghg_df = ghg_df.loc[[l for l, _ in b3_ghg_rows]].reset_index(drop=True)
                        ghg_df[b3_ghg_col_udm] = [u for _, u in b3_ghg_rows]

                        # Numero sedi per la tabella GHG (condiviso con la tabella energia).
                        b3_ghg_num_sites_key = 'b3_ghg_num_sites'
                        if b3_ghg_num_sites_key not in st.session_state:
                            st.session_state[b3_ghg_num_sites_key] = 1
                        ghg_site_numbers = list(range(1, st.session_state[b3_ghg_num_sites_key] + 1))

                        for site_idx in ghg_site_numbers:
                            pc = _ghg_sede_prev_col(site_idx)
                            cc = _ghg_sede_curr_col(site_idx)
                            if pc not in ghg_df.columns:
                                ghg_df[pc] = ""
                            if cc not in ghg_df.columns:
                                ghg_df[cc] = ""

                        # Parser numerico robusto (supporta 1.234,56 e 1234.56).
                        def _parse_number(raw_val):
                            try:
                                raw = str(raw_val).strip().replace(" ", "")
                                if not raw:
                                    return None
                                if "," in raw and "." in raw:
                                    # Caso EU: 1.234,56
                                    clean = raw.replace(".", "").replace(",", ".")
                                elif "," in raw:
                                    clean = raw.replace(",", ".")
                                else:
                                    clean = raw
                                return float(clean)
                            except Exception:
                                return None

                        # Leggi Fatturato dalla tabella Dimensioni (riga 1) per anno corrente e precedente.
                        def _get_fatturato_safe(year_col):
                            try:
                                dim_df = st.session_state.vsme_disclosure_tables.get('vsme_table_B1_dimensioni')
                                if dim_df is None or dim_df.shape[0] < 2:
                                    return None
                                return _parse_number(dim_df.iloc[1].get(year_col, ""))
                            except Exception:
                                return None

                        # Costruisci visualizzazione con Intensità GHG calcolata e variazioni.
                        ghg_ordered_cols = [b3_ghg_col_label, b3_ghg_col_udm]
                        ghg_column_config = {
                            b3_ghg_col_label: st.column_config.TextColumn(
                                b3_ghg_col_label,
                                disabled=True,
                                width="medium",
                                help=b3_ghg_col_label,
                            ),
                            b3_ghg_col_udm: st.column_config.TextColumn(
                                b3_ghg_col_udm,
                                disabled=True,
                                width="small",
                                help="Unità di misura",
                            ),
                        }

                        ghg_df_display = ghg_df[[b3_ghg_col_label, b3_ghg_col_udm]].copy()
                        fat_prev = _get_fatturato_safe(str(compilation_year - 2))
                        fat_curr = _get_fatturato_safe(str(compilation_year - 1))

                        for site_idx in sorted(ghg_site_numbers):
                            pc = _ghg_sede_prev_col(site_idx)
                            cc = _ghg_sede_curr_col(site_idx)
                            vc = _ghg_sede_var_col(site_idx)

                            ghg_df_display[pc] = ghg_df[pc].copy()
                            ghg_df_display[cc] = ghg_df[cc].copy()

                            # Calcola Intensità GHG: (Scope1 + Scope2) / Fatturato
                            try:
                                s1_prev = _parse_number(ghg_df.loc[ghg_df[b3_ghg_col_label] == GHG_SCOPE1_LABEL, pc].values[0])
                            except Exception:
                                s1_prev = None
                            try:
                                s2_prev = _parse_number(ghg_df.loc[ghg_df[b3_ghg_col_label] == GHG_SCOPE2_LABEL, pc].values[0])
                            except Exception:
                                s2_prev = None
                            try:
                                s1_curr = _parse_number(ghg_df.loc[ghg_df[b3_ghg_col_label] == GHG_SCOPE1_LABEL, cc].values[0])
                            except Exception:
                                s1_curr = None
                            try:
                                s2_curr = _parse_number(ghg_df.loc[ghg_df[b3_ghg_col_label] == GHG_SCOPE2_LABEL, cc].values[0])
                            except Exception:
                                s2_curr = None

                            if s1_prev is not None and s2_prev is not None and fat_prev and fat_prev != 0:
                                ghg_intensity_prev = f"{(s1_prev + s2_prev) / fat_prev:.6g}"
                            else:
                                ghg_intensity_prev = ""
                            if s1_curr is not None and s2_curr is not None and fat_curr and fat_curr != 0:
                                ghg_intensity_curr = f"{(s1_curr + s2_curr) / fat_curr:.6g}"
                            else:
                                ghg_intensity_curr = ""

                            # Forza i valori di Intensità GHG nel display df
                            intensity_mask = ghg_df_display[b3_ghg_col_label] == GHG_INTENSITY_LABEL
                            ghg_df_display.loc[intensity_mask, pc] = ghg_intensity_prev
                            ghg_df_display.loc[intensity_mask, cc] = ghg_intensity_curr

                            # Calcola variazione per tutte le righe
                            def _var_ghg_row(row, p=pc, c=cc):
                                try:
                                    pv = float(str(row[p]).replace(",", ".").strip())
                                    cv = float(str(row[c]).replace(",", ".").strip())
                                    if pv == 0:
                                        return ""
                                    return f"{round((cv - pv) / abs(pv) * 100, 2):+.2f}%"
                                except Exception:
                                    return ""
                            ghg_df_display[vc] = ghg_df_display.apply(_var_ghg_row, axis=1)

                            ghg_ordered_cols.extend([pc, cc, vc])
                            ghg_column_config[pc] = st.column_config.TextColumn(pc, width="small", help=f"Valore relativo all'anno {compilation_year - 2}")
                            ghg_column_config[cc] = st.column_config.TextColumn(cc, width="small", help=f"Valore relativo all'anno {compilation_year - 1}")
                            ghg_column_config[vc] = st.column_config.TextColumn(vc, disabled=True, width="small", help="Variazione % calcolata automaticamente")

                        # Rendi Intensità GHG e Variazione % non editabili nel display.
                        # Le colonn editable sono tutte eccetto Intensità GHG (pc/cc) e var.
                        ghg_disabled_rows_info = ""
                        if fat_prev is None or fat_curr is None:
                            ghg_disabled_rows_info = "ℹ️ Il valore Intensità GHG è calcolato automaticamente da (Scope 1 + Scope 2) / Fatturato. Inserire il Fatturato nella tabella Dimensioni (sezione B1)."

                        if ghg_disabled_rows_info:
                            st.caption(ghg_disabled_rows_info)

                        edited_ghg = st.data_editor(
                            ghg_df_display[ghg_ordered_cols],
                            key=f"editor_{b3_ghg_key}_{st.session_state[b3_ghg_num_sites_key]}",
                            column_config=ghg_column_config,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                            disabled=[b3_ghg_col_label, b3_ghg_col_udm] + [_ghg_sede_var_col(s) for s in ghg_site_numbers],
                        )

                        # Salva solo le righe non-intensity (intensity è calcolata, non persistita)
                        ghg_editable_cols = [b3_ghg_col_label, b3_ghg_col_udm]
                        for site_idx in sorted(ghg_site_numbers):
                            ghg_editable_cols.extend([_ghg_sede_prev_col(site_idx), _ghg_sede_curr_col(site_idx)])
                        saved_ghg = edited_ghg[[c for c in ghg_editable_cols if c in edited_ghg.columns]].copy()
                        # Ricalcola e persisti Intensità GHG dai valori appena inseriti.
                        for site_idx in sorted(ghg_site_numbers):
                            pc = _ghg_sede_prev_col(site_idx)
                            cc = _ghg_sede_curr_col(site_idx)
                            try:
                                s1_prev = _parse_number(saved_ghg.loc[saved_ghg[b3_ghg_col_label] == GHG_SCOPE1_LABEL, pc].values[0])
                            except Exception:
                                s1_prev = None
                            try:
                                s2_prev = _parse_number(saved_ghg.loc[saved_ghg[b3_ghg_col_label] == GHG_SCOPE2_LABEL, pc].values[0])
                            except Exception:
                                s2_prev = None
                            try:
                                s1_curr = _parse_number(saved_ghg.loc[saved_ghg[b3_ghg_col_label] == GHG_SCOPE1_LABEL, cc].values[0])
                            except Exception:
                                s1_curr = None
                            try:
                                s2_curr = _parse_number(saved_ghg.loc[saved_ghg[b3_ghg_col_label] == GHG_SCOPE2_LABEL, cc].values[0])
                            except Exception:
                                s2_curr = None

                            prev_intensity = ""
                            curr_intensity = ""
                            if s1_prev is not None and s2_prev is not None and fat_prev not in (None, 0):
                                prev_intensity = f"{(s1_prev + s2_prev) / fat_prev:.6g}"
                            if s1_curr is not None and s2_curr is not None and fat_curr not in (None, 0):
                                curr_intensity = f"{(s1_curr + s2_curr) / fat_curr:.6g}"

                            intensity_mask = saved_ghg[b3_ghg_col_label] == GHG_INTENSITY_LABEL
                            saved_ghg.loc[intensity_mask, pc] = prev_intensity
                            saved_ghg.loc[intensity_mask, cc] = curr_intensity
                        saved_ghg = saved_ghg.where(pd.notna(saved_ghg), "")
                        st.session_state.vsme_disclosure_tables[b3_ghg_key] = saved_ghg

                        if st.button("+ Aggiungi sede", key="b3_ghg_add_site_btn"):
                            st.session_state[b3_ghg_num_sites_key] += 1
                            st.rerun()

                        if pillar_code == "E" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B3")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=is_b3_checklist_question,
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B3",
                                question_filter=is_b3_checklist_question,
                            )

                    elif b4_is_custom:
                        compilation_year = datetime.now().year
                        prev_year = str(compilation_year - 2)
                        curr_year = str(compilation_year - 1)

                        b4_key = "vsme_table_B4_inquinamento"
                        b4_columns = [
                            "Inquinante",
                            "Rilasciato in",
                            f"Emissioni [kg] - {prev_year}",
                            f"Emissioni [kg] - {curr_year}",
                        ]
                        b4_release_options = ["Acqua", "Aria", "Suolo"]

                        if b4_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[b4_key] = pd.DataFrame([
                                {
                                    "Inquinante": "",
                                    "Rilasciato in": "Acqua",
                                    f"Emissioni [kg] - {prev_year}": "",
                                    f"Emissioni [kg] - {curr_year}": "",
                                },
                                {
                                    "Inquinante": "",
                                    "Rilasciato in": "Aria",
                                    f"Emissioni [kg] - {prev_year}": "",
                                    f"Emissioni [kg] - {curr_year}": "",
                                },
                                {
                                    "Inquinante": "",
                                    "Rilasciato in": "Suolo",
                                    f"Emissioni [kg] - {prev_year}": "",
                                    f"Emissioni [kg] - {curr_year}": "",
                                },
                            ])

                        b4_df = st.session_state.vsme_disclosure_tables[b4_key].copy()
                        for col in b4_columns:
                            if col not in b4_df.columns:
                                b4_df[col] = ""
                        b4_df = b4_df[b4_columns].where(pd.notna(b4_df[b4_columns]), "")
                        if b4_df.empty:
                            b4_df = pd.DataFrame([
                                {
                                    "Inquinante": "",
                                    "Rilasciato in": "Acqua",
                                    f"Emissioni [kg] - {prev_year}": "",
                                    f"Emissioni [kg] - {curr_year}": "",
                                }
                            ])
                        b4_df["Rilasciato in"] = b4_df["Rilasciato in"].apply(
                            lambda value: value if str(value).strip() in b4_release_options else "Acqua"
                        )

                        b4_column_config = {
                            "Inquinante": st.column_config.TextColumn(
                                "Inquinante",
                                width="medium",
                                help="Nome del contaminante rilasciato nell'ambiente"
                            ),
                            "Rilasciato in": st.column_config.SelectboxColumn(
                                "Rilasciato in",
                                options=b4_release_options,
                                required=True,
                                width="small",
                                help="Mezzo di rilascio (Aria, Acqua, Suolo)"
                            ),
                            f"Emissioni [kg] - {prev_year}": st.column_config.TextColumn(
                                f"Emissioni [kg] - {prev_year}",
                                width="small",
                                help=f"Quantità rilasciata in kg (anno {prev_year})"
                            ),
                            f"Emissioni [kg] - {curr_year}": st.column_config.TextColumn(
                                f"Emissioni [kg] - {curr_year}",
                                width="small",
                                help=f"Quantità rilasciata in kg (anno {curr_year})"
                            ),
                        }

                        edited_b4 = st.data_editor(
                            b4_df,
                            key=f"editor_{b4_key}",
                            column_config=b4_column_config,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                        )
                        st.session_state.vsme_disclosure_tables[b4_key] = edited_b4.where(pd.notna(edited_b4), "")

                        if st.button("Aggiungi inquinante", key="b4_add_row_btn"):
                            current_b4 = st.session_state.vsme_disclosure_tables[b4_key].copy()
                            new_row = pd.DataFrame([
                                {
                                    "Inquinante": "",
                                    "Rilasciato in": "Acqua",
                                    f"Emissioni [kg] - {prev_year}": "",
                                    f"Emissioni [kg] - {curr_year}": "",
                                }
                            ])
                            current_b4 = pd.concat([current_b4, new_row], ignore_index=True)
                            st.session_state.vsme_disclosure_tables[b4_key] = current_b4
                            st.rerun()

                        if pillar_code == "E" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B4")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=is_b4_environmental_monitoring_question,
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B4",
                                question_filter=is_b4_environmental_monitoring_question,
                            )

                    elif b5_is_custom:
                        compilation_year = datetime.now().year
                        prev_year = str(compilation_year - 2)
                        curr_year = str(compilation_year - 1)

                        # Prima tabella: righe per sede, colonne per metrica
                        b5_sites_key = "vsme_table_B5_sedi_biodiversita"
                        b5_metrics = [
                            "Coordinate (geolocalizzazione)",
                            "Area (ha)",
                            "Area protetta (biodiversity sensitive area)",
                            "Note",
                        ]

                        geo_df = st.session_state.vsme_disclosure_tables.get("vsme_table_B1_geolocalizzazione_siti")
                        site_pairs = []
                        if geo_df is not None and isinstance(geo_df, pd.DataFrame) and not geo_df.empty:
                            normalized_geo_df = geo_df.where(pd.notna(geo_df), "")
                            used_names = set()
                            fallback_idx = 1
                            for _, row in normalized_geo_df.iterrows():
                                site_name = str(row.get("Sito (di proprietà, in locazione o gestito)", "")).strip()
                                coordinates = str(row.get("Coordinate GPS", "")).strip()
                                if not site_name and not coordinates:
                                    continue
                                if not site_name:
                                    site_name = f"Sede {fallback_idx}"
                                base_name = site_name
                                suffix = 2
                                while site_name in used_names:
                                    site_name = f"{base_name} ({suffix})"
                                    suffix += 1
                                used_names.add(site_name)
                                site_pairs.append((site_name, coordinates))
                                fallback_idx += 1

                        if not site_pairs:
                            site_pairs = [("Sede 1", "")]

                        site_names = [name for name, _ in site_pairs]
                        coordinates_by_site = {name: coordinates for name, coordinates in site_pairs}

                        existing_b5_sites_df = st.session_state.vsme_disclosure_tables.get(b5_sites_key)
                        if existing_b5_sites_df is None or not isinstance(existing_b5_sites_df, pd.DataFrame) or existing_b5_sites_df.empty:
                            synced_b5_sites_df = pd.DataFrame({"Sedi": site_names})
                            for metric in b5_metrics:
                                synced_b5_sites_df[metric] = ""
                        else:
                            synced_b5_sites_df = existing_b5_sites_df.copy()
                            if "Sedi" not in synced_b5_sites_df.columns:
                                first_col = synced_b5_sites_df.columns[0]
                                synced_b5_sites_df = synced_b5_sites_df.rename(columns={first_col: "Sedi"})
                            for metric in b5_metrics:
                                if metric not in synced_b5_sites_df.columns:
                                    synced_b5_sites_df[metric] = ""
                            synced_b5_sites_df = synced_b5_sites_df[["Sedi"] + b5_metrics]

                        # Assicura che tutte le sedi da B1 siano presenti come righe
                        existing_sites = set(synced_b5_sites_df["Sedi"].tolist())
                        for site_name in site_names:
                            if site_name not in existing_sites:
                                new_row = pd.DataFrame({"Sedi": [site_name], **{metric: [""] for metric in b5_metrics}})
                                synced_b5_sites_df = pd.concat([synced_b5_sites_df, new_row], ignore_index=True)
                        
                        synced_b5_sites_df = synced_b5_sites_df.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})

                        # Sincronizza coordinate da B1
                        for idx, site_name in enumerate(synced_b5_sites_df["Sedi"]):
                            coord = coordinates_by_site.get(str(site_name).strip(), "")
                            synced_b5_sites_df.loc[idx, "Coordinate (geolocalizzazione)"] = coord

                        st.markdown("##### Siti in aree sensibili alla biodiversità")
                        
                        b5_column_config = {
                            "Sedi": st.column_config.TextColumn(
                                "Sedi",
                                disabled=True,
                                width="medium",
                                help="Nome della sede"
                            ),
                            "Coordinate (geolocalizzazione)": st.column_config.TextColumn(
                                "Coordinate (geolocalizzazione)",
                                width="medium",
                                help="Coordinate GPS della sede (sincronizzate da Geolocalizzazione B1)"
                            ),
                            "Area (ha)": st.column_config.TextColumn(
                                "Area (ha)",
                                width="small",
                                help="Area in ettari"
                            ),
                            "Area protetta (biodiversity sensitive area)": st.column_config.SelectboxColumn(
                                "Area protetta (biodiversity sensitive area)",
                                options=["", "Sì", "No"],
                                required=False,
                                width="medium",
                                help="Indicare se l'area è protetta"
                            ),
                            "Note": st.column_config.TextColumn(
                                "Note",
                                width="medium",
                                help="Note aggiuntive"
                            ),
                        }

                        edited_b5_sites = st.data_editor(
                            synced_b5_sites_df,
                            key=f"editor_{b5_sites_key}",
                            column_config=b5_column_config,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                            disabled=["Sedi"] + (["Coordinate (geolocalizzazione)"] if len(synced_b5_sites_df) > 0 and all(synced_b5_sites_df["Sedi"].iloc[i] in site_names for i in range(min(len(site_names), len(synced_b5_sites_df)))) else []),
                        )
                        
                        # Bonifica e valida i valori "Area protetta" - togli "None"
                        if "Area protetta (biodiversity sensitive area)" in edited_b5_sites.columns:
                            for idx in range(len(edited_b5_sites)):
                                raw_val = str(edited_b5_sites.loc[idx, "Area protetta (biodiversity sensitive area)"]).strip().lower()
                                if raw_val == "none":
                                    edited_b5_sites.loc[idx, "Area protetta (biodiversity sensitive area)"] = ""
                                elif raw_val in {"si", "sì", "yes", "y", "1", "true"}:
                                    edited_b5_sites.loc[idx, "Area protetta (biodiversity sensitive area)"] = "Sì"
                                elif raw_val in {"no", "n", "0", "false"}:
                                    edited_b5_sites.loc[idx, "Area protetta (biodiversity sensitive area)"] = "No"
                                elif raw_val not in {"", "sì", "si", "no"}:
                                    edited_b5_sites.loc[idx, "Area protetta (biodiversity sensitive area)"] = ""
                        
                        edited_b5_sites = edited_b5_sites.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})
                        st.session_state.vsme_disclosure_tables[b5_sites_key] = edited_b5_sites

                        if st.button("+ Aggiungi sede", key="b5_add_site_btn"):
                            new_site_name = f"Sede {len(edited_b5_sites) + 1}"
                            new_row = pd.DataFrame({"Sedi": [new_site_name], **{metric: [""] for metric in b5_metrics}})
                            updated_df = pd.concat([edited_b5_sites, new_row], ignore_index=True)
                            st.session_state.vsme_disclosure_tables[b5_sites_key] = updated_df
                            st.rerun()

                        # Seconda tabella: 4 metriche con colonne N-2, N-1 e variazione % auto-calcolata.
                        st.markdown("---")
                        b5_metrics_key = "vsme_table_B5_metriche_biodiversita"
                        b5_rows = [
                            "Uso del suolo totale [ha]",
                            "Totale superficie impermeabilizzata",
                            "Totale superficie nature-oriented on-site",
                            "Totale superficie nature-oriented off-site",
                        ]
                        first_site_name = site_names[0] if site_names else "Sede 1"
                        b5_label_col = first_site_name

                        existing_b5_metrics_df = st.session_state.vsme_disclosure_tables.get(b5_metrics_key)
                        if existing_b5_metrics_df is None or not isinstance(existing_b5_metrics_df, pd.DataFrame) or existing_b5_metrics_df.empty:
                            b5_metrics_df = pd.DataFrame({
                                b5_label_col: b5_rows,
                                prev_year: "",
                                curr_year: "",
                            })
                        else:
                            b5_metrics_df = existing_b5_metrics_df.copy().fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})
                            candidate_label_cols = [
                                col for col in b5_metrics_df.columns
                                if col not in {prev_year, curr_year, "Variazione %", first_site_name}
                            ]
                            if b5_label_col not in b5_metrics_df.columns:
                                if candidate_label_cols:
                                    b5_metrics_df = b5_metrics_df.rename(columns={candidate_label_cols[0]: b5_label_col})
                                else:
                                    b5_metrics_df[b5_label_col] = b5_rows
                            if prev_year not in b5_metrics_df.columns:
                                b5_metrics_df[prev_year] = ""
                            if curr_year not in b5_metrics_df.columns:
                                b5_metrics_df[curr_year] = ""
                            b5_metrics_df = b5_metrics_df[[b5_label_col, prev_year, curr_year]]
                            if b5_metrics_df.empty:
                                b5_metrics_df = pd.DataFrame({
                                    b5_label_col: b5_rows,
                                    prev_year: "",
                                    curr_year: "",
                                })

                        # Struttura righe fissa in ordine coerente con il template.
                        b5_metrics_df = b5_metrics_df.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})
                        b5_metrics_df = b5_metrics_df.set_index(b5_label_col, drop=False)
                        for row_label in b5_rows:
                            if row_label not in b5_metrics_df.index:
                                b5_metrics_df.loc[row_label, b5_label_col] = row_label
                                b5_metrics_df.loc[row_label, prev_year] = ""
                                b5_metrics_df.loc[row_label, curr_year] = ""
                        b5_metrics_df = b5_metrics_df.loc[b5_rows].reset_index(drop=True)
                        b5_metrics_df = b5_metrics_df.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})

                        b5_display_df = b5_metrics_df.copy()
                        b5_display_df["Variazione %"] = b5_display_df.apply(
                            lambda row: _compute_variation_percent(row.get(prev_year, ""), row.get(curr_year, "")),
                            axis=1,
                        )
                        b5_display_df = b5_display_df.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})

                        b5_column_config = {
                            b5_label_col: st.column_config.TextColumn(
                                b5_label_col,
                                disabled=True,
                                width="large",
                                help="Metrica di uso del suolo"
                            ),
                            prev_year: st.column_config.TextColumn(
                                prev_year,
                                width="medium",
                                help=f"Valore anno {prev_year}"
                            ),
                            curr_year: st.column_config.TextColumn(
                                curr_year,
                                width="medium",
                                help=f"Valore anno {curr_year}"
                            ),
                            "Variazione %": st.column_config.TextColumn(
                                "Variazione %",
                                disabled=True,
                                width="medium",
                                help="Variazione percentuale calcolata automaticamente"
                            ),
                        }

                        st.markdown("##### Uso del suolo")
                        edited_b5_metrics_df = st.data_editor(
                            b5_display_df[[b5_label_col, prev_year, curr_year, "Variazione %"]],
                            key=f"editor_{b5_metrics_key}_{b5_label_col}_{prev_year}_{curr_year}",
                            column_config=b5_column_config,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                            disabled=[b5_label_col, "Variazione %"],
                        )
                        edited_b5_metrics_df = edited_b5_metrics_df.fillna("").astype(str).replace({"None": "", "nan": "", "NaN": ""})
                        st.session_state.vsme_disclosure_tables[b5_metrics_key] = edited_b5_metrics_df[[b5_label_col, prev_year, curr_year]].copy()

                        if pillar_code == "E" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B5")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=is_b5_biodiversity_metrics_question,
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B5",
                                question_filter=is_b5_biodiversity_metrics_question,
                            )

                    elif b6_is_custom:
                        compilation_year = datetime.now().year
                        prev_year = str(compilation_year - 2)
                        curr_year = str(compilation_year - 1)

                        b6_key = "vsme_table_B6_uso_acqua"
                        B6_LABEL_COL = "Voce"
                        B6_UDM_COL = "UdM"
                        B6_STRESS_KEY = "b6_stress_idrico"
                        B6_PRELIEVO_ROWS = [
                            "Prelievo da acquedotto",
                            "Prelievo da pozzo",
                            "Prelievo acque superficiali (es. laghi, fiumi)",
                            "Prelievo di acqua da altre fonti (specificare)",
                        ]
                        B6_TOTALE_ROW = "Totale prelievo idrico"
                        B6_STRESS_ROW = "La zona di prelievo è soggetta a stress idrico?"
                        B6_STRESS_QTA_ROW = "Se sì, specificare la quantità di acqua prelevata in siti soggetti ad alto stress idrico"
                        B6_SCARICO_ROW = "Quantità di acqua scaricata"
                        B6_CONSUMO_ROW = "Consumo di acqua"
                        B6_ALL_ROWS = (
                            B6_PRELIEVO_ROWS
                            + [B6_TOTALE_ROW, B6_STRESS_ROW, B6_STRESS_QTA_ROW, B6_SCARICO_ROW, B6_CONSUMO_ROW]
                        )
                        B6_UDM_OPTIONS = [
                            "", "m³", "L", "kL", "ML", "Gl", "ft³", "gal (US)", "gal (UK)"
                        ]
                        B6_READONLY_ROWS = {B6_TOTALE_ROW, B6_CONSUMO_ROW}
                        B6_STRESS_ROW_OPTIONS = ["", "Sì", "No"]

                        # Inizializza o recupera df da session_state
                        if b6_key not in st.session_state.vsme_disclosure_tables:
                            st.session_state.vsme_disclosure_tables[b6_key] = pd.DataFrame([
                                {B6_LABEL_COL: row, B6_UDM_COL: "", prev_year: "", curr_year: ""}
                                for row in B6_ALL_ROWS
                            ])
                        else:
                            b6_existing = st.session_state.vsme_disclosure_tables[b6_key].copy()
                            for col in [B6_LABEL_COL, B6_UDM_COL, prev_year, curr_year]:
                                if col not in b6_existing.columns:
                                    b6_existing[col] = ""
                            existing_labels = b6_existing[B6_LABEL_COL].tolist()
                            for row_lbl in B6_ALL_ROWS:
                                if row_lbl not in existing_labels:
                                    b6_existing = pd.concat([
                                        b6_existing,
                                        pd.DataFrame([{B6_LABEL_COL: row_lbl, B6_UDM_COL: "", prev_year: "", curr_year: ""}])
                                    ], ignore_index=True)
                            b6_existing = b6_existing.where(pd.notna(b6_existing), "")
                            b6_existing = b6_existing.set_index(B6_LABEL_COL, drop=False)
                            b6_existing = b6_existing.loc[[r for r in B6_ALL_ROWS if r in b6_existing.index]].reset_index(drop=True)
                            st.session_state.vsme_disclosure_tables[b6_key] = b6_existing

                        b6_df = st.session_state.vsme_disclosure_tables[b6_key].copy()
                        b6_df = b6_df.where(pd.notna(b6_df), "")

                        # Helper: parse numero
                        def _b6_parse(val):
                            try:
                                raw = str(val or "").strip().replace(" ", "")
                                if not raw:
                                    return None
                                if "," in raw and "." in raw:
                                    raw = raw.replace(".", "").replace(",", ".")
                                elif "," in raw:
                                    raw = raw.replace(",", ".")
                                return float(raw)
                            except Exception:
                                return None

                        def _b6_sum_prelievo(df, year_col):
                            total = 0.0
                            any_val = False
                            for lbl in B6_PRELIEVO_ROWS:
                                rows = df[df[B6_LABEL_COL] == lbl]
                                if not rows.empty:
                                    v = _b6_parse(rows.iloc[0].get(year_col, ""))
                                    if v is not None:
                                        total += v
                                        any_val = True
                            return str(round(total, 6)) if any_val else "0"

                        def _b6_calc_consumo(df, year_col):
                            totale_val = _b6_parse(_b6_sum_prelievo(df, year_col))
                            scarico_rows = df[df[B6_LABEL_COL] == B6_SCARICO_ROW]
                            scarico_val = None
                            if not scarico_rows.empty:
                                scarico_val = _b6_parse(scarico_rows.iloc[0].get(year_col, ""))
                            if totale_val is None:
                                return "0"
                            if scarico_val is None:
                                return str(round(totale_val, 6))
                            return str(round(totale_val - scarico_val, 6))

                        # Aggiorna totale e consumo prima di mostrare
                        for yc in [prev_year, curr_year]:
                            totale_val = _b6_sum_prelievo(b6_df, yc)
                            consumo_val = _b6_calc_consumo(b6_df, yc)
                            b6_df.loc[b6_df[B6_LABEL_COL] == B6_TOTALE_ROW, yc] = totale_val
                            b6_df.loc[b6_df[B6_LABEL_COL] == B6_CONSUMO_ROW, yc] = consumo_val

                        # Gestione stress idrico: se No → riga sotto svuotata e disabilitata
                        stress_value = st.session_state.get(B6_STRESS_KEY, "")
                        stress_row_idx = b6_df[b6_df[B6_LABEL_COL] == B6_STRESS_ROW].index
                        if not stress_row_idx.empty:
                            b6_df.loc[stress_row_idx[0], prev_year] = stress_value
                            b6_df.loc[stress_row_idx[0], curr_year] = stress_value

                        # Righe non editabili: totale, consumo, stress idrico
                        b6_disabled_rows = B6_READONLY_ROWS | {B6_STRESS_ROW}
                        if stress_value != "Sì":
                            b6_disabled_rows.add(B6_STRESS_QTA_ROW)
                            b6_df.loc[b6_df[B6_LABEL_COL] == B6_STRESS_QTA_ROW, prev_year] = ""
                            b6_df.loc[b6_df[B6_LABEL_COL] == B6_STRESS_QTA_ROW, curr_year] = ""

                        st.markdown("##### Uso di acqua")

                        b6_col_config = {
                            B6_LABEL_COL: st.column_config.TextColumn(
                                "Voce",
                                disabled=True,
                                width="large",
                                help="Voce di prelievo o consumo idrico",
                            ),
                            B6_UDM_COL: st.column_config.SelectboxColumn(
                                "UdM",
                                options=B6_UDM_OPTIONS,
                                required=False,
                                width="small",
                                help="Unità di misura",
                            ),
                            prev_year: st.column_config.TextColumn(
                                prev_year,
                                width="small",
                                help=f"Valore anno {prev_year}",
                            ),
                            curr_year: st.column_config.TextColumn(
                                curr_year,
                                width="small",
                                help=f"Valore anno {curr_year}",
                            ),
                        }

                        edited_b6 = st.data_editor(
                            b6_df[[B6_LABEL_COL, B6_UDM_COL, prev_year, curr_year]],
                            key=f"editor_{b6_key}_{prev_year}_{curr_year}",
                            column_config=b6_col_config,
                            hide_index=True,
                            num_rows="fixed",
                            width='stretch',
                            disabled=[B6_LABEL_COL] + list(b6_disabled_rows - {B6_LABEL_COL}),
                        )

                        # Dropdown stress idrico separato (non modificabile in data_editor per riga)
                        st.markdown("**La zona di prelievo è soggetta a stress idrico?**")
                        stress_sel = st.selectbox(
                            "Stress idrico",
                            options=B6_STRESS_ROW_OPTIONS,
                            index=B6_STRESS_ROW_OPTIONS.index(stress_value) if stress_value in B6_STRESS_ROW_OPTIONS else 0,
                            key=B6_STRESS_KEY,
                            label_visibility="collapsed",
                        )

                        # Ricalcola totale e consumo dopo edit utente
                        edited_b6 = edited_b6.where(pd.notna(edited_b6), "")
                        for yc in [prev_year, curr_year]:
                            totale_val = _b6_sum_prelievo(edited_b6, yc)
                            consumo_val = _b6_calc_consumo(edited_b6, yc)
                            edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_TOTALE_ROW, yc] = totale_val
                            edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_CONSUMO_ROW, yc] = consumo_val

                        edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_STRESS_ROW, prev_year] = stress_sel
                        edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_STRESS_ROW, curr_year] = stress_sel
                        if stress_sel != "Sì":
                            edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_STRESS_QTA_ROW, prev_year] = ""
                            edited_b6.loc[edited_b6[B6_LABEL_COL] == B6_STRESS_QTA_ROW, curr_year] = ""

                        st.session_state.vsme_disclosure_tables[b6_key] = edited_b6

                        if pillar_code == "E" and questions_for_pillar is not None and gap_scale_options is not None:
                            st.markdown("---")
                            st.markdown("##### Documenti e informazioni - Disclosure B6")
                            render_gap_list(
                                questions_for_pillar,
                                pillar_code,
                                st.container(),
                                gap_scale_options,
                                gap_prefix,
                                question_filter=is_b6_water_question,
                                manage_cleanup=False,
                            )
                            render_disclosure_readiness(
                                questions_for_pillar,
                                pillar_code,
                                gap_prefix,
                                "Disclosure B6",
                                question_filter=is_b6_water_question,
                            )

                    if not b1_is_custom and not b2_is_custom and not b3_is_custom and not b4_is_custom and not b5_is_custom and not b6_is_custom:
                        if disclosure.get("tables"):
                            for table in disclosure["tables"]:
                                editor_storage_key = f"vsme_table_{pillar_code}_{selected_mode}_{disclosure['code']}_{table['sheet_name']}"
                                if editor_storage_key not in st.session_state.vsme_disclosure_tables:
                                    st.session_state.vsme_disclosure_tables[editor_storage_key] = table["data"].copy()

                                edited_df = st.data_editor(
                                    st.session_state.vsme_disclosure_tables[editor_storage_key],
                                    key=f"editor_{editor_storage_key}",
                                    width='stretch',
                                    hide_index=True,
                                    num_rows="dynamic",
                                )
                                st.session_state.vsme_disclosure_tables[editor_storage_key] = edited_df
                        else:
                            st.info("Nessuna tabella di raccolta dati trovata per questa disclosure nel file Excel caricato in workspace.")

            st.divider()

    if status_normativo != "VSME":
        st.info("La diagnosi VSME e disponibile solo con esito: Rendicontazione VSME.")
    else:
        module_labels = {
            "base": "Modulo Base",
            "comprehensive": "Modulo Completo",
            "absolute": "Analisi VSME Assoluta",
        }
        module_choice = st.radio(
            "Percorso di diagnosi VSME:",
            [module_labels["base"], module_labels["comprehensive"], module_labels["absolute"]],
            horizontal=True,
            key="vsme_module_choice",
        )
        selected_mode = next((k for k, v in module_labels.items() if v == module_choice), "base")

        checklist_data, vsme_scale_options, err = load_vsme_checklist_from_excel()
        disclosure_reference, disclosure_reference_err = load_vsme_disclosure_reference()
        if err:
            st.warning(f"WARNING: {err}. Uso le domande VSME predefinite.")
            checklist_data = None
            vsme_scale_options = VSME_SCALE_OPTIONS
        if disclosure_reference_err:
            st.warning(f"WARNING: impossibile leggere la raccolta dati VSME: {disclosure_reference_err}")

        # Se il caricamento è fallito o non ha dati, usa le domande predefinite.
        default_questions = {
            "GEN": ["Overview della Governance generale/sostenibilità presente?"],
            "E": [
                "Consumi Energetici suddivisi (rinnovabili vs fossili)?",
                "Emissioni GHG Scope 1 e Scope 2 misurate?",
                "Registro dei rifiuti aggiornato?",
            ],
            "S": [
                "Dati organico per genere, contratto e orario?",
                "Monitoraggio infortuni sul lavoro?",
                "Ore di formazione medie annue calcolate?",
            ],
            "G": ["Policy scritta su etica e diritti umani?", "Referente interno per la sostenibilità?"],
        }
        questions_by_module = {
            "base": {pillar: list(questions) for pillar, questions in default_questions.items()},
            "comprehensive": {pillar: list(questions) for pillar, questions in default_questions.items()},
        }

        if not checklist_data or not checklist_data.get('base'):
            pass
        else:
            questions_by_module = {
                "base": {
                    "GEN": checklist_data["base"].get("GEN", []),
                    "E": checklist_data["base"].get("E", []),
                    "S": checklist_data["base"].get("S", []),
                    "G": checklist_data["base"].get("G", []),
                },
                "comprehensive": {
                    "GEN": checklist_data["comprehensive"].get("GEN", []),
                    "E": checklist_data["comprehensive"].get("E", []),
                    "S": checklist_data["comprehensive"].get("S", []),
                    "G": checklist_data["comprehensive"].get("G", []),
                },
            }

        pillar_titles = {
            "GEN": "Informazioni generali",
            "E": "Ambiente",
            "S": "Sociale",
            "G": "Governance",
        }
        response_colors = {
            "Sì": "#2ecc71",
            "Sì, ma necessita integrazione": "#f39c12",
            "No, ma pianificato": "#3498db",
            "No": "#e74c3c",
        }
        default_palette = ["#2ecc71", "#f39c12", "#3498db", "#e74c3c", "#9b59b6", "#16a085"]
        for idx, option in enumerate(vsme_scale_options):
            response_colors.setdefault(to_vsme_plot_response_label(option), default_palette[idx % len(default_palette)])

        def build_vsme_results(module_payloads, scale_options):
            valid_keys = set()
            key_to_module = {}
            for payload in module_payloads:
                prefix = payload["prefix"]
                module_name = payload["module_name"]
                questions = payload["questions"]
                for pillar in ["GEN", "E", "S", "G"]:
                    if pillar in questions:
                        for i in range(len(questions[pillar])):
                            key = f"{prefix}_{pillar}_{i}"
                            valid_keys.add(key)
                            key_to_module[key] = module_name

            # "B2" è un pillar virtuale per le domande HSE AP / HSE Policies / Transition Plan
            checklist_label_map = {
                "GEN": "Disclosure B1",
                "B2": "Disclosure B2",
                "E": pillar_titles["E"],
                "S": pillar_titles["S"],
                "G": pillar_titles["G"],
            }
            all_pillar_keys = list(pillar_titles.keys()) + ["B2"]

            ordered_options = [label for label in VSME_PLOT_RESPONSE_ORDER if label in response_colors]
            pillar_totals = {p: {option: 0 for option in ordered_options} for p in all_pillar_keys}
            pillar_details = {p: [] for p in all_pillar_keys}
            for key, data in st.session_state.gap_answers.items():
                if key not in valid_keys:
                    continue
                pillar = data.get("pillar")
                answer = data.get("ans")
                answer_label = to_vsme_plot_response_label(answer)
                question = data.get("q", "")
                module_name = key_to_module.get(key, "")
                datapoints = get_question_datapoints(question)
                clean_question = clean_question_label(question)
                # Domande B2 vengono assegnate al pillar virtuale "B2" invece di "GEN"
                effective_pillar = "B2" if (pillar == "GEN" and is_b2_checklist_question(question)) else pillar
                checklist_label = checklist_label_map.get(effective_pillar, pillar_titles.get(effective_pillar, effective_pillar))
                if effective_pillar in pillar_totals and answer_label in pillar_totals[effective_pillar]:
                    pillar_totals[effective_pillar][answer_label] += datapoints
                if effective_pillar in pillar_details:
                    pillar_details[effective_pillar].append({
                        "Modulo": module_name,
                        "Checklist": checklist_label,
                        "Domanda": clean_question,
                        "Risposta": answer_label,
                        "Datapoints": datapoints,
                    })

            global_total_datapoints = sum(sum(pillar_totals[p].values()) for p in all_pillar_keys)
            response_totals = {option: 0 for option in ordered_options}
            for p in all_pillar_keys:
                for option in ordered_options:
                    response_totals[option] += pillar_totals[p][option]

            response_distribution_rows = []
            for option in ordered_options:
                datapoints = response_totals[option]
                pct_dataset = (datapoints / global_total_datapoints * 100) if global_total_datapoints else 0.0
                response_distribution_rows.append({
                    "Risposta": option,
                    "Datapoints": datapoints,
                    "% sul totale dataset": pct_dataset,
                })

            # Titoli da usare nelle intestazioni dei grafici per-pillar
            pillar_display_title = {
                "GEN": "Informazioni generali – Disclosure B1",
                "B2": "Informazioni generali – Disclosure B2",
                "E": pillar_titles["E"],
                "S": pillar_titles["S"],
                "G": pillar_titles["G"],
            }

            per_pillar_summary = {}
            per_pillar_details = {}
            stacked_rows = []
            all_rows = []
            for p in all_pillar_keys:
                title = pillar_display_title.get(p, p)
                total_datapoints = sum(pillar_totals[p].values())
                summary_rows = []
                detail_rows = []
                for option in ordered_options:
                    datapoints = pillar_totals[p][option]
                    pct = (datapoints / total_datapoints * 100) if total_datapoints else 0.0
                    global_pct = (datapoints / global_total_datapoints * 100) if global_total_datapoints else 0.0
                    summary_row = {
                        "Checklist": title,
                        "Risposta": option,
                        "Datapoints": datapoints,
                        "% sul totale checklist": pct,
                        "% sul totale dataset": global_pct,
                    }
                    summary_rows.append(summary_row)
                    stacked_rows.append(summary_row)
                for row in pillar_details[p]:
                    detail_row = row.copy()
                    detail_row["% sul totale checklist"] = (
                        detail_row["Datapoints"] / total_datapoints * 100 if total_datapoints else 0.0
                    )
                    detail_row["% sul totale dataset"] = (
                        detail_row["Datapoints"] / global_total_datapoints * 100 if global_total_datapoints else 0.0
                    )
                    detail_rows.append(detail_row)
                    all_rows.append(detail_row)
                df_summary = pd.DataFrame(summary_rows)
                df_details = pd.DataFrame(detail_rows)
                per_pillar_summary[p] = df_summary
                per_pillar_details[p] = df_details

            df_stacked = pd.DataFrame(stacked_rows)
            df_all_details = pd.DataFrame(all_rows)
            df_response_distribution = pd.DataFrame(response_distribution_rows)

            return {
                "summary": per_pillar_summary,
                "details": per_pillar_details,
                "stacked": df_stacked,
                "all_details": df_all_details,
                "response_distribution": df_response_distribution,
                "pillar_display_title": pillar_display_title,
            }

        tab_summary = None
        if selected_mode in ["base", "comprehensive"]:
            module_questions = questions_by_module[selected_mode]
            module_prefix = f"vsme_{selected_mode}"

            st.markdown(
                """
                <style>
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"] {
                    font-size: 1rem;
                    font-weight: 600;
                    border-radius: 8px 8px 0 0;
                    padding: 0.5rem 0.95rem;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(1) {
                    color: #1f4e79;
                    background: #eaf3ff;
                    border: 1px solid #bfd7f2;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(1) {
                    color: #0f3554;
                    background: #d6e9ff;
                    border-bottom-color: #d6e9ff;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(2) {
                    color: #2f6b3b;
                    background: #e7f4ea;
                    border: 1px solid #b9d8c0;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(2) {
                    color: #214d2a;
                    background: #d7eddc;
                    border-bottom-color: #d7eddc;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(3) {
                    color: #a54b3f;
                    background: #fde8e4;
                    border: 1px solid #f2c4ba;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(3) {
                    color: #7d362d;
                    background: #f9d7cf;
                    border-bottom-color: #f9d7cf;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(4) {
                    color: #8a6a00;
                    background: #fff4d6;
                    border: 1px solid #ecd89b;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(4) {
                    color: #6c5200;
                    background: #f8e7b0;
                    border-bottom-color: #f8e7b0;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(5) {
                    color: #5f4b8b;
                    background: #efe7ff;
                    border: 1px solid #d6c7f3;
                }
                div[data-testid="stVerticalBlock"] > div:has(.vsme-main-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(5) {
                    color: #49366f;
                    background: #e2d6fb;
                    border-bottom-color: #e2d6fb;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            st.markdown("<div class='vsme-main-tabs-marker'></div>", unsafe_allow_html=True)
            tab_gen, c_v_E, c_v_S, c_v_G, tab_summary = st.tabs(["Informazioni generali", "Ambiente", "Sociale", "Governance", "Sintesi diagnosi"])
            if module_questions["GEN"]:
                render_vsme_disclosure_tables(
                    "GEN",
                    selected_mode,
                    tab_gen,
                    disclosure_reference,
                    questions_for_pillar=module_questions["GEN"],
                    gap_scale_options=vsme_scale_options,
                    gap_prefix=module_prefix,
                )
            else:
                render_vsme_disclosure_tables("GEN", selected_mode, tab_gen, disclosure_reference)
            render_vsme_disclosure_tables(
                "E",
                selected_mode,
                c_v_E,
                disclosure_reference,
                questions_for_pillar=module_questions["E"],
                gap_scale_options=vsme_scale_options,
                gap_prefix=module_prefix,
            )
            render_vsme_disclosure_tables("S", selected_mode, c_v_S, disclosure_reference)
            render_gap_list(module_questions["S"], "S", c_v_S, vsme_scale_options, module_prefix)
            render_vsme_disclosure_tables("G", selected_mode, c_v_G, disclosure_reference)
            render_gap_list(module_questions["G"], "G", c_v_G, vsme_scale_options, module_prefix)
        if selected_mode == "absolute":
            module_payloads = [
                {
                    "module_name": module_labels["base"],
                    "prefix": "vsme_base",
                    "questions": questions_by_module["base"],
                },
                {
                    "module_name": module_labels["comprehensive"],
                    "prefix": "vsme_comprehensive",
                    "questions": questions_by_module["comprehensive"],
                },
            ]
        else:
            module_payloads = [
                {
                    "module_name": module_labels[selected_mode],
                    "prefix": f"vsme_{selected_mode}",
                    "questions": questions_by_module[selected_mode],
                }
            ]

        results = build_vsme_results(module_payloads, vsme_scale_options)

        target_container = tab_summary if tab_summary is not None else st.container()
        with target_container:
            st.divider()
            module_prefixes = [payload["prefix"] for payload in module_payloads]
            gap_summary = summarize_gap_answers(module_prefixes)
            completeness_label = f"{gap_summary['coverage']:.0f}%"
            if gap_summary["coverage"] >= 80 and gap_summary["missing_datapoints"] == 0:
                readiness_label = "Alta"
            elif gap_summary["coverage"] >= 50:
                readiness_label = "Media"
            else:
                readiness_label = "Bassa"

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Esito", status_normativo)
            m2.metric("Percorso", module_labels[selected_mode])
            m3.metric("Preparazione", readiness_label)
            m4.metric("Completezza dati", completeness_label)

            c_diag1, c_diag2 = st.columns(2)
            with c_diag1:
                st.markdown("### Cosa si applica")
                st.write(f"Esito automatico: {status_normativo}")
                st.write(f"Domande valutate: {gap_summary['total']}")
                st.write(f"Risposte pronte o parziali: {gap_summary['ready'] + gap_summary['partial']}")
            with c_diag2:
                st.markdown("### Cosa ti manca")
                if gap_summary["top_gaps"]:
                    for idx, question in enumerate(gap_summary["top_gaps"], start=1):
                        st.write(f"{idx}. {question}")
                else:
                    st.write("Nessun gap prioritario rilevato al momento.")

            st.markdown("### Prossime azioni consigliate")
            for idx, action in enumerate(build_priority_actions(), start=1):
                st.write(f"{idx}. {action}")

            if selected_mode == "absolute":
                st.subheader("📊 Vista aggregata VSME")
                st.write("Confronto stacked per checklist aggregando Modulo Base e Modulo Completo.")
            else:
                st.subheader(f"📊 Riepilogo diagnosi - {module_labels[selected_mode]}")

            df_distribution = results["response_distribution"].copy()
            fig_distribution = px.bar(
                df_distribution,
                x="Risposta",
                y="% sul totale dataset",
                color="Risposta",
                text=df_distribution["% sul totale dataset"].apply(lambda x: f"{x:.2f}%"),
                color_discrete_map=response_colors,
                category_orders={"Risposta": VSME_PLOT_RESPONSE_ORDER},
            )
            fig_distribution.update_layout(
                title="Percentuale sul totale dataset",
                yaxis_title="Percentuale sul totale dataset",
                xaxis_title="Risposta",
                legend_title="Risposta",
                showlegend=False,
            )
            st.plotly_chart(fig_distribution, width='stretch', key=f"vsme_distribution_{selected_mode}")

            table_distribution = df_distribution.copy()
            if len(table_distribution) > 0:
                raw_percentages = table_distribution["% sul totale dataset"].fillna(0).astype(float).tolist()
                rounded_percentages = [round(value, 2) for value in raw_percentages]
                total_datapoints_distribution = float(table_distribution["Datapoints"].fillna(0).sum())

                # Con datapoints presenti, forza la somma visibile a 100% correggendo il delta di arrotondamento.
                if total_datapoints_distribution > 0:
                    rounding_delta = round(100.0 - sum(rounded_percentages), 2)
                    if abs(rounding_delta) >= 0.01:
                        pivot_idx = max(range(len(raw_percentages)), key=lambda idx: raw_percentages[idx])
                        rounded_percentages[pivot_idx] = round(rounded_percentages[pivot_idx] + rounding_delta, 2)

                table_distribution["% sul totale dataset"] = [f"{max(value, 0.0):.2f}%" for value in rounded_percentages]

            st.dataframe(table_distribution, width='stretch', hide_index=True)

            st.subheader("📋 Checklist completa")
            all_details_table = results["all_details"].copy()
            if len(all_details_table) > 0:
                all_details_table["% sul totale checklist"] = all_details_table["% sul totale checklist"].apply(lambda x: f"{x:.2f}%")
                details_columns_order = [
                    "Modulo",
                    "Checklist",
                    "Domanda",
                    "Datapoints",
                    "% sul totale checklist",
                    "Risposta",
                ]
                existing_order = [col for col in details_columns_order if col in all_details_table.columns]
                remaining_cols = [col for col in all_details_table.columns if col not in existing_order]
                remaining_cols = [col for col in remaining_cols if col != "% sul totale dataset"]
                all_details_table = all_details_table[existing_order + remaining_cols]
            st.dataframe(all_details_table, width='stretch', hide_index=True)

# =====================================================================
# TAB 2: ANALISI RISCHI (MAPPA FISICA, IPCC E NGFS)
# =====================================================================
with t_rischi:
    st.caption("Percorso 2 di 3")
    st.subheader("Piano di azione & rischi")
    st.write("Qui trasformi la diagnosi in un piano operativo, aggiungi le sedi rilevanti e valuti il profilo di rischio climatico e di transizione.")
    st.markdown(
        """
        <style>
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"] {
            font-size: 1rem;
            font-weight: 600;
            border-radius: 8px 8px 0 0;
            padding: 0.5rem 0.95rem;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(1) {
            color: #1f4e79;
            background: #eaf3ff;
            border: 1px solid #bfd7f2;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(1) {
            color: #0f3554;
            background: #d6e9ff;
            border-bottom-color: #d6e9ff;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(2) {
            color: #2f6b3b;
            background: #e7f4ea;
            border: 1px solid #b9d8c0;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(2) {
            color: #214d2a;
            background: #d7eddc;
            border-bottom-color: #d7eddc;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"]:nth-child(3) {
            color: #8a6a00;
            background: #fff4d6;
            border: 1px solid #ecd89b;
        }
        div[data-testid="stVerticalBlock"] > div:has(.risk-tabs-marker) + div div[data-baseweb="tab-list"] button[role="tab"][aria-selected="true"]:nth-child(3) {
            color: #6c5200;
            background: #f8e7b0;
            border-bottom-color: #f8e7b0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<div class='risk-tabs-marker'></div>", unsafe_allow_html=True)
    rt_fisico, rt_transizione, rt_credito = st.tabs(["🗺️ Piano & Mappa Asset", "🔄 Emissioni & Transizione", "💰 Stress Test Finanziario"])
    
    with rt_fisico:
        def ensure_hq_asset_on_map():
            hq_address = st.session_state.get("hq_address", "").strip()
            current_assets = st.session_state.portfolio_df.drop(columns=["Display_Size"], errors="ignore")
            if not current_assets.empty and "ID" in current_assets.columns:
                current_assets = current_assets[current_assets["ID"] != "HQ-SEDE"]

            if not hq_address:
                st.session_state.portfolio_df = process_portfolio_dataframe(current_assets) if not current_assets.empty else pd.DataFrame()
                return

            manual_hq_lat = st.session_state.get("hq_lat")
            manual_hq_lon = st.session_state.get("hq_lon")
            if manual_hq_lat is not None and manual_hq_lon is not None:
                try:
                    st.session_state.hq_geocoded_address = hq_address or "Coordinate manuali sede"
                    st.session_state.hq_lat = float(manual_hq_lat)
                    st.session_state.hq_lon = float(manual_hq_lon)
                except (TypeError, ValueError):
                    return
            elif OFFLINE_MODE:
                return
            elif st.session_state.get("hq_geocoded_address") != hq_address or st.session_state.get("hq_lat") is None or st.session_state.get("hq_lon") is None:
                try:
                    geolocator = Nominatim(user_agent="smes-reporting-support-hq")
                    location = geolocator.geocode(hq_address, timeout=10)
                    if not location:
                        return
                    st.session_state.hq_geocoded_address = hq_address
                    st.session_state.hq_lat = float(location.latitude)
                    st.session_state.hq_lon = float(location.longitude)
                except Exception:
                    return

            hq_row = pd.DataFrame([{
                "Name": st.session_state.get("company_name", "Sede Legale") or "Sede Legale",
                "Address": hq_address,
                "Lat": float(st.session_state.hq_lat),
                "Lon": float(st.session_state.hq_lon),
                "Operator": "SEDE",
                "ID": "HQ-SEDE",
                "Size": 7000,
            }])

            combined_assets = pd.concat([current_assets, hq_row], ignore_index=True)
            st.session_state.portfolio_df = process_portfolio_dataframe(combined_assets)
        
        if st.session_state.portfolio_df.empty and os.path.exists("centrali_operative_esatte.csv"):
            try:
                df_map = pd.read_csv("centrali_operative_esatte.csv")
                st.session_state.portfolio_df = process_portfolio_dataframe(df_map)
            except: pass
            
        def load_portfolio_file(uploaded_file):
            """Carica file CSV, XLSX o XLS e ritorna un DataFrame."""
            try:
                if uploaded_file.name.endswith('.csv'):
                    return pd.read_csv(uploaded_file)
                elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                    return pd.read_excel(uploaded_file)
                else:
                    st.error(f"Formato file non supportato: {uploaded_file.name}. Usa CSV, XLSX o XLS.")
                    return pd.DataFrame()
            except Exception as e:
                st.error(f"Errore nel caricamento del file: {str(e)}")
                return pd.DataFrame()
            
        ensure_hq_asset_on_map()

        if not st.session_state.portfolio_df.empty:
            df_render = st.session_state.portfolio_df.copy()

            np.random.seed(42)
            base_risk = np.clip(100 - (df_render['Lat'] - 35) * 6 + np.random.randint(-10, 15, size=len(df_render)), 10, 100)

            st.subheader(
                "🌍 Selezione Scenario Climatico (IPCC AR6)",
                help="I 3 scenari (Shared Socioeconomic Pathways) elaborati dalle Nazioni Unite. Più lo scenario è fossile (SSP5), più il rischio fisico calcolato per l'impianto aumenta."
            )
            ipcc_scenario_mappa = st.radio(
                "Scenario IPCC",
                ["SSP1-2.6 (+1.5°C)", "SSP2-4.5 (+2.4°C)", "SSP5-8.5 (+4.0°C)"],
                horizontal=True,
                label_visibility="collapsed"
            )

            if "SSP1" in ipcc_scenario_mappa: risk_multiplier = 0.7
            elif "SSP2" in ipcc_scenario_mappa: risk_multiplier = 1.1
            else: risk_multiplier = 1.6
            df_render['Risk_Score'] = np.clip(base_risk * risk_multiplier, 10, 100).astype(int)

            top_risk_row = df_render.sort_values("Risk_Score", ascending=False).iloc[0]
            top_risk_name = top_risk_row.get("Name", "Asset")
            top_risk_score = int(top_risk_row.get("Risk_Score", 0))

            st.markdown("### Azioni prioritarie")
            for idx, action in enumerate(build_priority_actions(top_risk_score, top_risk_name), start=1):
                st.write(f"{idx}. {action}")

            r1, r2, r3 = st.columns(3)
            r1.metric("Asset mappati", len(df_render))
            r2.metric("Asset piu esposto", top_risk_name)
            r3.metric("Rischio massimo", f"{top_risk_score}/100")

            fig_portfolio = px.scatter_mapbox(
                df_render, lat="Lat", lon="Lon", hover_name="Name",
                hover_data={"Lat": False, "Lon": False, "ID": True, "Operator": True, "Address": True, "Risk_Score": True, "Size": True, "Display_Size": False},
                color="Risk_Score", size="Display_Size", color_continuous_scale=px.colors.diverging.RdYlGn_r,
                range_color=[10, 100], size_max=25, zoom=4.5 if len(df_render) > 1 else 10, mapbox_style="carto-positron", height=600
            )
            fig_portfolio.update_layout(margin={"r":0,"t":0,"l":0,"b":0})
            st.plotly_chart(fig_portfolio, use_container_width=True)
        with st.expander("🏠 Aggiungi immobile da indirizzo", expanded=False):
            with st.form("manual_asset_address_form"):
                c_addr1, c_addr2 = st.columns([2, 1])
                with c_addr1:
                    manual_address = st.text_input(
                        "Indirizzo immobile",
                        key="manual_asset_address",
                        placeholder="Es. Via Roma 1, Milano, Italia"
                    )
                with c_addr2:
                    manual_asset_name = st.text_input(
                        "Nome immobile (opzionale)",
                        key="manual_asset_name",
                        placeholder="Es. Sede Milano"
                    )

                submit_manual_address = st.form_submit_button("Aggiungi immobile in mappa")

            if submit_manual_address:
                if not manual_address or not manual_address.strip():
                    st.warning("Inserisci un indirizzo valido.")
                elif OFFLINE_MODE:
                    st.warning("La geolocalizzazione da indirizzo richiede internet. In offline usa il form GPS o carica un file con colonne Lat e Lon.")
                else:
                    try:
                        geolocator = Nominatim(user_agent="smes-reporting-support-map")
                        location = geolocator.geocode(manual_address.strip(), timeout=10)
                        if not location:
                            st.error("Indirizzo non trovato. Prova a inserire via, numero civico, citta e paese.")
                        else:
                            new_asset = pd.DataFrame([{
                                "Name": manual_asset_name.strip() if manual_asset_name and manual_asset_name.strip() else "Immobile",
                                "Address": manual_address.strip(),
                                "Lat": float(location.latitude),
                                "Lon": float(location.longitude),
                                "Operator": "IMMOBILE",
                                "ID": f"IMM-{int(time.time())}",
                                "Size": 5000,
                            }])

                            current_assets = st.session_state.portfolio_df.drop(columns=["Display_Size"], errors="ignore")
                            combined_assets = pd.concat([current_assets, new_asset], ignore_index=True)
                            st.session_state.portfolio_df = process_portfolio_dataframe(combined_assets)
                            st.success("Immobile aggiunto. Il punto e ora incluso nella mappa rischi IPCC.")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Errore durante la geolocalizzazione: {e}")

        with st.expander("📍 Aggiungi immobile da coordinate GPS", expanded=False):
            with st.form("manual_asset_gps_form"):
                c_gps1, c_gps2, c_gps3 = st.columns([1, 1, 1])
                with c_gps1:
                    gps_lat = st.number_input("Latitudine", key="manual_asset_lat", format="%.6f")
                with c_gps2:
                    gps_lon = st.number_input("Longitudine", key="manual_asset_lon", format="%.6f")
                with c_gps3:
                    gps_name = st.text_input("Nome immobile", key="manual_asset_gps_name", placeholder="Es. Sede GPS")

                submit_manual_gps = st.form_submit_button("Aggiungi coordinate in mappa")

            if submit_manual_gps:
                if gps_lat == 0.0 and gps_lon == 0.0:
                    st.warning("Inserisci coordinate GPS valide.")
                else:
                    gps_asset = pd.DataFrame([{
                        "Name": gps_name.strip() if gps_name and gps_name.strip() else "Immobile GPS",
                        "Address": f"GPS: {gps_lat:.6f}, {gps_lon:.6f}",
                        "Lat": float(gps_lat),
                        "Lon": float(gps_lon),
                        "Operator": "IMMOBILE",
                        "ID": f"GPS-{int(time.time())}",
                        "Size": 5000,
                    }])
                    current_assets = st.session_state.portfolio_df.drop(columns=["Display_Size"], errors="ignore")
                    combined_assets = pd.concat([current_assets, gps_asset], ignore_index=True)
                    st.session_state.portfolio_df = process_portfolio_dataframe(combined_assets)
                    st.success("Immobile GPS aggiunto alla mappa rischi IPCC.")
                    st.rerun()

        with st.expander("🔄 Carica un portfolio impianti diverso (Upload file)"):
            uploaded_portfolio = st.file_uploader("Carica File (CSV, XLSX, XLS)", type=['csv', 'xlsx', 'xls'], help="Il file deve contenere colonne chiamate Lat e Lon.")

            if st.button("Genera Mappa da nuovo file"):
                df_map = pd.DataFrame()
                if uploaded_portfolio:
                    df_map = load_portfolio_file(uploaded_portfolio)
                
                if not df_map.empty:
                    st.session_state.portfolio_df = process_portfolio_dataframe(df_map)

        if not st.session_state.portfolio_df.empty:
            df_render_table = st.session_state.portfolio_df.copy()
            with st.expander("Mostra dati tabellari"):
                table_columns = ["ID", "Name", "Operator", "Address", "Lat", "Lon", "Risk_Score"]
                available_columns = [column for column in table_columns if column in df_render_table.columns]
                st.dataframe(df_render_table[available_columns], use_container_width=True)

    with rt_transizione:
        st.subheader("Calcolatore GHG (Fattori ISPRA/DEFRA)", help="Incrocia i dati di consumo grezzi (es. Metri cubi) con i 'Fattori di Emissione' approvati a livello ministeriale per dedurre in automatico l'impronta di carbonio (tCO2).")
        
        EMISSION_FACTORS = {
            "Scope 1 - Gas Naturale": {"scope": "scope1", "unita": {"Metri Cubi (Sm3)": 1.98}},
            "Scope 1 - Gasolio": {"scope": "scope1", "unita": {"Litri (L)": 2.68}},
            "Scope 2 - Elettricità (Mix Italia)": {"scope": "scope2", "unita": {"MWh": 259.0}},
            "Scope 3 - Trasporto Merci": {"scope": "scope3", "unita": {"Tonnellate-km": 0.11}}
        }
        
        c_calc1, c_calc2, c_calc3 = st.columns([2, 1, 1])
        fonte_sel = c_calc1.selectbox("Categoria Consumo", list(EMISSION_FACTORS.keys()), help="Suddivisione secondo lo standard globale 'GHG Protocol'.")
        unita_sel = c_calc2.selectbox("Unità", list(EMISSION_FACTORS[fonte_sel]["unita"].keys()))
        fattore = EMISSION_FACTORS[fonte_sel]["unita"][unita_sel]
        scope_target = EMISSION_FACTORS[fonte_sel]["scope"]
        consumo = c_calc1.number_input("Volume Annuo", min_value=0.0, step=100.0)
        co2_calc = (consumo * fattore) / 1000 
        
        c_calc3.metric("Emissioni", f"{co2_calc:,.2f} tCO2", help=f"Fattore: {fattore} kgCO2 per unità. Diviso 1000 per avere le tonnellate.")
        if st.button(f"➕ Somma allo {scope_target.capitalize()}"):
            if co2_calc > 0:
                st.session_state[scope_target] += int(co2_calc)
                sync_from_scopes()
                st.success("Aggiunte!")
                time.sleep(1)
                st.rerun()

        st.divider()
        c_ghg1, c_ghg2 = st.columns(2)
        with c_ghg1:
            st.number_input("Scope 1 (tCO2)", value=st.session_state.scope1, step=500, key='scope1', on_change=sync_from_scopes, help="Emissioni dirette da caldaie, forni, e auto aziendali.")
            st.number_input("Scope 2 (tCO2)", value=st.session_state.scope2, step=500, key='scope2', on_change=sync_from_scopes, help="Emissioni indirette dalla corrente elettrica acquistata.")
            st.number_input("Scope 3 (tCO2)", value=st.session_state.scope3, step=500, key='scope3', on_change=sync_from_scopes, help="Emissioni catena del valore (fornitori, logistica, uso dei prodotti).")
        with c_ghg2:
            st.info(f"### Impronta Lorda\n# {get_tot_emissions():,} tCO2")
            st.slider("Efficacia Decarbonizzazione (%)", 0, 100, key='perc_red', on_change=sync_from_perc, help="Simula una politica di riduzione (es. acquisto impianti green o quote compensative). Genera le Emissioni Nette, usate negli Stress Test finanziari.")
            st.success(f"**Emissioni Nette:** {st.session_state.em_final:,} tCO2")

    with rt_credito:
        st.subheader("Stress Test Finanziario (Scenari NGFS)", help="Interroga il database Network for Greening the Financial System (NGFS) che traccia il prezzo ombra del carbonio fino al 2050 diviso per paese e scenario.")
        c_cred1, c_cred2 = st.columns(2)
        with c_cred1: st.number_input("Rata Prestito Transizione (€)", value=st.session_state.rata_prestito, step=100000, key='rata_prestito', help="Sottrae un costo annuo fisso dal calcolo dell'EBITDA.")
        with c_cred2: st.slider("Severità Policy", 1.0, 3.0, value=st.session_state.policy_multiplier, step=0.1, key='policy_multiplier', help="Moltiplica il prezzo base della Carbon Tax per simulare scenari normativi regionali estremamente restrittivi.")

        country_data = df_base[df_base['Paese'] == st.session_state.selected_country].copy()
        plot_data = []
        for _, row in country_data.iterrows():
            eff_price = row['Prezzo Carbonio Base'] * st.session_state.policy_multiplier
            profit = st.session_state.revenue - st.session_state.opex - (eff_price * st.session_state.em_final) - st.session_state.rata_prestito
            plot_data.append({"Anno": row['Anno'], "Utile Netto (€)": profit, "Scenario": row['Scenario']})
        
        st.plotly_chart(px.line(pd.DataFrame(plot_data), x="Anno", y="Utile Netto (€)", color="Scenario", title="EBITDA Netto post-Carbon Tax"), use_container_width=True)

# =====================================================================
# TAB 5: DOWNLOAD
# =====================================================================
with t_down:
    st.caption("Percorso 3 di 3")
    st.subheader("Deliverable finale")
    st.write("Questa sezione prepara un output leggibile per management, consulenza operativa o uso tecnico interno.")

    report_type = st.radio(
        "Formato deliverable",
        ["Executive", "Operativo", "Tecnico"],
        horizontal=True,
        key="deliverable_type",
    )

    export_gap_summary = summarize_gap_answers(["vsme_base", "vsme_comprehensive"])
    export_actions = build_priority_actions()
    asset_count = 0 if st.session_state.portfolio_df.empty else len(st.session_state.portfolio_df)

    st.markdown("### Anteprima deliverable")
    preview_lines = []
    if report_type == "Executive":
        preview_lines = [
            f"Esito azienda: {st.session_state.get('status_normativo', 'VSME')}",
            f"Completezza dati stimata: {export_gap_summary['coverage']:.0f}%",
            f"Gap prioritari aperti: {export_gap_summary['missing']}",
            f"Asset o sedi considerate: {asset_count}",
        ]
    elif report_type == "Operativo":
        preview_lines = [
            f"Checklist valutate: {export_gap_summary['total']}",
            f"Risposte pronte: {export_gap_summary['ready']}",
            f"Risposte parziali: {export_gap_summary['partial']}",
            f"Risposte assenti: {export_gap_summary['missing']}",
        ]
    else:
        preview_lines = [
            f"Sedi o asset geolocalizzati: {asset_count}",
            f"Emissioni lorde stimate: {get_tot_emissions():,} tCO2",
            f"Emissioni nette stimate: {st.session_state.em_final:,} tCO2",
            f"Scenario policy multiplier: {st.session_state.policy_multiplier:.1f}x",
        ]

    for idx, line in enumerate(preview_lines, start=1):
        st.write(f"{idx}. {line}")

    st.markdown("### Azioni incluse nel report")
    for idx, action in enumerate(export_actions, start=1):
        st.write(f"{idx}. {action}")

    if st.button("🪄 Genera deliverable PDF", help="Compila un report PDF sintetico a partire dai dati gia inseriti nell'app."):
        report_titles = {
            "Executive": "Executive ESG Brief",
            "Operativo": "Operational ESG Action Plan",
            "Tecnico": "Technical ESG Output",
        }
        report_filenames = {
            "Executive": "Deliverable_Executive.pdf",
            "Operativo": "Deliverable_Operativo.pdf",
            "Tecnico": "Deliverable_Tecnico.pdf",
        }

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 18)
        pdf.cell(200, 15, txt=report_titles[report_type], ln=True, align='C')
        pdf.set_font("Arial", size=11)
        pdf.ln(4)
        pdf.multi_cell(0, 8, txt=f"Esito: {st.session_state.get('status_normativo', 'VSME')}")
        pdf.multi_cell(0, 8, txt=f"Completezza dati: {export_gap_summary['coverage']:.0f}%")
        pdf.multi_cell(0, 8, txt=f"Asset o sedi considerate: {asset_count}")
        pdf.ln(2)
        for idx, action in enumerate(export_actions, start=1):
            safe_action = action.encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 8, txt=f"{idx}. {safe_action}")

        st.download_button("Scarica PDF", pdf.output(dest='S').encode('latin-1'), report_filenames[report_type])

project_name_for_autosave = (
    st.session_state.get('project_name_input', '').strip()
    or st.session_state.get('current_project_name', '').strip()
    or st.session_state.get('company_name', '').strip()
)
if project_name_for_autosave:
    autosave_project_if_needed(project_name_for_autosave)
