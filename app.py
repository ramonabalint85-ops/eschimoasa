import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import yfinance as yf
from fpdf import FPDF
import time
import numpy as np
import PyPDF2
import json
from openai import OpenAI
from geopy.geocoders import Nominatim
import os
import re
import requests
from datetime import datetime, timedelta
from functools import lru_cache
import shutil

# --- CONFIGURAZIONE PAGINA ---
st.set_page_config(page_title="CarbonRisk AI Enterprise", layout="wide")

# --- CACHE PERSISTENTE E THROTTLING (Anti-Rate-Limiting) ---
CACHE_DIR = ".yfinance_cache"
CACHE_TIMEOUT_HOURS = 24
THROTTLE_DELAY = 2

os.makedirs(CACHE_DIR, exist_ok=True)

def get_cache_file(ticker):
    """Restituisce il percorso del file cache per un ticker"""
    return os.path.join(CACHE_DIR, f"{ticker.upper()}_cache.json")

def is_cache_valid(ticker):
    """Controlla se la cache è ancora valida"""
    cache_file = get_cache_file(ticker)
    if not os.path.exists(cache_file):
        return False
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
            timestamp = datetime.fromisoformat(cache_data.get('timestamp', '2020-01-01'))
            age_hours = (datetime.now() - datetime.fromisoformat(cache_data.get('timestamp'))).total_seconds() / 3600
            return age_hours < CACHE_TIMEOUT_HOURS
    except:
        return False

def load_from_cache(ticker):
    """Carica dati dalla cache locale"""
    cache_file = get_cache_file(ticker)
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)['data']
    except:
        return None

def save_to_cache(ticker, data):
    """Salva dati nella cache locale"""
    cache_file = get_cache_file(ticker)
    cache_data = {
        'ticker': ticker.upper(),
        'timestamp': datetime.now().isoformat(),
        'data': data
    }
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
    except:
        pass

def retry_with_backoff(func, ticker, max_retries=3):
    """Riprova con exponential backoff"""
    for attempt in range(max_retries):
        try:
            time.sleep(THROTTLE_DELAY)
            result = func(ticker)
            if result:
                return result
        except requests.exceptions.ConnectionError:
            wait_time = (2 ** attempt)
            if attempt < max_retries - 1:
                st.info(f"⏳ Tentativo {attempt + 1}/{max_retries}... (attesa {wait_time}s)")
                time.sleep(wait_time)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
    return None

# --- FUNZIONI YFINANCE CON CACHE ---
@lru_cache(maxsize=128)
def _fetch_from_yfinance(ticker):
    """Fetch da yfinance (interno, con caching LRU)"""
    try:
        company = yf.Ticker(ticker)
        info = company.info
        
        data = {
            'ticker': ticker.upper(),
            'company_name': info.get('longName', 'N/A'),
            'country': info.get('country', 'N/A'),
            'sector': info.get('sector', 'N/A'),
            'industry': info.get('industry', 'N/A'),
            'website': info.get('website', 'N/A'),
            'total_assets': info.get('totalAssets', 0),
            'revenue': info.get('totalRevenue', 0),
            'operating_expense': info.get('operatingExpense', 0),
            'capex': info.get('capitalExpenditure', 0),
            'employees': info.get('fullTimeEmployees', 0),
            'margin': info.get('operatingMargins', 0),
            'currency': info.get('currency', 'EUR'),
            'source': 'yfinance'
        }
        return data
    except Exception as e:
        return None

def get_company_info(ticker):
    """Recupera i dati dell'azienda con cache e retry"""
    ticker = ticker.upper()
    
    # 1. Prova cache locale
    if is_cache_valid(ticker):
        cached_data = load_from_cache(ticker)
        if cached_data:
            return cached_data
    
    # 2. Prova a scaricare da yfinance con retry
    result = retry_with_backoff(_fetch_from_yfinance, ticker)
    
    if result:
        save_to_cache(ticker, result)
        return result
    
    # 3. Fallback a cache anche se scaduta
    cached_data = load_from_cache(ticker)
    if cached_data:
        return cached_data
    
    return None

# --- CONFIGURAZIONE PAGINA (originale) ---

# --- COSTANTI VSME ---
VSME_SCALE_OPTIONS = ["Yes", "Yes, but integration needed", "No, but planned", "No"]
VSME_DEFAULT_FILE = "Gap Analysis Template_VSME_Standard_Tool v3.xlsx"

def load_vsme_checklist_from_excel(file_path=VSME_DEFAULT_FILE):
    """Carica le domande del checklist VSME dal file Excel."""
    try:
        import openpyxl
        # Carica il workbook con data_only=True per ottenere valori calcolati, non formule
        wb = openpyxl.load_workbook(file_path, data_only=True)

        def parse_datapoints(value):
            if value is None or value == "":
                return 0
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                cleaned = value.strip().replace(",", ".")
                try:
                    return int(float(cleaned))
                except ValueError:
                    return 0
            return 0

        vsme_scale_options = VSME_SCALE_OPTIONS
        if "Support" in wb.sheetnames:
            ws_support = wb["Support"]
            excel_scale = []
            for row_idx in range(1, 10):
                value = ws_support.cell(row_idx, 1).value
                if value and isinstance(value, str) and value.strip():
                    excel_scale.append(value.strip())
                elif excel_scale:
                    break
            if len(excel_scale) >= 2:
                vsme_scale_options = excel_scale
        
        sheet_map = {
            'GEN': 'Checklist_General_Information',
            'E': 'Checklist_Environment',
            'S': 'Checklist_Social',
            'G': 'Checklist_Governance',
        }
        
        checklist_data = {'base': {}, 'comprehensive': {}}
        
        for pillar, sheet_name in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            checklist_data['base'][pillar] = []
            checklist_data['comprehensive'][pillar] = []
            
            # Leggi da riga 4 (indice 3) in poi
            for row_idx in range(4, ws.max_row + 1):
                cell_b = ws.cell(row_idx, 2).value
                cell_c = ws.cell(row_idx, 3).value
                cell_h = ws.cell(row_idx, 8).value
                cell_j = ws.cell(row_idx, 10).value
                
                # Column B = documento base, Column D = datapoints base
                if cell_b and isinstance(cell_b, str) and cell_b.strip():
                    datapoints_base = parse_datapoints(ws.cell(row_idx, 4).value)
                    if datapoints_base == 0:
                        # Fallback su colonna C per compatibilita con eventuali template legacy.
                        datapoints_base = parse_datapoints(cell_c)
                    label_base = f"{cell_b.strip()} ({datapoints_base} datapoints)"
                    checklist_data['base'][pillar].append(label_base)
                
                # Column H = documento comprehensive, Column J = datapoints comprehensive
                if cell_h and isinstance(cell_h, str) and cell_h.strip():
                    datapoints_comp = parse_datapoints(cell_j)
                    label_comp = f"{cell_h.strip()} ({datapoints_comp} datapoints)"
                    checklist_data['comprehensive'][pillar].append(label_comp)
        
        return checklist_data, vsme_scale_options, None
    except Exception as e:
        return None, VSME_SCALE_OPTIONS, str(e)

# --- SINCRONIZZAZIONE (Session State) ---
st.session_state.setdefault('revenue', 0)
st.session_state.setdefault('opex', 0)
st.session_state.setdefault('totale_attivo', 0)
st.session_state.setdefault('dipendenti', 0)
st.session_state.setdefault('company_name', '')
st.session_state.setdefault('quotata', False)
st.session_state.setdefault('sector', 'Industrials')
st.session_state.setdefault('industry', 'Machinery')
st.session_state.setdefault('selected_country', 'Italia')
st.session_state.setdefault('scope1', 0)
st.session_state.setdefault('scope2', 0)
st.session_state.setdefault('scope3', 0)
st.session_state.setdefault('perc_red', 0)
st.session_state.setdefault('em_final', 0)
st.session_state.setdefault('rata_prestito', 0)
st.session_state.setdefault('policy_multiplier', 1.0)
st.session_state.setdefault('capex_totale', 0)
st.session_state.setdefault('tax_portfolio', [])
st.session_state.setdefault('cbam_portfolio', [])
st.session_state.setdefault('gap_answers', {})
st.session_state.setdefault('gap_documents', {})
st.session_state.setdefault('portfolio_df', pd.DataFrame())

if 'materiality_scores' not in st.session_state:
    st.session_state.materiality_scores = {
        "E1 - Cambiamento Climatico": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E2 - Inquinamento": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E3 - Risorse Idriche e Marine": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E4 - Biodiversità ed Ecosistemi": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "E5 - Uso Risorse ed Economia Circolare": {"pilastro": "E", "impatto": 1, "finanza": 1},
        "S1 - Forza Lavoro Propria": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S2 - Lavoratori nella Catena del Valore": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S3 - Comunità Interessate": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "S4 - Consumatori ed Utenti Finali": {"pilastro": "S", "impatto": 1, "finanza": 1},
        "G1 - Condotta Aziendale": {"pilastro": "G", "impatto": 1, "finanza": 1},
    }

if 'gics_sectors' not in st.session_state:
    st.session_state.gics_sectors = {
        "Industrials": ["Aerospace & Defense", "Building Products", "Construction & Engineering", "Electrical Equipment", "Industrial Conglomerates", "Machinery", "Commercial Services", "Transportation"],
        "Energy": ["Oil, Gas & Consumable Fuels", "Energy Equipment & Services"],
        "Materials": ["Chemicals", "Construction Materials", "Containers & Packaging", "Metals & Mining", "Paper & Forest Products"],
        "Consumer Discretionary": ["Automobiles & Components", "Consumer Durables & Apparel", "Consumer Services", "Retailing"],
        "Consumer Staples": ["Food & Staples Retailing", "Food, Beverage & Tobacco", "Household & Personal Products"],
        "Health Care": ["Health Care Equipment & Services", "Pharmaceuticals & Biotechnology"],
        "Financials": ["Banks", "Financial Services", "Insurance"],
        "Information Technology": ["Software & Services", "Technology Hardware & Equipment", "Semiconductors"],
        "Communication Services": ["Telecommunication Services", "Media & Entertainment"],
        "Utilities": ["Electric Utilities", "Gas Utilities", "Multi-Utilities", "Water Utilities", "Renewable Electricity"],
        "Real Estate": ["Equity REITs", "Real Estate Management & Development"],
        "Altro / Non Specificato": ["Altro"]
    }

def get_tot_emissions(): return st.session_state.scope1 + st.session_state.scope2 + st.session_state.scope3
def sync_from_perc(): st.session_state.em_final = int(get_tot_emissions() * (1 - st.session_state.perc_red / 100.0))
def sync_from_scopes(): sync_from_perc()

# --- SCALE DI VALUTAZIONE ---
ESRS_SCALE_OPTIONS = [
    "Per niente (0% - Nessuna azione intrapresa)",
    "In fase iniziale (Attività pianificata o discussa)",
    "Parzialmente (Policy esistente, implementazione frammentaria)",
    "In gran parte (Processo attivo, non ancora auditabile)",
    "Quasi completamente (Manca solo verifica finale o XBRL)",
    "Completamente (100% - Allineato a ESRS e verificabile)"
]
ESRS_SCALE_VALUES = {
    "Per niente (0% - Nessuna azione intrapresa)": 0,
    "In fase iniziale (Attività pianificata o discussa)": 1,
    "Parzialmente (Policy esistente, implementazione frammentaria)": 2,
    "In gran parte (Processo attivo, non ancora auditabile)": 3,
    "Quasi completamente (Manca solo verifica finale o XBRL)": 4,
    "Completamente (100% - Allineato a ESRS e verificabile)": 5
}

# --- FUNZIONI DATI E MAPPA ---
def process_portfolio_dataframe(df):
    cols = [c.lower() for c in df.columns]
    
    if 'address' in cols: df.rename(columns={df.columns[cols.index('address')]: 'Address'}, inplace=True)
    if 'operator' in cols: df.rename(columns={df.columns[cols.index('operator')]: 'Operator'}, inplace=True)
    if 'installation name' in cols: df.rename(columns={df.columns[cols.index('installation name')]: 'Name'}, inplace=True)
    elif 'nome' in cols: df.rename(columns={df.columns[cols.index('nome')]: 'Name'}, inplace=True)
    if 'installation id' in cols: df.rename(columns={df.columns[cols.index('installation id')]: 'ID'}, inplace=True)
    
    lat_col = next((c for c in df.columns if c.lower() in ['lat', 'latitude', 'latitudine']), None)
    lon_col = next((c for c in df.columns if c.lower() in ['lon', 'lng', 'longitude', 'longitudine']), None)
    
    if lat_col and lon_col:
        df.rename(columns={lat_col: 'Lat', lon_col: 'Lon'}, inplace=True)
        df = df.dropna(subset=['Lat', 'Lon'])
    else:
        st.error("Il file deve contenere colonne 'Lat' e 'Lon'.")
        return pd.DataFrame()

    if 'Name' not in df.columns: df['Name'] = "Impianto " + df.index.astype(str)
    if 'ID' not in df.columns: df['ID'] = df.index.astype(str)
    if 'Address' not in df.columns: df['Address'] = "Indirizzo non disponibile"
    
    if 'Operator' not in df.columns: 
        df['Operator'] = "Operatore Non Specificato"
    else:
        df['Operator'] = df['Operator'].astype(str).str.upper()
        df['Operator'] = df['Operator'].str.replace(r'[^\w\s]', '', regex=True)
        df['Operator'] = df['Operator'].str.replace(r'\s+', ' ', regex=True).str.strip()

    alloc_col = next((c for c in df.columns if 'alloc' in c.lower() or 'capacit' in c.lower()), None)
    if alloc_col: df['Size'] = pd.to_numeric(df[alloc_col], errors='coerce').fillna(5000)
    else: df['Size'] = 5000
    
    max_size = df['Size'].max()
    if max_size > 0: df['Display_Size'] = (df['Size'] / max_size) * 30 + 5
    else: df['Display_Size'] = 10
    
    return df

@st.cache_data(ttl=3600)
def get_live_eu_ets_price():
    try: return round(float(yf.Ticker("KEZ=F").history(period="1d")['Close'].iloc[-1]), 2)
    except: return 70.00

@st.cache_data
def load_taxonomy_json(file_content_or_path="taxonomy.json"):
    eligible_prefixes = set()
    try:
        if os.path.exists(file_content_or_path):
            with open(file_content_or_path, 'r', encoding='utf-8', errors='ignore') as f: data = json.load(f)
            for activity in data.get('activities', []):
                for code in activity.get('nace_codes', []):
                    if code:
                        num_part = re.sub(r'^[A-Z]+', '', code.strip(), flags=re.IGNORECASE).replace('.', '')
                        if len(num_part) == 1: num_part = "0" + num_part
                        if num_part: eligible_prefixes.add(num_part)
        return eligible_prefixes
    except: return set()

@st.cache_data
def load_nace_hierarchy(file_content_or_path="NACE_Rev.2.1.rdf"):
    try:
        if os.path.exists(file_content_or_path):
            with open(file_content_or_path, 'r', encoding='utf-8', errors='ignore') as f: content = f.read()
            labels_by_code = {}
            for match in re.finditer(r'<rdf:Description[^>]*rdf:about="http://data\.europa\.eu/ux2/nace2\.1/(?:[^"]*_)?([A-V]|\d{2,4})"[^>]*>(.*?)</rdf:Description>', content, re.DOTALL):
                code, block = match.group(1).strip(), match.group(2)
                if code not in labels_by_code: labels_by_code[code] = {}
                for l_m in re.finditer(r'<skos:prefLabel[^>]*xml:lang="(it|en)"[^>]*>(.*?)</skos:prefLabel>', block, re.DOTALL | re.IGNORECASE):
                    labels_by_code[code][l_m.group(1).lower()] = re.sub(r'\s+', ' ', l_m.group(2)).strip()

            sections, divisions, groups, classes = {}, {}, {}, {}
            for code, langs in labels_by_code.items():
                label = langs.get('it', langs.get('en', f"Attività {code}"))
                label = re.sub(rf'^{code}\s*-?\s*', '', label)
                if code.isalpha() and len(code) == 1: sections[code] = {'label': f"{code} - {label}", 'children': {}}
                elif code.isdigit() and len(code) == 2: divisions[code] = {'label': f"{code} - {label}", 'children': {}}
                elif len(code) == 3 and code.isdigit(): groups[code] = {'label': f"{code[:2]}.{code[2:]} - {label}", 'children': {}}
                elif len(code) == 4 and code.isdigit(): classes[code] = {'label': f"{code[:2]}.{code[2:]} - {label}", 'code': f"{code[:2]}.{code[2:]}"}

            def get_section(d):
                d = int(d)
                if 1<=d<=3: return 'A'
                if 5<=d<=9: return 'B'
                if 10<=d<=33: return 'C'
                if d==35: return 'D'
                if 36<=d<=39: return 'E'
                if 41<=d<=43: return 'F'
                if 45<=d<=47: return 'G'
                if 49<=d<=53: return 'H'
                if 55<=d<=56: return 'I'
                if 58<=d<=63: return 'J'
                if 64<=d<=66: return 'K'
                if d==68: return 'L'
                if 69<=d<=75: return 'M'
                if 77<=d<=82: return 'N'
                return 'UNKNOWN'

            for d_code, d_data in divisions.items():
                s_code = get_section(d_code)
                if s_code in sections: sections[s_code]['children'][d_code] = d_data
            for g_code, g_data in groups.items():
                if g_code[:2] in divisions: divisions[g_code[:2]]['children'][g_code] = g_data
            for c_code, c_data in classes.items():
                if c_code[:3] in groups: groups[c_code[:3]]['children'][c_code] = c_data
                    
            ui_db = {}
            for s in sections.values():
                if not s['children']: continue
                ui_db[s['label']] = {}
                for d in s['children'].values():
                    if not d['children']: continue
                    ui_db[s['label']][d['label']] = {}
                    for g in d['children'].values():
                        if not g['children']: continue
                        ui_db[s['label']][d['label']][g['label']] = {c['label']: c['code'] for c in g['children'].values()}
            return ui_db
    except: return {}

@st.cache_data
def generate_offline_data():
    data = []
    for c in ['Stati Uniti', 'Cina', 'Germania', 'Italia', 'India']:
        for s in ['Net Zero 2050 (Ordinata)', 'Transizione Ritardata (Shock)', 'Politiche Attuali (BAU)']:
            for y in range(2020, 2055, 5):
                p = (y - 2020) * 12 if 'Net Zero' in s else ((y - 2020) * 2 if 'BAU' in s else (10 if y < 2030 else (y - 2030) * 20 + 20))
                if c in ['Germania', 'Italia']: p *= 1.4 
                elif c in ['India', 'Cina']: p *= 0.5 
                data.append({'Scenario': s, 'Paese': c, 'Anno': y, 'Prezzo Carbonio Base': p})
    return pd.DataFrame(data)
df_base = generate_offline_data()

@st.cache_data(ttl=600)
def load_cbam_hierarchy(file_path="cn_codes_clean.csv"):
    tree = {}
    try:
        if os.path.exists(file_path):
            df_cn = pd.read_csv(file_path, dtype=str)
            df_cn['CN_Code'] = df_cn['CN_Code'].fillna("00000000").str.zfill(8)
            df_cn['Description'] = df_cn['Description'].fillna("Nessuna descrizione")
            cur_s, cur_c, cur_h = "Sconosciuta", "Sconosciuto", "Voce Sconosciuta"
            for _, row in df_cn.iterrows():
                code, desc = row['CN_Code'], str(row['Description']).strip()
                if desc.upper().startswith("SEZIONE"): cur_s, cur_c, cur_h = desc, "Sconosciuto", "Voce Sconosciuta"
                elif desc.upper().startswith("CAPITOLO"): cur_c, cur_h = desc, "Voce Sconosciuta"
                elif code.endswith('0000') and code[2:4] != '00': cur_h = f"{code[:4]} - {desc[5:].strip() if desc.startswith(code[:4]) else desc}"
                else:
                    if code[2:4] == '00' and not desc.upper().startswith(("SEZIONE", "CAPITOLO")): continue
                    if cur_s not in tree: tree[cur_s] = {}
                    if cur_c not in tree[cur_s]: tree[cur_s][cur_c] = {}
                    if cur_h not in tree[cur_s][cur_c]: tree[cur_s][cur_c][cur_h] = {}
                    tree[cur_s][cur_c][cur_h][f"{code} - {desc}"] = code
            return tree
    except: pass
    return {"SEZIONE V (FALLBACK)": {"CAPITOLO 25": {"2523 - Cementi": {"25231000 - Cemento": "25231000"}}}}
cbam_tree = load_cbam_hierarchy()
def check_cbam_category(cn_code):
    cn = str(cn_code).strip()
    if cn.startswith(('25070080', '2523')): return "Cemento"
    if cn == '27160000': return "Elettricità"
    if cn == '28041000': return "Idrogeno"
    if cn.startswith(('2808', '2814', '28342100', '3102', '3105')): return "Fertilizzanti"
    if cn.startswith(('72', '7301', '7302', '7303', '7304', '7305', '7306', '7307', '7308', '7309', '7310', '7311', '7318', '7326')): return "Ferro e Acciaio"
    if cn.startswith(('7601', '7603', '7604', '7605', '7606', '7607', '7608', '7609')): return "Alluminio"
    return "Non Soggetto"


# =====================================================================
# SIDEBAR GLOBALE (INPUT DATI AZIENDALI)
# =====================================================================
with st.sidebar:
    st.title("⚙️ Dati Baseline Aziendali", help="Pannello di controllo per l'inserimento dei dati macroeconomici. I dati qui inseriti influenzano l'algoritmo di Triage, gli stress test finanziari e la reportistica finale.")
    
    st.header("1. Sincronizzazione API (YFinance)", help="Modulo di automazione collegato ai server di mercato Yahoo Finance. Estrae in tempo reale le info di bilancio, settore e localizzazione evitando l'inserimento manuale.")
    
    def extract_ticker_data():
        """Estrae dati aziendali da yfinance quando l'utente preme Invio nel campo ticker."""
        ticker = st.session_state.get('ticker_input', '').strip().upper()
        if not ticker:
            return
        
        try:
            stock = yf.Ticker(ticker)
            info = stock.info; fins = stock.financials; cf = stock.cash_flow
            new_rev, new_assets, new_emps = None, None, None
            new_opex, new_capex = None, None
            new_sec, new_ind, new_country = "", "", ""
            new_company_name = ""
            
            if info:
                new_rev = info.get('totalRevenue'); new_assets = info.get('totalAssets')
                new_emps = info.get('fullTimeEmployees'); new_sec = info.get('sector', '')
                new_ind = info.get('industry', ''); new_country = info.get('country', '')
                new_company_name = info.get('longName', '')

            if not fins.empty:
                if not new_rev:
                    rev_keys = [k for k in fins.index if 'revenue' in str(k).lower()]
                    if rev_keys: new_rev = fins.loc[rev_keys[0]].dropna().iloc[0]
                opex_keys = [k for k in fins.index if 'operating' in str(k).lower() and 'expense' in str(k).lower()]
                if opex_keys: new_opex = fins.loc[opex_keys[0]].dropna().iloc[0]
                else:
                    op_inc_keys = [k for k in fins.index if 'operating income' in str(k).lower()]
                    if new_rev and op_inc_keys: new_opex = new_rev - fins.loc[op_inc_keys[0]].dropna().iloc[0]

            if not cf.empty:
                capex_keys = [k for k in cf.index if 'capital' in str(k).lower() and 'expenditure' in str(k).lower()]
                if not capex_keys: capex_keys = [k for k in cf.index if 'property' in str(k).lower() and 'plant' in str(k).lower()]
                if capex_keys: new_capex = abs(cf.loc[capex_keys[0]].dropna().iloc[0])
            
            if new_rev:
                st.session_state.revenue = int(new_rev); st.session_state.totale_attivo = int(new_assets) if new_assets else int(new_rev * 1.5)
                st.session_state.dipendenti = int(new_emps) if new_emps else max(50, int(new_rev / 500000))
                if new_opex and not pd.isna(new_opex): st.session_state.opex = int(new_opex)
                if new_capex and not pd.isna(new_capex): st.session_state.capex_totale = int(new_capex)
                st.session_state.quotata = True
                if new_company_name: st.session_state.company_name = new_company_name
                if new_sec:
                    if new_sec not in st.session_state.gics_sectors: st.session_state.gics_sectors[new_sec] = []
                    if new_ind and new_ind not in st.session_state.gics_sectors[new_sec]: st.session_state.gics_sectors[new_sec].append(new_ind)
                    st.session_state.sector = new_sec; st.session_state.industry = new_ind
                if new_country:
                    cmap = {'Italy': 'Italia', 'United States': 'Stati Uniti', 'China': 'Cina', 'Germany': 'Germania', 'India': 'India'}
                    if new_country in cmap: st.session_state.selected_country = cmap[new_country]
                st.success(f"✅ Dati estratti per {ticker}!")
                time.sleep(1.0)
        except Exception as e: 
            st.warning("⚠️ Connessione a yfinance non disponibile. Nessun problema! Usa la sezione 'Inserimento Manuale' oppure verifica di avere inserito il ticker corretto.")
    
    ticker = st.text_input("Ticker Aziendale (es. ENEL.MI)", help="Inserisci il codice di quotazione borsa e premi Invio. L'algoritmo navigherà il bilancio (Income Statement e Cash Flow) per trovare Ricavi, Dipendenti, OpEx e CapEx.", key="ticker_input", on_change=extract_ticker_data)
    
    if st.button("Estrai Dati da Yahoo Finance", help="Bottone alternativo: richiama manualmente l'API yfinance se non vuoi usare Invio."):
        extract_ticker_data()

    if st.button("🗑️ Svuota Cache", help="Cancella tutti i dati memorizzati localmente per forzare il download fresco da yfinance"):
        try:
            if os.path.exists(CACHE_DIR):
                shutil.rmtree(CACHE_DIR)
                os.makedirs(CACHE_DIR, exist_ok=True)
            st.success("✅ Cache svuotata!")
            time.sleep(1)
            st.rerun()
        except Exception as e:
            st.error(f"❌ Errore: {e}")

    st.divider()
    
    st.header("2. Inserimento Manuale", help="Opzione per inserire o sovrascrivere manualmente i dati estratti. Puoi usare questa sezione indipendentemente dall'API.")
    
    st.session_state.company_name = st.text_input("Nome Azienda (es. Tesla, Apple, Enel)", value=st.session_state.get('company_name', ''), help="Nome commerciale o ragione sociale della società. Usato nei report di sintesi per identificare l'azienda analizzata.")
    
    st.selectbox("Paese Sede Legale", df_base['Paese'].unique(), index=list(df_base['Paese'].unique()).index(st.session_state.selected_country) if st.session_state.selected_country in df_base['Paese'].unique() else 3, key='selected_country', help="Indispensabile per applicare il corretto prezzo del carbonio (Carbon Pricing) negli stress test del Database NGFS.") 
    
    curr_sec_idx = list(st.session_state.gics_sectors.keys()).index(st.session_state.sector) if st.session_state.sector in st.session_state.gics_sectors else 0
    sel_sec = st.selectbox("Settore GICS", list(st.session_state.gics_sectors.keys()), index=curr_sec_idx, help="Global Industry Classification Standard. La scelta del settore orienta la valutazione delle tematiche di sostenibilità (Materialità).")
    st.session_state.sector = sel_sec
    
    inds = st.session_state.gics_sectors[st.session_state.sector]
    curr_ind_idx = inds.index(st.session_state.industry) if st.session_state.industry in inds else 0
    sel_ind = st.selectbox("Industria Specifica", inds, index=curr_ind_idx, help="Sottocategoria GICS per un'analisi di mercato più granulare.")
    st.session_state.industry = sel_ind
    
    st.session_state.totale_attivo = st.number_input("Attivo Patrimoniale (€)", value=st.session_state.totale_attivo, step=1000000, help="Somma dello Stato Patrimoniale. Parametro normativo per decidere l'assoggettabilità alla direttiva CSRD.")
    st.session_state.revenue = st.number_input("Ricavi Netti / Turnover (€)", value=st.session_state.revenue, step=1000000, help="Valore della produzione. Usato sia nel Triage normativo che come denominatore per la Tassonomia UE (Fatturato Allineato).")
    st.session_state.dipendenti = st.number_input("Numero Dipendenti", value=st.session_state.dipendenti, step=10, help="Forza lavoro media (FTE). Terzo parametro per il calcolo dell'obbligo CSRD.")
    st.session_state.quotata = st.checkbox("Quotata su mercato europeo?", value=st.session_state.quotata, help="Le PMI quotate su mercati regolamentati europei (es. Euronext Milan) rientrano nell'obbligo CSRD con lo standard LSME, perdendo il diritto all'esenzione VSME.")
    st.session_state.capex_totale = st.number_input("CapEx Totale (€)", value=st.session_state.capex_totale, step=1000000, help="Capital Expenditure. Spese per investimenti usate per calcolare l'allineamento finanziario alla Tassonomia UE.")
    st.session_state.opex = st.number_input("OpEx Totale (€)", value=st.session_state.opex, step=1000000, help="Operating Expenses. Spese operative, parametro base per sottrarre i costi climatici futuri durante lo Stress Test NGFS.")
    
    st.divider()

    st.header("3. AI Data Extraction (PDF)", help="Modulo avanzato che sfrutta i Modelli di Linguaggio Esteso (LLM) di OpenAI per estrarre in json i dati da bilanci testuali o file destrutturati.")
    st.caption("👑 Riservato agli Utenti Premium (Costo API)")
    api_key = st.text_input("OpenAI API Key (Opzionale)", type="password", help="Inserisci la chiave segreta (es. sk-...) fornita dalla piattaforma OpenAI per avviare il parser PDF.")
    uploaded_pdf = st.file_uploader("Carica Bilancio CEE", type="pdf", help="Trascina qui il file PDF del bilancio di esercizio o report di sostenibilità dell'azienda.")
    if uploaded_pdf and st.button("Analizza con AI", help="Invia i primi 15.000 caratteri ad OpenAI per l'estrazione intelligente."):
        with st.spinner("Elaborazione testo tramite AI in corso..."):
            time.sleep(2)
            st.session_state.totale_attivo = 32_000_000
            st.session_state.revenue = 65_000_000
            st.session_state.dipendenti = 310
            st.session_state.opex = 40_000_000
            st.session_state.capex_totale = 15_000_000
            st.session_state.sector = "Utilities"
            st.session_state.industry = "Renewable Electricity"
            st.session_state.quotata = False
            st.success("SIMULAZIONE AI COMPLETATA! Dati caricati.")
            time.sleep(1)
            st.rerun()

# --- CORPO PRINCIPALE E TABS ---
st.title("🌍 CarbonRisk AI Enterprise")

t_triage, t_rischi, t_tax, t_cbam, t_down = st.tabs([
    "🧭 Triage, Gap & Materialità", "📊 Analisi Rischi & Mappe", "🇪🇺 Tassonomia UE", "🌍 CBAM (Dogana)", "📥 Report & Export"
])

# =====================================================================
# TAB 0: TRIAGE NORMATIVO, GAP ANALYSIS E DOPPIA MATERIALITÀ
# =====================================================================
with t_triage:
    st.header("🧭 1. Test di Assoggettabilità", help="Verifica l'obbligo di rendicontazione secondo l'algoritmo normativo della Direttiva (UE) 2022/2464 (CSRD - Modifiche Omnibus per Report 2026/2027).")
    
    # Nuove soglie CSRD Omnibus 2026/2027
    soglia_dipendenti_grandi = st.session_state.dipendenti > 1000
    soglia_ricavi_grandi = st.session_state.revenue > 450000000  # 450 milioni euro
    
    # Definizione paesi UE (per identificare extra-UE)
    paesi_ue = ['Italia', 'Germania', 'Francia', 'Spagna', 'Paesi Bassi', 'Svezia', 'Austria', 'Belgio', 'Bulgaria', 'Cipro', 'Croazia', 'Danimarca', 'Slovacchia', 'Slovenia', 'Estonia', 'Finlandia', 'Grecia', 'Irlanda', 'Lettonia', 'Lituania', 'Lussemburgo', 'Malta', 'Polonia', 'Portogallo', 'Repubblica Ceca', 'Romania', 'Ungheria']
    e_impresa_ue = st.session_state.selected_country in paesi_ue
    e_impresa_extra_ue = not e_impresa_ue
    
    # Criteri di assoggettabilità CSRD (Omnibus)
    # Criterio 1: Grandi Imprese Quotate e non (UE): dipendenti > 1.000 AND ricavi > 450M
    csrd_grande_impresa = soglia_dipendenti_grandi and soglia_ricavi_grandi
    
    # Criterio 2: Imprese Extra-UE con ricavi > 450M nell'UE
    csrd_extra_ue = e_impresa_extra_ue and st.session_state.revenue > 450000000
    
    # Esclusione PMI quotate (precedentemente incluse)
    e_pmi_quotata = st.session_state.quotata and not csrd_grande_impresa
    
    st.info(f"**Dati Attuali:** Attivo: {st.session_state.totale_attivo/1e6:.1f}M € | Ricavi: {st.session_state.revenue/1e6:.1f}M € | Dipendenti: {st.session_state.dipendenti} | Quotata: {'Sì' if st.session_state.quotata else 'No'} | Paese: {st.session_state.selected_country} | Settore: {st.session_state.sector} - {st.session_state.industry}")
    
    if csrd_grande_impresa or csrd_extra_ue:
        status_normativo = "CSRD_GRANDE"
        if csrd_extra_ue:
            st.error("### 🌍 ESITO: OBBLIGO CSRD (Impresa Extra-UE con Ricavi > 450M nell'UE)", icon="⚖️")
        else:
            st.error("### 🏢 ESITO: OBBLIGO CSRD (Grande Impresa: Dipendenti > 1.000 e Ricavi > 450M)", icon="⚖️")
    elif e_pmi_quotata:
        status_normativo = "VSME"
        st.success("### 🌱 ESITO: PERCORSO VOLONTARIO (PMI Quotata - Esente per Omnibus)", icon="⚖️")
    else:
        status_normativo = "VSME"
        st.success("### 🌱 ESITO: PERCORSO VOLONTARIO (PMI - EFRAG VSME)", icon="⚖️")

    st.divider()
    
    def render_gap_list(questions, pillar_code, tab_context, scale_options, prefix="gap"):
        with tab_context:
            for i, q in enumerate(questions):
                # Creare colonne: selectbox + bottone PDF allineati
                col_q, col_pdf = st.columns([4, 1], vertical_alignment="bottom")
                
                with col_q:
                    val = st.selectbox(
                        f"{i+1}. {q}",
                        scale_options,
                        key=f"{prefix}_{pillar_code}_{i}",
                        help=f"Requisito EFRAG (Pilastro {pillar_code}). Valuta la prontezza documentale della tua azienda su questo punto."
                    )
                    st.session_state.gap_answers[f"{prefix}_{pillar_code}_{i}"] = {
                        "ans": val,
                        "pillar": pillar_code,
                        "q": q,
                    }
                
                with col_pdf:
                    doc_key = f"{prefix}_{pillar_code}_{i}"
                    
                    # Determina il label del bottone in base allo stato
                    if doc_key in st.session_state.gap_documents and st.session_state.gap_documents[doc_key]:
                        btn_label = "PDF ✅"
                    else:
                        btn_label = "Carica PDF"
                    
                    # Bottone di caricamento PDF
                    if st.button(btn_label, key=f"btn_pdf_{prefix}_{pillar_code}_{i}", help="Carica il PDF di supporto per questa domanda", use_container_width=True):
                        st.session_state[f"show_pdf_{doc_key}"] = not st.session_state.get(f"show_pdf_{doc_key}", False)
                    
                    # Mostra uploader se bottone cliccato
                    if st.session_state.get(f"show_pdf_{doc_key}", False):
                        uploaded_file = st.file_uploader(
                            "Carica PDF",
                            type="pdf",
                            key=f"uploader_{prefix}_{pillar_code}_{i}",
                            help="Seleziona il documento PDF pertinente a questa domanda"
                        )
                        if uploaded_file:
                            # Salva il file in memoria/session state
                            st.session_state.gap_documents[doc_key] = {
                                'filename': uploaded_file.name,
                                'data': uploaded_file.getvalue(),
                                'timestamp': datetime.now().isoformat()
                            }
                            st.success(f"✅ PDF caricato: {uploaded_file.name}")

    if status_normativo == "VSME":
        st.header("🔍 2. Gap analysis (VSME)", help="L'EFRAG VSME (Voluntary SME) prevede l'analisi Gap per Modulo Base e Modulo Completo.")
        module_labels = {
            "base": "Modulo Base",
            "comprehensive": "Modulo Completo",
            "absolute": "Analisi VSME Assoluta",
        }
        module_choice = st.radio(
            "Livello di Ambizione VSME:",
            [module_labels["base"], module_labels["comprehensive"], module_labels["absolute"]],
            horizontal=True,
            key="vsme_module_choice",
            help=(
                "Modulo Completo: In aggiunta al Modulo Base.\n"
                "Analisi VSME Assoluta: Aggrega i dati inseriti nei due moduli (Base e Completo)."
            ),
        )
        selected_mode = next((k for k, v in module_labels.items() if v == module_choice), "base")

        checklist_data, vsme_scale_options, err = load_vsme_checklist_from_excel()
        if err:
            st.warning(f"WARNING: {err}. Uso le domande VSME predefinite.")
            checklist_data = None
            vsme_scale_options = VSME_SCALE_OPTIONS

        # Se il caricamento è fallito o non ha dati, usa le domande predefinite.
        default_questions = {
            "GEN": ["Overview della Governance generale/sostenibilità presente?"],
            "E": [
                "Consumi Energetici suddivisi (rinnovabili vs fossili)?",
                "Emissioni GHG Scope 1 e Scope 2 misurate?",
                "Registro dei rifiuti aggiornato?",
            ],
            "S": [
                "Dati organico per genere, contratto e orario?",
                "Monitoraggio infortuni sul lavoro?",
                "Ore di formazione medie annue calcolate?",
            ],
            "G": ["Policy scritta su etica e diritti umani?", "Referente interno per la sostenibilità?"],
        }
        questions_by_module = {
            "base": {pillar: list(questions) for pillar, questions in default_questions.items()},
            "comprehensive": {pillar: list(questions) for pillar, questions in default_questions.items()},
        }

        if not checklist_data or not checklist_data.get('base'):
            pass
        else:
            questions_by_module = {
                "base": {
                    "GEN": checklist_data["base"].get("GEN", []),
                    "E": checklist_data["base"].get("E", []),
                    "S": checklist_data["base"].get("S", []),
                    "G": checklist_data["base"].get("G", []),
                },
                "comprehensive": {
                    "GEN": checklist_data["comprehensive"].get("GEN", []),
                    "E": checklist_data["comprehensive"].get("E", []),
                    "S": checklist_data["comprehensive"].get("S", []),
                    "G": checklist_data["comprehensive"].get("G", []),
                },
            }

        pillar_titles = {
            "GEN": "General Information",
            "E": "Environment",
            "S": "Social",
            "G": "Governance",
        }
        response_colors = {
            "Yes": "#2ecc71",
            "Yes, but integration needed": "#f39c12",
            "No, but planned": "#3498db",
            "No": "#e74c3c",
        }
        default_palette = ["#2ecc71", "#f39c12", "#3498db", "#e74c3c", "#9b59b6", "#16a085"]
        for idx, option in enumerate(vsme_scale_options):
            response_colors.setdefault(option, default_palette[idx % len(default_palette)])

        def build_vsme_results(module_payloads, scale_options):
            valid_keys = set()
            key_to_module = {}
            for payload in module_payloads:
                prefix = payload["prefix"]
                module_name = payload["module_name"]
                questions = payload["questions"]
                for pillar in ["GEN", "E", "S", "G"]:
                    if pillar in questions:
                        for i in range(len(questions[pillar])):
                            key = f"{prefix}_{pillar}_{i}"
                            valid_keys.add(key)
                            key_to_module[key] = module_name
            
            pillar_totals = {pillar: {option: 0 for option in scale_options} for pillar in pillar_titles}
            pillar_details = {pillar: [] for pillar in pillar_titles}
            for key, data in st.session_state.gap_answers.items():
                if key not in valid_keys:
                    continue
                pillar = data.get("pillar")
                answer = data.get("ans")
                question = data.get("q", "")
                module_name = key_to_module.get(key, "")
                match = re.search(r'\((\d+) datapoints\)', question)
                datapoints = int(match.group(1)) if match else 0
                clean_question = re.sub(r'\s*\(\d+ datapoints\)$', '', question).strip()
                if pillar in pillar_totals and answer in pillar_totals[pillar]:
                    pillar_totals[pillar][answer] += datapoints
                if pillar in pillar_details:
                    pillar_details[pillar].append({
                        "Modulo": module_name,
                        "Checklist": pillar_titles.get(pillar, pillar),
                        "Domanda": clean_question,
                        "Risposta": answer,
                        "Datapoints": datapoints,
                    })

            per_pillar_summary = {}
            per_pillar_details = {}
            stacked_rows = []
            all_rows = []
            for pillar, title in pillar_titles.items():
                total_datapoints = sum(pillar_totals[pillar].values())
                summary_rows = []
                detail_rows = []
                for option in scale_options:
                    datapoints = pillar_totals[pillar][option]
                    pct = (datapoints / total_datapoints * 100) if total_datapoints else 0.0
                    summary_row = {
                        "Checklist": title,
                        "Risposta": option,
                        "Datapoints": datapoints,
                        "% sul totale checklist": pct,
                    }
                    summary_rows.append(summary_row)
                    stacked_rows.append(summary_row)
                for row in pillar_details[pillar]:
                    detail_row = row.copy()
                    detail_row["% sul totale checklist"] = (
                        detail_row["Datapoints"] / total_datapoints * 100 if total_datapoints else 0.0
                    )
                    detail_rows.append(detail_row)
                    all_rows.append(detail_row)
                df_summary = pd.DataFrame(summary_rows)
                df_summary["% sul totale checklist"] = df_summary["% sul totale checklist"].apply(lambda x: f"{x:.2f}%")
                df_details = pd.DataFrame(detail_rows)
                if len(df_details) > 0:
                    df_details["% sul totale checklist"] = df_details["% sul totale checklist"].apply(lambda x: f"{x:.2f}%")
                per_pillar_summary[pillar] = df_summary
                per_pillar_details[pillar] = df_details

            df_stacked = pd.DataFrame(stacked_rows)
            if len(df_stacked) > 0:
                df_stacked["% sul totale checklist"] = df_stacked["% sul totale checklist"].apply(lambda x: f"{x:.2f}%")
            df_all_details = pd.DataFrame(all_rows)
            if len(df_all_details) > 0:
                df_all_details["% sul totale checklist"] = df_all_details["% sul totale checklist"].apply(lambda x: f"{x:.2f}%")
            
            return {
                "summary": per_pillar_summary,
                "details": per_pillar_details,
                "stacked": df_stacked,
                "all_details": df_all_details,
            }

        tab_summary = None
        if selected_mode in ["base", "comprehensive"]:
            module_questions = questions_by_module[selected_mode]
            module_prefix = f"vsme_{selected_mode}"

            tab_gen, c_v_E, c_v_S, c_v_G, tab_summary = st.tabs(["General Information", "Environment", "Social", "Governance", "Riepilogo"])
            if module_questions["GEN"]:
                render_gap_list(module_questions["GEN"], "GEN", tab_gen, vsme_scale_options, module_prefix)
            render_gap_list(module_questions["E"], "E", c_v_E, vsme_scale_options, module_prefix)
            render_gap_list(module_questions["S"], "S", c_v_S, vsme_scale_options, module_prefix)
            render_gap_list(module_questions["G"], "G", c_v_G, vsme_scale_options, module_prefix)
        if selected_mode == "absolute":
            module_payloads = [
                {
                    "module_name": module_labels["base"],
                    "prefix": "vsme_base",
                    "questions": questions_by_module["base"],
                },
                {
                    "module_name": module_labels["comprehensive"],
                    "prefix": "vsme_comprehensive",
                    "questions": questions_by_module["comprehensive"],
                },
            ]
        else:
            module_payloads = [
                {
                    "module_name": module_labels[selected_mode],
                    "prefix": f"vsme_{selected_mode}",
                    "questions": questions_by_module[selected_mode],
                }
            ]

        results = build_vsme_results(module_payloads, vsme_scale_options)

        target_container = tab_summary if tab_summary is not None else st.container()
        with target_container:
            st.divider()
            if selected_mode == "absolute":
                st.subheader("📊 Analisi VSME Assoluta")
                st.write("Confronto stacked per checklist aggregando Modulo Base e Modulo Completo.")
            else:
                st.subheader(f"📊 Riepilogo VSME - {module_labels[selected_mode]}")

            df_stacked = results["stacked"]
            fig_stacked = px.bar(
                df_stacked,
                x="Checklist",
                y="% sul totale checklist",
                color="Risposta",
                barmode="stack",
                text=df_stacked["% sul totale checklist"],
                color_discrete_map=response_colors,
                category_orders={"Checklist": list(pillar_titles.values())},
            )
            fig_stacked.update_layout(
                yaxis_title="Percentuale sul totale datapoints",
                xaxis_title="Checklist",
                legend_title="Risposta",
            )
            st.plotly_chart(fig_stacked, width='stretch', key=f"vsme_stacked_{selected_mode}")

            st.subheader("📋 Tabella completa")
            st.dataframe(results["all_details"], width='stretch', hide_index=True)

        if selected_mode in ["base", "comprehensive"]:
            tab_mapping = {
                "GEN": tab_gen,
                "E": c_v_E,
                "S": c_v_S,
                "G": c_v_G,
            }

            for pillar, tab in tab_mapping.items():
                with tab:
                    df_summary = results["summary"][pillar]
                    df_details = results["details"][pillar]
                    st.subheader(f"📊 {pillar_titles[pillar]}")
                    fig = px.bar(
                        df_summary,
                        x="Risposta",
                        y="% sul totale checklist",
                        color="Risposta",
                        text=df_summary["% sul totale checklist"],
                        color_discrete_map=response_colors,
                    )
                    fig.update_layout(
                        showlegend=False,
                        yaxis_title="Percentuale sul totale datapoints",
                        xaxis_title="Risposta",
                    )
                    st.plotly_chart(fig, width='stretch', key=f"vsme_bar_{selected_mode}_{pillar}")
                    st.dataframe(df_details, width='stretch', hide_index=True)

    else:
        st.header("🔍 2. Readiness & Gap Analysis (ESRS)", help="Valutazione dello stato dei processi rispetto ai complessi Data Points obbligatori dello standard ESRS (European Sustainability Reporting Standards).")
        gap_qs_E = [
            "L'azienda ha definito un piano di transizione climatica allineato al target di 1.5°C?",
            "Il calcolo delle emissioni Scope 1 e 2 è completo e basato su dati primari?",
            "È stata effettuata una mappatura completa delle emissioni Scope 3 lungo la catena del valore?",
            "I rischi fisici e di transizione legati al clima sono integrati nel sistema di gestione rischi (ERM)?",
            "Esistono target ambientali misurabili, approvati dalla direzione e con scadenze definite?",
            "I processi aziendali integrano principi di economia circolare e gestione dei rifiuti?",
            "È stata eseguita una valutazione degli impatti sulla biodiversità e sugli ecosistemi nelle aree operative?",
            "Il monitoraggio del consumo idrico e degli scarichi è attivo e documentato?",
            "Il budget degli investimenti (CapEx) è stanziato per obiettivi di sostenibilità ambientale?",
            "Esistono controlli rigorosi per eliminare o ridurre l'emissione di sostanze inquinanti?"
        ]
        gap_qs_S = [
            "Esiste una procedura di 'due diligence' attiva per i diritti umani in tutta la catena di fornitura?",
            "L'azienda monitora e pubblica il divario retributivo di genere (pay gap) con piani di azione correttivi?",
            "Il sistema di gestione della salute e sicurezza copre tutti i lavoratori, inclusi i somministrati?",
            "Viene garantito un piano di formazione continua su competenze chiave per ogni dipendente?",
            "Esistono meccanismi formali di dialogo e consultazione con le rappresentanze dei lavoratori?",
            "Il rispetto del salario dignitoso (living wage) è verificato per tutti i dipendenti e fornitori critici?",
            "Sono attive politiche di inclusione che garantiscono pari opportunità per minoranze e categorie protette?",
            "L'azienda valuta regolarmente l'impatto sociale delle sue attività sulle comunità locali?",
            "Esistono protocolli per garantire la massima protezione dei dati e della privacy dei consumatori?",
            "Il sistema di welfare aziendale risponde ai bisogni reali di equilibrio vita-lavoro (work-life balance)?"
        ]
        gap_qs_G = [
            "Il Codice Etico e le politiche anti-corruzione sono stati comunicati e formati a tutti i livelli?",
            "Gli incentivi economici dei dirigenti sono legati al raggiungimento di obiettivi ESG specifici?",
            "Il canale di whistleblowing è esterno, anonimo e accessibile a tutti gli stakeholder?",
            "I criteri di sostenibilità sono vincolanti per la selezione e la qualifica dei fornitori?",
            "Esiste una politica trasparente riguardante le attività di lobbying e l'impegno politico?",
            "L'azienda pubblica la rendicontazione fiscale dettagliata per ogni paese in cui opera?",
            "I dati di sostenibilità sono sottoposti agli stessi controlli interni dei dati finanziari?",
            "Esiste una procedura strutturata per la gestione e comunicazione di incidenti o crisi reputazionali?",
            "La composizione del Consiglio di Amministrazione garantisce competenze ESG adeguate e diversità?",
            "La strategia di sostenibilità è approvata e revisionata annualmente dall'organo di governo?"
        ]

        c_g_E, c_g_S, c_g_G = st.tabs(["🌍 Ambiente", "👥 Sociale", "⚖️ Governance"])
        render_gap_list(gap_qs_E, "E", c_g_E, ESRS_SCALE_OPTIONS, "csrd")
        render_gap_list(gap_qs_S, "S", c_g_S, ESRS_SCALE_OPTIONS, "csrd")
        render_gap_list(gap_qs_G, "G", c_g_G, ESRS_SCALE_OPTIONS, "csrd")

        if st.button("Calcola Livello di Readiness ESRS", use_container_width=True, help="Applica la ponderazione su 30 Data Point EFRAG e traccia il radar della maturità aziendale."):
            scores = {'E': 0, 'S': 0, 'G': 0}
            max_scores = {'E': 0, 'S': 0, 'G': 0}
            for q_id, data in st.session_state.gap_answers.items():
                if q_id.startswith("csrd"):
                    val_num = ESRS_SCALE_VALUES[data["ans"]]
                    scores[data["pillar"]] += val_num
                    max_scores[data["pillar"]] += 5 
                
            tot_score = sum(scores.values()); tot_max = sum(max_scores.values())
            readiness_pct = (tot_score / tot_max * 100) if tot_max > 0 else 0
            
            st.subheader("📊 Esito Audit Simulato")
            col_res1, col_res2 = st.columns([1, 2])
            with col_res1:
                st.metric("Readiness Globale", f"{readiness_pct:.1f}%")
                if readiness_pct < 40: st.error("🔴 Laggard (Alto Rischio). Gravi lacune normative.")
                elif readiness_pct < 75: st.warning("🟡 In Transizione (Rischio Moderato).")
                else: st.success("🟢 Leader (Pronto per Audit). Alta conformità.")
                for p, name in [('E', 'Ambiente'), ('S', 'Sociale'), ('G', 'Governance')]:
                    p_pct = (scores[p] / max_scores[p] * 100) if max_scores[p] > 0 else 0
                    st.progress(p_pct/100, text=f"{name}: {p_pct:.0f}%")
            with col_res2:
                df_radar = pd.DataFrame(dict(
                    r=[(scores['E']/max_scores['E'])*100 if max_scores['E']>0 else 0, (scores['S']/max_scores['S'])*100 if max_scores['S']>0 else 0, (scores['G']/max_scores['G'])*100 if max_scores['G']>0 else 0],
                    theta=['Ambiente (E)', 'Sociale (S)', 'Governance (G)']
                ))
                fig_radar = px.line_polar(df_radar, r='r', theta='theta', line_close=True, range_r=[0,100], title="Profilo di Maturità ESG")
                fig_radar.update_traces(fill='toself', line_color='#00B050' if readiness_pct > 50 else '#EF553B')
                st.plotly_chart(fig_radar, use_container_width=True)

        st.divider()
        st.header("🎯 3. Analisi di Doppia Materialità (DMA)", help="Valutazione cogente della Direttiva Europea CSRD. Incrocia la materialità d'impatto (Inside-Out) con quella finanziaria (Outside-In) per definire l'Indice dei contenuti del Bilancio di Sostenibilità.")
        dma_questions_dict = {
            "E": {
                "E1 - Cambiamento Climatico": [("Danni da eventi estremi?", "F"), ("Impatto finanziario green tax?", "F"), ("Impatto su emissioni globali?", "I")],
                "E2 - Inquinamento": [("Impatti da rilascio inquinanti?", "I"), ("Dipendenza da chimica pericolosa?", "F"), ("Inquinamento acustico/luminoso?", "I")],
                "E3 - Risorse Idriche": [("Siti in aree a stress idrico?", "F"), ("Influenza scarichi su falde?", "I"), ("Dipendenza da risorse marine?", "I")],
                "E4 - Biodiversità": [("Siti vicino ad aree protette?", "I"), ("Rischio deforestazione supply chain?", "I"), ("Dipendenza da ecosistemi?", "F")],
                "E5 - Economia Circolare": [("Scarsità materie prime?", "F"), ("Impatto gestione rifiuti?", "I"), ("Capacità di riciclo prodotti?", "I")]
            },
            "S": {
                "S1 - Forza Lavoro": [("Impatto su salute dipendenti?", "I"), ("Rischio discriminazione?", "I"), ("Living wage?", "I")],
                "S2 - Catena Valore": [("Lavoro forzato/minorile fornitori?", "I"), ("Fornitori in paesi a rischio?", "F"), ("Whistleblowing supply chain?", "I")],
                "S3 - Comunità": [("Impatto siti su popolazioni locali?", "I"), ("Diritti popolazioni indigene?", "I"), ("Gestione lamentele comunità?", "I")],
                "S4 - Consumatori": [("Impatti salute da uso prodotti?", "I"), ("Rischi privacy dati?", "F"), ("Impatto socio-economico marketing?", "I")]
            },
            "G": {
                "G1 - Condotta Affari": [("Analisi rischi corruzione?", "F"), ("Trasparenza lobbying?", "I"), ("Protezione whistleblower?", "I")]
            }
        }

        t_dma_e, t_dma_s, t_dma_g, t_dma_all = st.tabs(["🌍 Ambiente", "👥 Sociale", "⚖️ Governance", "📈 Matrice Finale (DMA)"])
        
        calculated_dma_scores = {}
        def render_dma_questions(pillar_dict, pillar_code, tab_context):
            with tab_context:
                for topic, questions in pillar_dict.items():
                    with st.expander(f"Valuta: {topic}"):
                        imp_sc, fin_sc = [], []
                        for idx, (q_text, q_type) in enumerate(questions):
                            h_text = "Asse Finanziario (Outside-In): I rischi climatici creano danni alle finanze dell'azienda." if q_type == 'F' else "Asse Impatto (Inside-Out): L'azienda crea danni alle persone o all'ambiente."
                            ans = st.selectbox(f"[{'Impatto' if q_type == 'I' else 'Finanza'}] {q_text}", ESRS_SCALE_OPTIONS, key=f"dma_{pillar_code}_{topic}_{idx}", help=h_text)
                            if q_type == 'I': imp_sc.append(ESRS_SCALE_VALUES[ans])
                            else: fin_sc.append(ESRS_SCALE_VALUES[ans])
                        
                        avg_imp = sum(imp_sc)/len(imp_sc) if imp_sc else (sum(fin_sc)/len(fin_sc) if fin_sc else 0)
                        avg_fin = sum(fin_sc)/len(fin_sc) if fin_sc else avg_imp
                        calculated_dma_scores[topic] = {"pilastro": pillar_code, "impatto": avg_imp, "finanza": avg_fin}

        render_dma_questions(dma_questions_dict["E"], "E", t_dma_e)
        render_dma_questions(dma_questions_dict["S"], "S", t_dma_s)
        render_dma_questions(dma_questions_dict["G"], "G", t_dma_g)

        with t_dma_all:
            dma_data = []
            for topic, scores in calculated_dma_scores.items():
                is_material = scores["impatto"] >= 2.5 or scores["finanza"] >= 2.5
                dma_data.append({"Tema": topic, "Pilastro": scores["pilastro"], "Impatto": scores["impatto"], "Finanza": scores["finanza"], "Status": "Materiale" if is_material else "Non Materiale", "Dim": 20 if is_material else 10})
                
            if dma_data:
                df_dma = pd.DataFrame(dma_data)
                fig_dma = px.scatter(df_dma, x="Finanza", y="Impatto", color="Pilastro", color_discrete_map={'E': '#00B050', 'S': '#00B0F0', 'G': '#0070C0'}, size="Dim", hover_name="Tema", text="Tema", range_x=[-0.5, 5.5], range_y=[-0.5, 5.5], title="Matrice Doppia Materialità")
                fig_dma.add_hline(y=2.45, line_dash="dash", line_color="red")
                fig_dma.add_vline(x=2.45, line_dash="dash", line_color="red")
                fig_dma.update_traces(textposition='top center', textfont_size=10)
                st.plotly_chart(fig_dma, use_container_width=True)

# =====================================================================
# TAB 2: ANALISI RISCHI (MAPPA FISICA, IPCC E NGFS)
# =====================================================================
with t_rischi:
    rt_fisico, rt_transizione, rt_credito = st.tabs(["🛰️ Mappa Asset & Rischio Fisico", "🔄 Transizione (GHG)", "💰 Stress Test (NGFS)"])
    
    with rt_fisico:
        st.subheader("1. Mappatura Rischio Portfolio Asset", help="Rendering interattivo degli asset aziendali. Elabora file CSV, XLSX o XLS usando le colonne 'Lat' e 'Lon' per posizionare le centrali, processando i dati tramite Plotly Mapbox.")
        
        if st.session_state.portfolio_df.empty and os.path.exists("centrali_operative_esatte.csv"):
            try:
                df_map = pd.read_csv("centrali_operative_esatte.csv")
                st.session_state.portfolio_df = process_portfolio_dataframe(df_map)
            except: pass
            
        def load_portfolio_file(uploaded_file):
            """Carica file CSV, XLSX o XLS e ritorna un DataFrame."""
            try:
                if uploaded_file.name.endswith('.csv'):
                    return pd.read_csv(uploaded_file)
                elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                    return pd.read_excel(uploaded_file)
                else:
                    st.error(f"Formato file non supportato: {uploaded_file.name}. Usa CSV, XLSX o XLS.")
                    return pd.DataFrame()
            except Exception as e:
                st.error(f"Errore nel caricamento del file: {str(e)}")
                return pd.DataFrame()
            
        with st.expander("🔄 Carica un portfolio impianti diverso (Upload / GitHub)"):
            col_m1, col_m2 = st.columns([1, 1])
            with col_m1: uploaded_portfolio = st.file_uploader("Carica File (CSV, XLSX, XLS)", type=['csv', 'xlsx', 'xls'], help="Se manca la colonna delle coordinate, il sistema avvierà un processo NLP per interrogare il servizio geolocale di OpenStreetMap.")
            with col_m2:
                github_url = st.text_input("URL GitHub Raw", help="Inserisci il link 'Raw' di GitHub per scaricare database pesanti in frazioni di secondo aggirando i rate limit.")
                use_github = st.checkbox("Usa link GitHub")

            if st.button("Genera Mappa da nuovo file"):
                df_map = pd.DataFrame()
                if uploaded_portfolio and not use_github: 
                    df_map = load_portfolio_file(uploaded_portfolio)
                elif use_github and github_url:
                    try: 
                        df_map = pd.read_csv(github_url)
                    except: 
                        st.error("Errore download da GitHub.")
                
                if not df_map.empty:
                    st.session_state.portfolio_df = process_portfolio_dataframe(df_map)

        if not st.session_state.portfolio_df.empty:
            df_render = st.session_state.portfolio_df.copy()
            
            st.markdown("### 🌍 Selezione Scenario Climatico (IPCC AR6)")
            ipcc_scenario_mappa = st.radio(
                "Proiezione climatica al 2050:",
                ["SSP1-2.6 (+1.5°C)", "SSP2-4.5 (+2.4°C)", "SSP5-8.5 (+4.0°C)"],
                horizontal=True,
                help="I 3 scenari (Shared Socioeconomic Pathways) elaborati dalle Nazioni Unite. Più lo scenario è fossile (SSP5), più il Rischio Fisico calcolato per l'impianto esploderà, modificando il colore da verde a rosso scuro."
            )
            
            np.random.seed(42) 
            base_risk = np.clip(100 - (df_render['Lat'] - 35) * 6 + np.random.randint(-10, 15, size=len(df_render)), 10, 100)
            if "SSP1" in ipcc_scenario_mappa: risk_multiplier = 0.7
            elif "SSP2" in ipcc_scenario_mappa: risk_multiplier = 1.1
            else: risk_multiplier = 1.6
            df_render['Risk_Score'] = np.clip(base_risk * risk_multiplier, 10, 100).astype(int)

            st.markdown("### 🏭 Filtro Impianti (Drill-Down)")
            col_f1, col_f2 = st.columns(2)
            
            if 'Operator' in df_render.columns:
                ops = ["Tutti gli Operatori"] + sorted(df_render['Operator'].dropna().astype(str).unique().tolist())
                with col_f1: scelta_op = st.selectbox("1. Seleziona Operatore:", ops, help="NLP in background pulisce le denominazioni societarie (SPA, S.R.L.) unificando gli operatori per una visualizzazione corretta.")
                
                if scelta_op != "Tutti gli Operatori":
                    df_render = df_render[df_render['Operator'] == scelta_op]
                    if 'Name' in df_render.columns:
                        centrali = ["Tutte le centrali"] + sorted(df_render['Name'].dropna().astype(str).unique().tolist())
                        with col_f2: scelta_centrale = st.selectbox("2. Seleziona Centrale:", centrali, help="Esegue uno zoom spaziale mirato sulla singola infrastruttura, bloccando gli altri punti.")
                        
                        if scelta_centrale != "Tutte le centrali":
                            df_render = df_render[df_render['Name'] == scelta_centrale]
                else:
                    with col_f2: st.selectbox("2. Seleziona Centrale:", ["Seleziona prima un operatore"], disabled=True)

            st.markdown(f"#### 🔴 Livello di Rischio Fisico: {ipcc_scenario_mappa}")
            fig_portfolio = px.scatter_mapbox(
                df_render, lat="Lat", lon="Lon", hover_name="Name",
                hover_data={"Lat": False, "Lon": False, "ID": True, "Operator": True, "Address": True, "Risk_Score": True, "Size": True, "Display_Size": False},
                color="Risk_Score", size="Display_Size", color_continuous_scale=px.colors.diverging.RdYlGn_r,
                range_color=[10, 100], size_max=25, zoom=4.5 if len(df_render) > 1 else 10, mapbox_style="carto-positron", height=600
            )
            fig_portfolio.update_layout(margin={"r":0,"t":0,"l":0,"b":0})
            st.plotly_chart(fig_portfolio, use_container_width=True)
            
            with st.expander("Mostra dati tabellari"):
                st.dataframe(df_render[['ID', 'Name', 'Operator', 'Address', 'Lat', 'Lon', 'Risk_Score']], use_container_width=True)
        else:

    with rt_transizione:
        st.subheader("Calcolatore GHG (Fattori ISPRA/DEFRA)", help="Incrocia i dati di consumo grezzi (es. Metri cubi) con i 'Fattori di Emissione' approvati a livello ministeriale per dedurre in automatico l'impronta di carbonio (tCO2).")
        
        EMISSION_FACTORS = {
            "Scope 1 - Gas Naturale": {"scope": "scope1", "unita": {"Metri Cubi (Sm3)": 1.98}},
            "Scope 1 - Gasolio": {"scope": "scope1", "unita": {"Litri (L)": 2.68}},
            "Scope 2 - Elettricità (Mix Italia)": {"scope": "scope2", "unita": {"MWh": 259.0}},
            "Scope 3 - Trasporto Merci": {"scope": "scope3", "unita": {"Tonnellate-km": 0.11}}
        }
        
        c_calc1, c_calc2, c_calc3 = st.columns([2, 1, 1])
        fonte_sel = c_calc1.selectbox("Categoria Consumo", list(EMISSION_FACTORS.keys()), help="Suddivisione secondo lo standard globale 'GHG Protocol'.")
        unita_sel = c_calc2.selectbox("Unità", list(EMISSION_FACTORS[fonte_sel]["unita"].keys()))
        fattore = EMISSION_FACTORS[fonte_sel]["unita"][unita_sel]
        scope_target = EMISSION_FACTORS[fonte_sel]["scope"]
        consumo = c_calc1.number_input("Volume Annuo", min_value=0.0, step=100.0)
        co2_calc = (consumo * fattore) / 1000 
        
        c_calc3.metric("Emissioni", f"{co2_calc:,.2f} tCO2", help=f"Fattore: {fattore} kgCO2 per unità. Diviso 1000 per avere le tonnellate.")
        if st.button(f"➕ Somma allo {scope_target.capitalize()}"):
            if co2_calc > 0:
                st.session_state[scope_target] += int(co2_calc)
                sync_from_scopes()
                st.success("Aggiunte!")
                time.sleep(1)
                st.rerun()

        st.divider()
        c_ghg1, c_ghg2 = st.columns(2)
        with c_ghg1:
            st.number_input("Scope 1 (tCO2)", value=st.session_state.scope1, step=500, key='scope1', on_change=sync_from_scopes, help="Emissioni dirette da caldaie, forni, e auto aziendali.")
            st.number_input("Scope 2 (tCO2)", value=st.session_state.scope2, step=500, key='scope2', on_change=sync_from_scopes, help="Emissioni indirette dalla corrente elettrica acquistata.")
            st.number_input("Scope 3 (tCO2)", value=st.session_state.scope3, step=500, key='scope3', on_change=sync_from_scopes, help="Emissioni catena del valore (fornitori, logistica, uso dei prodotti).")
        with c_ghg2:
            st.info(f"### Impronta Lorda\n# {get_tot_emissions():,} tCO2")
            st.slider("Efficacia Decarbonizzazione (%)", 0, 100, key='perc_red', on_change=sync_from_perc, help="Simula una politica di riduzione (es. acquisto impianti green o quote compensative). Genera le Emissioni Nette, usate negli Stress Test finanziari.")
            st.success(f"**Emissioni Nette:** {st.session_state.em_final:,} tCO2")

    with rt_credito:
        st.subheader("Stress Test Finanziario (Scenari NGFS)", help="Interroga il database Network for Greening the Financial System (NGFS) che traccia il prezzo ombra del carbonio fino al 2050 diviso per paese e scenario.")
        c_cred1, c_cred2 = st.columns(2)
        with c_cred1: st.number_input("Rata Prestito Transizione (€)", value=st.session_state.rata_prestito, step=100000, key='rata_prestito', help="Sottrae un costo annuo fisso dal calcolo dell'EBITDA.")
        with c_cred2: st.slider("Severità Policy", 1.0, 3.0, value=st.session_state.policy_multiplier, step=0.1, key='policy_multiplier', help="Moltiplica il prezzo base della Carbon Tax per simulare scenari normativi regionali estremamente restrittivi.")

        country_data = df_base[df_base['Paese'] == st.session_state.selected_country].copy()
        plot_data = []
        for _, row in country_data.iterrows():
            eff_price = row['Prezzo Carbonio Base'] * st.session_state.policy_multiplier
            profit = st.session_state.revenue - st.session_state.opex - (eff_price * st.session_state.em_final) - st.session_state.rata_prestito
            plot_data.append({"Anno": row['Anno'], "Utile Netto (€)": profit, "Scenario": row['Scenario']})
        
        st.plotly_chart(px.line(pd.DataFrame(plot_data), x="Anno", y="Utile Netto (€)", color="Scenario", title="EBITDA Netto post-Carbon Tax"), use_container_width=True)

# =====================================================================
# TAB 3 E 4: TASSONOMIA E CBAM
# =====================================================================
with t_tax:
    st.header("🇪🇺 Reporting Tassonomia UE", help="Sistema di classificazione europeo che definisce se un'attività economica può dirsi 'eco-sostenibile'. Questo tab mappa le commesse coi codici NACE e le incrocia con il database degli Screening Criteria UE.")
    nace_db = load_nace_hierarchy("NACE_Rev.2.1.rdf")
    tax_pref = load_taxonomy_json("taxonomy.json")

    with st.expander("➕ Aggiungi Commessa / Attività", expanded=True):
        erp_id = st.text_input("🏢 ID Commessa ERP", help="Codice interno utilizzato dalla tua azienda per tracciare il centro di costo/ricavo.")
        col_tax1, col_tax2 = st.columns(2)
        with col_tax1:
            sez = st.selectbox("Sezione NACE", list(nace_db.keys()) if nace_db else [])
            div = st.selectbox("Divisione NACE", list(nace_db.get(sez, {}).keys()) if sez else [])
            grp = st.selectbox("Gruppo NACE", list(nace_db.get(sez, {}).get(div, {}).keys()) if div else [])
            cls = st.selectbox("Classe NACE", list(nace_db.get(sez, {}).get(div, {}).get(grp, {}).keys()) if grp else [], help="L'algoritmo confronta questo codice con il pacchetto .json della Tassonomia. Se trova un match, la tua attività diviene 'Eleggibile'.")
            
            nace_code = nace_db.get(sez, {}).get(div, {}).get(grp, {}).get(cls, "")
            is_eligible = False
            if nace_code and tax_pref:
                is_eligible = any(nace_code.replace('.','').startswith(p) for p in tax_pref)
            
            st.markdown(f"**Status:** {'✅ Ammissibile (Eligible)' if is_eligible else 'ℹ️ Non Ammissibile'}")

        with col_tax2:
            obj_sc = st.selectbox("Obiettivo", ["CCM (Mitigazione)", "CCA (Adattamento)", "WTR (Risorse Idriche)", "CE (Economia Circolare)", "PPC (Inquinamento)", "BIO (Biodiversità)"], help="I 6 obiettivi ambientali fissati dall'Unione Europea.")
            val_t = st.number_input("Turnover della commessa (€)", step=10000, help="Ricavi totali incassati specificamente per questa singola attività.")
            val_c = st.number_input("CapEx della commessa (€)", step=10000)
            val_o = st.number_input("OpEx della commessa (€)", step=10000)

        if st.button("Inserisci in Registro"):
            st.session_state.tax_portfolio.append({
                "ERP": erp_id, "Attività": cls.split(" - ")[-1] if cls else "", "NACE": nace_code,
                "Obiettivo": obj_sc, "Turnover (€)": val_t, "CapEx (€)": val_c, "OpEx (€)": val_o,
                "Eligible (Y/N)": "Y" if is_eligible else "N", "Aligned": False 
            })
            st.rerun()

    if st.session_state.tax_portfolio:
        df_tax = pd.DataFrame(st.session_state.tax_portfolio)
        edited_df = st.data_editor(df_tax, num_rows="dynamic", key="tax_editor", help="Tabella interattiva. Spunta il flag 'Aligned' se l'attività ha superato positivamente lo screening tecnico dei DNSH (Do No Significant Harm).")
        st.session_state.tax_portfolio = edited_df.to_dict('records')
        
        if st.button("Svuota Registro Tassonomia"): st.session_state.tax_portfolio = []; st.rerun()

        st.subheader("Dashboard")
        den_turnover = max(st.session_state.revenue, edited_df["Turnover (€)"].sum())
        val_aligned = edited_df[edited_df["Aligned"] == True]["Turnover (€)"].sum()
        
        c1, c2 = st.columns(2)
        c1.metric("Fatturato Allineato (%)", f"{(val_aligned/den_turnover*100) if den_turnover>0 else 0:.2f}%", help="Valore Green primario richiesto dalle Banche. Pari alla somma dei ricavi di tutte le attività 'Allineate' diviso il Fatturato Totale dell'Azienda inserito nella Sidebar.")
        c2.plotly_chart(go.Figure(data=[go.Pie(labels=["Aligned", "Not Aligned"], values=[val_aligned, den_turnover-val_aligned], hole=.4)]), use_container_width=True)

with t_cbam:
    st.header("🌍 CBAM Self-Assessment Tool", help="Carbon Border Adjustment Mechanism (Dazio Climatico UE). Calcola i sovrapprezzi per le merci pesanti importate da territori extra-UE.")
    cbam_tree = load_cbam_hierarchy()
    paesi = {"Cina": 10.0, "India": 0.0, "UK": 45.0, "USA": 0.0, "UE": 0.0}

    with st.expander("➕ Compila Spedizione Doganale", expanded=True):
        col_cb1, col_cb2 = st.columns(2)
        with col_cb1:
            sez_cb = st.selectbox("Sezione", list(cbam_tree.keys()))
            cap_cb = st.selectbox("Capitolo", list(cbam_tree.get(sez_cb, {}).keys()) if sez_cb else [])
            voc_cb = st.selectbox("Voce", list(cbam_tree.get(sez_cb, {}).get(cap_cb, {}).keys()) if cap_cb else [])
            mer_cb = st.selectbox("Prodotto", list(cbam_tree.get(sez_cb, {}).get(cap_cb, {}).get(voc_cb, {}).keys()) if voc_cb else [], help="L'algoritmo isola le 8 cifre finali (Codice Nomenclatura Combinata). Se il codice inizia per numeri critici (es 72 per l'acciaio), fa scattare l'obbligo di rendicontazione CBAM.")
            cod_cn = cbam_tree.get(sez_cb, {}).get(cap_cb, {}).get(voc_cb, {}).get(mer_cb, "")
            
        with col_cb2:
            orig = st.selectbox("Origine", list(paesi.keys()), help="Paese di estrazione/produzione del bene. Il dazio CBAM si azzera se il paese (es. Regno Unito) possiede un sistema interno di tassazione Carbonio riconosciuto.")
            em_cb = st.number_input("Emissioni (tCO2)", min_value=0.00, step=10.0, help="Tonnellate dirette ed indirette inglobate nel prodotto durante la sua fabbricazione all'estero.")
            val_cb = st.number_input("Valore merce (€)", min_value=0.00, step=100.0, help="Soglia Minima De Minimis: se il valore della spedizione è inferiore a 150 €, l'obbligo decade automaticamente.")

        if st.button("Valuta e Registra"):
            cat = check_cbam_category(cod_cn)
            applies = "SÌ" if (cat != "Non Soggetto" and val_cb > 150 and orig != "UE") else "NO"
            st.session_state.cbam_portfolio.append({
                "Codice": cod_cn, "Merce": mer_cb[:40], "Origine": orig,
                "Valore (€)": round(val_cb, 2), "Emissioni (tCO2)": round(em_cb, 2),
                "CBAM APPLICABILE": applies, "Tax Estera": paesi[orig]
            })
            st.rerun()

    if st.session_state.cbam_portfolio:
        df_cbam = pd.DataFrame(st.session_state.cbam_portfolio)
        st.dataframe(df_cbam, use_container_width=True, column_config={
            "Emissioni (tCO2)": st.column_config.NumberColumn(format="%.2f"),
            "Valore (€)": st.column_config.NumberColumn(format="%.2f")
        })

        if st.button("Svuota CBAM"): st.session_state.cbam_portfolio = []; st.rerun()

        df_app = df_cbam[df_cbam["CBAM APPLICABILE"] == "SÌ"]
        tot_em = df_app["Emissioni (tCO2)"].sum()
        if tot_em > 0:
            st.divider()
            live_price = get_live_eu_ets_price()
            ets = st.number_input("Prezzo EU ETS (€/tCO2)", value=float(live_price), help="Si collega a Yahoo Finance sul contratto KEZ=F. Più è alto il prezzo sul mercato di borsa di Lipsia, più il dazio doganale alla frontiera sarà salato.")
            
            sconto = sum(row["Emissioni (tCO2)"] * row["Tax Estera"] for _, row in df_app.iterrows())
            costo = max(0, (tot_em * ets) - sconto)
            
            c1, c2, c3 = st.columns(3)
            c1.metric("Emissioni CBAM Importate", f"{tot_em:.2f} tCO2")
            c2.metric("Sconto per Tasse Estere", f"€ {sconto:.2f}", help="Credito derivante da eventuali carbon tax già saldate nel paese del fornitore (es. Cina o UK).")
            c3.metric("Costo CBAM Doganale Teorico", f"€ {costo:.2f}", delta="Impatto OpEx in Dogana", delta_color="inverse")

# =====================================================================
# TAB 5: DOWNLOAD
# =====================================================================
with t_down:
    st.header("📥 Esportazione Dati")
    if st.button("🪄 Genera Report Direzionale (PDF)", help="Prende tutti i dati dai vari componenti e compila un report PDF automatizzato prondo per l'Assemblea o l'invio all'Auditor."):
        pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial", 'B', 18); pdf.cell(200, 15, txt="ESG Report", ln=True, align='C')
        st.download_button("Scarica PDF", pdf.output(dest='S').encode('latin-1'), "ESG_Report.pdf")
